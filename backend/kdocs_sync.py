"""Kdocs workbook publishing and in-place synchronization.

The first publish uploads the generated XLSX verbatim so the cloud workbook
keeps the exact layout produced by this application. Later publishes update
the same Kdocs file cell-by-cell, preserving the file id and access link.

Authentication is owned by ``kdocs-cli`` and stored in the operating-system
credential store. This module never reads or persists the token itself.
"""

from __future__ import annotations

import base64
import datetime as dt
import hashlib
import json
import os
import re
import shutil
import subprocess
import sys
import tempfile
import threading
import time
from pathlib import Path
from typing import Any, Callable, Iterable

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Color

from backend.utils.class_utils import normalize_program_name


if sys.platform == "darwin":
    APP_DIR = Path.home() / "Library" / "Application Support" / "DonCollege" / "StudentEvaluation"
elif os.name == "nt":
    APP_DIR = Path(os.environ.get("APPDATA", Path.home())) / "DonCollege" / "StudentEvaluation"
else:
    APP_DIR = Path(os.environ.get("XDG_CONFIG_HOME", Path.home() / ".config")) / "DonCollege" / "StudentEvaluation"
CONFIG_PATH = APP_DIR / "kdocs-workbooks.json"
# Kdocs rejects ``rangeData`` arrays longer than 100 operations. Keep this
# limit centralized so values, formats and merges all obey the same contract.
CHUNK_SIZE = 100
PROGRAM_ORDER = ("顿河交", "顿河土", "顿河信", "国电")
PROGRAM_ORDER_INDEX = {
    normalize_program_name(name): index for index, name in enumerate(PROGRAM_ORDER)
}
REORDER_RETRY_DELAYS = (2, 4)


class KdocsSyncError(RuntimeError):
    def __init__(self, message: str, *, code: int | None = None):
        super().__init__(message)
        self.code = code


_SYNC_LOCK = threading.Lock()
ProgressCallback = Callable[..., None]


def _emit_progress(
    callback: ProgressCallback | None,
    percent: float,
    stage: str,
    detail: str = "",
    **extra: Any,
) -> None:
    if callback is None:
        return
    try:
        callback(max(0, min(100, int(round(percent)))), stage, detail, **extra)
    except Exception:
        # Progress display must never interrupt a cloud write.
        pass


def _subprocess_window_options() -> dict:
    """Keep console-based helpers invisible inside the desktop application."""
    if os.name != "nt":
        return {}
    return {"creationflags": getattr(subprocess, "CREATE_NO_WINDOW", 0x08000000)}


def _find_cli() -> str:
    found = shutil.which("kdocs-cli")
    if found:
        return found
    if getattr(sys, "frozen", False):
        bundled = Path(getattr(sys, "_MEIPASS", "")) / "kdocs-cli.exe"
        if bundled.is_file():
            return str(bundled)
    if os.name == "nt":
        candidate = Path(os.environ.get("LOCALAPPDATA", Path.home())) / "kdocs-cli" / "kdocs-cli.exe"
        if candidate.is_file():
            return str(candidate)
    raise KdocsSyncError("未找到 kdocs-cli，请先安装或重新启动软件后再试。")


def _json_from_output(output: str) -> Any:
    text = (output or "").strip()
    if not text:
        return {}
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        decoder = json.JSONDecoder()
        for index, char in enumerate(text):
            if char not in "[{":
                continue
            try:
                value, _ = decoder.raw_decode(text[index:])
                return value
            except json.JSONDecodeError:
                continue
    raise KdocsSyncError("金山文档返回了无法识别的数据。")


def _run(
    args: list[str],
    payload: dict | None = None,
    timeout: int = 120,
    *,
    parse_json: bool = True,
) -> Any:
    command = [_find_cli(), *args]
    payload_path: str | None = None
    try:
        if payload is not None:
            handle = tempfile.NamedTemporaryFile("w", suffix=".json", encoding="utf-8", delete=False)
            payload_path = handle.name
            with handle:
                json.dump(payload, handle, ensure_ascii=False, separators=(",", ":"))
            command.extend(["--file", payload_path, "--compact"])
        env = os.environ.copy()
        env["PYTHONUTF8"] = "1"
        completed = subprocess.run(
            command,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=timeout,
            env=env,
            check=False,
            **_subprocess_window_options(),
        )
        if completed.returncode != 0:
            message = (completed.stderr or "").strip()
            # Never echo stdout here: authentication commands may print secrets.
            match = re.search(r"\b(\d{5,6})\b", message)
            raise KdocsSyncError(
                message or "金山文档命令执行失败。",
                code=int(match.group(1)) if match else None,
            )
        # OAuth commands intentionally print human-readable progress (and may
        # include sensitive credential material). Their output must be ignored;
        # authentication is confirmed separately with ``auth status``.
        if not parse_json:
            return {}
        return _json_from_output(completed.stdout)
    except subprocess.TimeoutExpired as exc:
        raise KdocsSyncError("连接金山文档超时，请检查网络后重试。") from exc
    finally:
        if payload_path:
            try:
                os.unlink(payload_path)
            except OSError:
                pass


def _api(service: str, action: str, payload: dict, timeout: int = 120) -> Any:
    result = _run([service, action], payload, timeout)
    current = result
    for _ in range(8):
        if not isinstance(current, dict):
            break
        code = current.get("code")
        if code not in (None, 0, "0"):
            try:
                numeric_code = int(code)
            except (TypeError, ValueError):
                numeric_code = None
            message = current.get("msg") or current.get("message") or f"金山文档接口错误：{code}"
            raise KdocsSyncError(str(message), code=numeric_code)
        nested = current.get("data")
        if not isinstance(nested, dict):
            break
        current = nested
    return result


def _is_auth_error(exc: BaseException) -> bool:
    if isinstance(exc, KdocsSyncError) and exc.code == 400006:
        return True
    message = str(exc).lower()
    return any(marker in message for marker in ("400006", "token expired", "authentication expired", "鉴权失败"))


def _data(value: Any) -> Any:
    """Unwrap one or more Kdocs API envelopes.

    kdocs-cli 2.5.13 wraps the service response in its own response object, so
    a successful call can arrive as ``data.data`` even though the service
    documentation shows a single ``data`` level.
    """
    current = value
    for _ in range(8):
        if not isinstance(current, dict) or "data" not in current:
            break
        if len(current) != 1 and not any(key in current for key in ("code", "msg", "message")):
            break
        current = current["data"]
    return current


def auth_status(verify_access: bool = True) -> dict:
    try:
        result = _run(["auth", "status"], timeout=20)
        status = {
            "success": True,
            "installed": True,
            "authenticated": bool(result.get("authenticated")) if isinstance(result, dict) else False,
            "credential_saved": bool(result.get("authenticated")) if isinstance(result, dict) else False,
            "access_verified": False,
        }
        if status["authenticated"] and verify_access:
            bindings = _load_config().get("bindings", {})
            first = next((item for item in bindings.values() if isinstance(item, dict) and item.get("file_id")), None)
            if first:
                try:
                    _api("drive", "get-file-info", {"file_id": str(first["file_id"])}, timeout=45)
                    status["access_verified"] = True
                except KdocsSyncError as exc:
                    if _is_auth_error(exc):
                        status.update({"authenticated": False, "needs_login": True, "error": "金山文档登录已失效，请重新登录。"})
                    else:
                        status["verification_error"] = str(exc)
        return status
    except KdocsSyncError as exc:
        return {"success": False, "installed": False, "authenticated": False, "error": str(exc)}


def login() -> dict:
    """Run the browser OAuth flow without returning any CLI output."""
    try:
        _run(
            ["auth", "login", "--oauth-timeout", "300000"],
            timeout=330,
            parse_json=False,
        )
        status = auth_status(verify_access=False)
        if not status.get("authenticated"):
            return {"success": False, "error": "登录尚未完成，请重试。"}
        return {"success": True, "authenticated": True}
    except KdocsSyncError as exc:
        return {"success": False, "error": str(exc)}


def logout() -> dict:
    """Remove the saved Kdocs credential from the OS keychain."""
    try:
        _run(["auth", "logout"], timeout=20, parse_json=False)
        return {"success": True, "authenticated": False}
    except KdocsSyncError as exc:
        return {"success": False, "error": str(exc)}


def _load_config() -> dict:
    if not CONFIG_PATH.is_file():
        return {"version": 1, "bindings": {}}
    try:
        raw = json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
        if not isinstance(raw, dict):
            raise ValueError
        raw.setdefault("version", 1)
        raw.setdefault("bindings", {})
        return raw
    except (OSError, ValueError, json.JSONDecodeError):
        return {"version": 1, "bindings": {}}


def _save_config(config: dict) -> None:
    APP_DIR.mkdir(parents=True, exist_ok=True)
    temp_path = CONFIG_PATH.with_suffix(".tmp")
    temp_path.write_text(json.dumps(config, ensure_ascii=False, indent=2), encoding="utf-8")
    os.replace(temp_path, CONFIG_PATH)


def _binding(cloud_key: str) -> dict | None:
    item = _load_config().get("bindings", {}).get(cloud_key)
    return item if isinstance(item, dict) and item.get("file_id") else None


def get_binding(cloud_key: str) -> dict:
    item = _binding(cloud_key)
    return {"success": True, "bound": bool(item), **(item or {})}


def get_sync_overview(cloud_key: str, major: str = "") -> dict:
    """Return a user-facing snapshot without exposing credentials."""
    try:
        item = _binding(cloud_key)
        if not item:
            return {"success": True, "bound": False, "sheet_count": 0, "major_sheets": []}
        infos = _sheet_infos(str(item["file_id"]))
        visible_names = [
            str(info.get("sheetName") or "")
            for info in infos
            if info.get("sheetName") and not str(info.get("sheetName")).startswith("_")
        ]
        major = str(major or "").strip()
        major_sheets = [name for name in visible_names if major and name.startswith(major)]
        return {
            "success": True,
            "bound": True,
            "name": item.get("name") or "学院共享表.xlsx",
            "link_url": item.get("link_url") or "",
            "updated_at": item.get("updated_at") or "",
            "last_local_file": item.get("last_local_file") or "",
            "sheet_count": len(visible_names),
            "sheet_names": visible_names,
            "major_sheets": major_sheets,
        }
    except (KdocsSyncError, OSError, ValueError) as exc:
        return {
            "success": False,
            "bound": True,
            "needs_login": _is_auth_error(exc),
            "error": str(exc),
        }


def bind_workbook(cloud_key: str, link_url: str) -> dict:
    """Bind this installation to an existing college workbook shared by a peer."""
    try:
        if not cloud_key or len(cloud_key) > 80:
            raise KdocsSyncError("云表标识无效。")
        if not isinstance(link_url, str) or not link_url.strip().lower().startswith(("http://", "https://")):
            raise KdocsSyncError("请输入有效的金山文档链接。")
        if not auth_status().get("authenticated"):
            return {"success": False, "needs_login": True, "error": "请先连接金山文档账号。"}
        resolved = _api("drive", "read-file", {"url": link_url.strip()}, timeout=180)
        file_id = _extract_file_id(resolved)
        if not file_id:
            raise KdocsSyncError("无法从该链接识别云表格，请确认当前账号具有访问权限。")
        info = _data(_api("drive", "get-file-info", {"file_id": file_id}))
        if not isinstance(info, dict):
            raise KdocsSyncError("云表格验证失败。")
        name = str(info.get("name") or "学院共享表.xlsx")
        canonical_link = _get_link(file_id) or link_url.strip()
        config = _load_config()
        config["bindings"][cloud_key] = {
            "file_id": file_id,
            "link_url": canonical_link,
            "name": name,
            "updated_at": dt.datetime.now().isoformat(timespec="seconds"),
        }
        _save_config(config)
        return {
            "success": True,
            "bound": True,
            "file_id": file_id,
            "link_url": canonical_link,
            "name": name,
        }
    except (KdocsSyncError, OSError, ValueError) as exc:
        return {"success": False, "error": str(exc)}


def _extract_file_id(result: Any) -> str:
    body = _data(result)
    if not isinstance(body, dict):
        return ""
    return str(body.get("file_id") or body.get("id") or "")


def _get_link(file_id: str) -> str:
    result = _api("drive", "get-file-link", {"file_id": file_id})
    body = _data(result)
    if not isinstance(body, dict):
        return ""
    return str(body.get("file_url") or body.get("link_url") or body.get("url") or "")


def _publish_new(path: Path, cloud_key: str, progress_callback: ProgressCallback | None = None) -> dict:
    _emit_progress(progress_callback, 10, "正在准备文件", path.name)
    raw = path.read_bytes()
    _emit_progress(progress_callback, 22, "正在上传 Excel", "首次创建学院云表")
    result = _api(
        "drive",
        "upload-file",
        {
            "name": path.name,
            "content_base64": base64.b64encode(raw).decode("ascii"),
            "file_sum": hashlib.sha256(raw).hexdigest(),
            "file_type": "sha256",
        },
        timeout=180,
    )
    _emit_progress(progress_callback, 72, "上传完成", "正在取得云文件信息")
    file_id = _extract_file_id(result)
    if not file_id:
        raise KdocsSyncError("文件已上传，但没有取得云文件 ID。")
    # Independent verification: confirm the newly created file exists.
    _emit_progress(progress_callback, 80, "正在校验云端文件", "确认文件已经创建")
    info = _api("drive", "get-file-info", {"file_id": file_id})
    if not _data(info):
        raise KdocsSyncError("文件上传后的云端校验失败。")
    ordering = _reorder_workbook(
        file_id,
        progress_callback,
        progress_start=82,
        progress_end=92,
    )
    _emit_progress(progress_callback, 94, "正在取得共享链接", "准备完成同步")
    link = _get_link(file_id)
    if not link:
        raise KdocsSyncError("文件已上传，但没有取得访问链接。")

    config = _load_config()
    config["bindings"][cloud_key] = {
        "file_id": file_id,
        "link_url": link,
        "name": path.name,
        "updated_at": dt.datetime.now().isoformat(timespec="seconds"),
    }
    _save_config(config)
    _emit_progress(progress_callback, 97, "正在保存同步记录", path.name)
    return {
        "success": True,
        "created": True,
        "file_id": file_id,
        "link_url": link,
        "name": path.name,
        "moved_sheets": ordering["moved_sheets"],
        "sheet_order": ordering["sheet_order"],
        "message": "已原样上传，云端模板与软件生成的 Excel 一致。",
    }


def _sheet_infos(file_id: str) -> list[dict]:
    result = _api("sheet", "get-sheets-info", {"file_id": file_id})
    body = _data(result)
    if isinstance(body, dict) and isinstance(body.get("detail"), dict):
        body = body["detail"]
    if isinstance(body, dict):
        infos = body.get("sheetsInfo") or body.get("sheets_info") or []
    else:
        infos = []
    return [item for item in infos if isinstance(item, dict)]


def _sheet_identity(info: dict) -> str:
    value = info.get("sheetId", info.get("sheet_id"))
    return str(value) if value is not None else ""


def _sheet_index(info: dict) -> int:
    value = info.get("sheetIdx", info.get("sheet_idx", 10**9))
    try:
        return int(value)
    except (TypeError, ValueError):
        return 10**9


def _sheet_sort_key(info: dict) -> tuple:
    """Return the college-wide canonical order for one worksheet.

    Recognised class sheets are grouped by programme and then ordered by grade
    and class number. Programme-grade ranking sheets (for example ``顿河交24``)
    use class number zero, so they remain grouped with the same programme.
    Helper sheets are kept at the end.
    """
    name = str(info.get("sheetName", info.get("sheet_name", "")) or "").strip()
    if name.startswith("_") or info.get("isVisible", info.get("is_visible", True)) is False:
        return (3, 0, "", 0, 0, name.casefold())

    cleaned = re.sub(r"班$", "", name).strip()
    match = re.fullmatch(r"(?P<program>.*?)(?P<grade>\d{2})(?P<class_num>\d{1,2})", cleaned)
    if match:
        program = match.group("program").strip()
        grade = int(match.group("grade"))
        class_num = int(match.group("class_num"))
    else:
        match = re.fullmatch(r"(?P<program>.*?)(?P<grade>\d{2})", cleaned)
        if not match:
            return (2, 0, "", 0, 0, name.casefold())
        program = match.group("program").strip()
        grade = int(match.group("grade"))
        class_num = 0

    normalized = normalize_program_name(program)
    known_index = PROGRAM_ORDER_INDEX.get(normalized)
    return (
        0 if known_index is not None else 1,
        known_index if known_index is not None else len(PROGRAM_ORDER),
        "" if known_index is not None else normalized,
        grade,
        class_num,
        name.casefold(),
    )


def _ordered_sheet_infos(infos: Iterable[dict]) -> list[dict]:
    return sorted(
        (item for item in infos if _sheet_identity(item)),
        key=lambda item: (_sheet_sort_key(item), _sheet_index(item)),
    )


def _worksheet_move_plan(current_ids: list[str], desired_ids: list[str]) -> list[tuple[str, str]]:
    """Plan the minimum sequence of ``move after`` operations."""
    order = list(current_ids)
    moves: list[tuple[str, str]] = []
    for position in range(1, len(desired_ids)):
        worksheet_id = desired_ids[position]
        previous_id = desired_ids[position - 1]
        if worksheet_id not in order or previous_id not in order:
            continue
        if order.index(worksheet_id) == order.index(previous_id) + 1:
            continue
        order.remove(worksheet_id)
        insert_at = order.index(previous_id) + 1
        order.insert(insert_at, worksheet_id)
        moves.append((worksheet_id, previous_id))
    return moves


def _reorder_workbook(
    file_id: str,
    progress_callback: ProgressCallback | None = None,
    *,
    progress_start: float = 4,
    progress_end: float = 98,
    retry_delays: tuple[int, ...] = REORDER_RETRY_DELAYS,
) -> dict:
    """Reorder all worksheets and verify the final cloud-side order.

    Multiple computers may run this simultaneously. Because every client uses
    the same canonical key, a fresh read plus bounded retry converges after the
    final concurrent sheet creation or move completes.
    """
    attempts = len(retry_delays) + 1
    moved_total = 0
    last_order: list[str] = []
    for attempt in range(attempts):
        span = max(progress_end - progress_start, 1)
        attempt_base = progress_start + span * (attempt / attempts)
        _emit_progress(
            progress_callback,
            attempt_base,
            "正在整理工作表顺序",
            "读取云端工作表" if attempt == 0 else f"检测到并发更新，正在第 {attempt + 1} 次校准",
        )
        infos = sorted(_sheet_infos(file_id), key=_sheet_index)
        desired_infos = _ordered_sheet_infos(infos)
        current_ids = [_sheet_identity(item) for item in infos if _sheet_identity(item)]
        desired_ids = [_sheet_identity(item) for item in desired_infos]
        plan = _worksheet_move_plan(current_ids, desired_ids)

        try:
            for move_index, (worksheet_id, previous_id) in enumerate(plan, start=1):
                percent = attempt_base + (span / attempts) * (move_index / max(len(plan), 1)) * 0.72
                name = next(
                    (str(item.get("sheetName") or "") for item in desired_infos if _sheet_identity(item) == worksheet_id),
                    "",
                )
                _emit_progress(
                    progress_callback,
                    percent,
                    "正在整理工作表顺序",
                    name or f"第 {move_index} 张工作表",
                    current_sheet=name,
                    sheet_index=move_index,
                    sheet_total=len(plan),
                )
                _api(
                    "sheet",
                    "update-worksheet",
                    {
                        "file_id": file_id,
                        "worksheet_id": int(worksheet_id),
                        "move_sheet_id": int(previous_id),
                        "move_type": "sheet_move_type_after",
                    },
                )
                moved_total += 1
        except KdocsSyncError as exc:
            message = str(exc).lower()
            if attempt >= attempts - 1 or not any(marker in message for marker in ("conflict", "lock", "冲突", "占用")):
                raise

        _emit_progress(progress_callback, min(progress_end - 1, attempt_base + span / attempts * 0.82), "正在校验工作表顺序", "回读云端排列")
        verified_infos = sorted(_sheet_infos(file_id), key=_sheet_index)
        verified_ids = [_sheet_identity(item) for item in verified_infos if _sheet_identity(item)]
        verified_desired = [_sheet_identity(item) for item in _ordered_sheet_infos(verified_infos)]
        last_order = [str(item.get("sheetName") or "") for item in verified_infos]
        if verified_ids == verified_desired:
            _emit_progress(progress_callback, progress_end, "工作表顺序已整理", f"共 {len(verified_ids)} 张工作表")
            return {
                "success": True,
                "moved_sheets": moved_total,
                "sheet_count": len(verified_ids),
                "sheet_order": last_order,
                "attempts": attempt + 1,
            }
        if attempt < len(retry_delays):
            time.sleep(retry_delays[attempt])

    raise KdocsSyncError("工作表顺序在并发同步后仍未稳定，请稍后点击“重新整理工作表”。")


def _cloud_cells(result: Any) -> dict[tuple[int, int], str]:
    body = _data(result)
    if isinstance(body, dict) and isinstance(body.get("detail"), dict):
        body = body["detail"]
    if not isinstance(body, dict):
        return {}
    cells = body.get("rangeData") or body.get("range_data") or []
    output: dict[tuple[int, int], str] = {}
    for cell in cells:
        if not isinstance(cell, dict):
            continue
        row = cell.get("rowFrom", cell.get("row_from"))
        col = cell.get("colFrom", cell.get("col_from"))
        if row is None or col is None:
            continue
        value = cell.get("fmlaText", cell.get("fmla_text"))
        if value in (None, ""):
            value = cell.get("originalCellValue", cell.get("original_cell_value", cell.get("cellText", "")))
        output[(int(row), int(col))] = _normalise(value)
    return output


def _normalise(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (dt.datetime, dt.date, dt.time)):
        return value.isoformat(sep=" ") if isinstance(value, dt.datetime) else value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _local_value(value: Any) -> str:
    return _normalise(value)


def _colour(value: Color | None) -> dict | None:
    if value is None:
        return None
    rgb = getattr(value, "rgb", None)
    if isinstance(rgb, str) and len(rgb) in (6, 8):
        if len(rgb) == 6:
            rgb = "FF" + rgb
        try:
            return {"type": 2, "value": int(rgb, 16), "tint": 0}
        except ValueError:
            return None
    return None


_BORDER_STYLE = {
    None: 0,
    "thin": 1,
    "medium": 2,
    "dashed": 3,
    "dotted": 4,
    "thick": 5,
    "double": 6,
    "dashDot": 7,
    "dashDotDot": 7,
    "mediumDashed": 3,
    "mediumDashDot": 7,
    "mediumDashDotDot": 7,
    "hair": 4,
}


def _cell_format(cell) -> dict:
    xf: dict[str, Any] = {}
    alignment = cell.alignment
    horizontal = {"left": 1, "center": 2, "right": 3, "fill": 4, "justify": 5, "centerContinuous": 6, "distributed": 7}
    vertical = {"top": 0, "center": 1, "bottom": 2, "justify": 3, "distributed": 4}
    if alignment.horizontal in horizontal:
        xf["alcH"] = horizontal[alignment.horizontal]
    if alignment.vertical in vertical:
        xf["alcV"] = vertical[alignment.vertical]
    if alignment.wrap_text is not None:
        xf["wrap"] = bool(alignment.wrap_text)
    if alignment.shrink_to_fit is not None:
        xf["shrinkToFit"] = bool(alignment.shrink_to_fit)
    if alignment.indent:
        xf["indent"] = int(alignment.indent)
    if alignment.text_rotation:
        xf["trot"] = int(alignment.text_rotation)

    font: dict[str, Any] = {}
    if cell.font.name:
        font["name"] = cell.font.name
    if cell.font.sz:
        font["dyHeight"] = int(round(float(cell.font.sz) * 20))
    if cell.font.bold:
        font["bls"] = True
    if cell.font.italic:
        font["italic"] = True
    if cell.font.strike:
        font["strikeout"] = True
    colour = _colour(cell.font.color)
    if colour:
        font["color"] = colour
    if font:
        xf["font"] = font

    if cell.number_format:
        xf["numfmt"] = cell.number_format

    fill_colour = _colour(cell.fill.fgColor)
    if cell.fill.fill_type == "solid" and fill_colour:
        xf["fill"] = {
            "type": 1,
            "back": fill_colour,
            "fore": {"type": 255, "value": 0, "tint": 0},
        }

    for side_name, api_name, colour_name in (
        ("left", "dgLeft", "clrLeft"),
        ("right", "dgRight", "clrRight"),
        ("top", "dgTop", "clrTop"),
        ("bottom", "dgBottom", "clrBottom"),
    ):
        side = getattr(cell.border, side_name)
        style = _BORDER_STYLE.get(side.style, 0)
        if style:
            xf[api_name] = style
            side_colour = _colour(side.color)
            if side_colour:
                xf[colour_name] = side_colour
    return xf


def _chunks(items: list[dict], size: int = CHUNK_SIZE) -> Iterable[list[dict]]:
    for index in range(0, len(items), size):
        yield items[index:index + size]


def _update_ops(file_id: str, worksheet_id: int, operations: list[dict]) -> None:
    for chunk in _chunks(operations):
        _api(
            "sheet",
            "update-range-data",
            {"file_id": file_id, "worksheet_id": worksheet_id, "rangeData": chunk},
            timeout=120,
        )


def _format_ops_for_sheet(ws, max_row: int, max_col: int) -> list[dict]:
    """Coalesce equal adjacent cell formats into row ranges.

    Kdocs limits each request to 100 rangeData entries. Generated worksheets
    commonly repeat one style across a whole row, so sending ranges instead of
    one operation per cell is both faster and less failure-prone.
    """
    operations: list[dict] = []
    for row_index in range(1, max_row + 1):
        run_start: int | None = None
        run_end: int | None = None
        run_xf: dict | None = None
        for col_index in range(1, max_col + 2):
            xf = None
            if col_index <= max_col:
                cell = ws.cell(row=row_index, column=col_index)
                if not isinstance(cell, MergedCell) and cell.style_id:
                    xf = _cell_format(cell) or None
            if xf is not None and xf == run_xf:
                run_end = col_index
                continue
            if run_xf is not None and run_start is not None and run_end is not None:
                operations.append({
                    "opType": "format",
                    "rowFrom": row_index - 1, "rowTo": row_index - 1,
                    "colFrom": run_start - 1, "colTo": run_end - 1,
                    "xf": run_xf,
                })
            run_xf = xf
            run_start = col_index if xf is not None else None
            run_end = col_index if xf is not None else None
    return operations


def _read_region(file_id: str, worksheet_id: int, max_row: int, max_col: int) -> dict[tuple[int, int], str]:
    if max_row <= 0 or max_col <= 0:
        return {}
    result = _api(
        "sheet",
        "get-range-data",
        {
            "file_id": file_id,
            "worksheet_id": worksheet_id,
            "range": {"rowFrom": 0, "rowTo": max_row - 1, "colFrom": 0, "colTo": max_col - 1},
        },
    )
    return _cloud_cells(result)


def _add_sheet(file_id: str, name: str) -> dict:
    _api(
        "sheet",
        "add-sheet",
        {"file_id": file_id, "name": name, "end": True, "count": 1},
    )
    refreshed = _sheet_infos(file_id)
    match = next((item for item in refreshed if item.get("sheetName") == name), None)
    if not match:
        raise KdocsSyncError(f"已请求新增工作表“{name}”，但云端校验没有找到它。")
    return match


def _sync_sheet(
    file_id: str,
    ws,
    info: dict,
    created: bool,
    progress_callback: ProgressCallback | None = None,
) -> dict:
    _emit_progress(progress_callback, 4, "正在读取班级表", ws.title)
    worksheet_id = int(info.get("sheetId"))
    max_row = max(int(ws.max_row or 0), int(info.get("rowTo", -1)) + 1)
    max_col = max(int(ws.max_column or 0), int(info.get("colTo", -1)) + 1)
    before = {} if created else _read_region(file_id, worksheet_id, max_row, max_col)

    _emit_progress(progress_callback, 20, "正在分析单元格变化", ws.title)
    local: dict[tuple[int, int], str] = {}
    value_ops: list[dict] = []
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            key = (cell.row - 1, cell.column - 1)
            value = _local_value(cell.value)
            local[key] = value
            if before.get(key, "") != value:
                value_ops.append({
                    "opType": "formula",
                    "rowFrom": key[0], "rowTo": key[0],
                    "colFrom": key[1], "colTo": key[1],
                    "formula": value,
                })
    # Clear cells that existed in the cloud but are blank in the regenerated workbook.
    for key, old_value in before.items():
        if old_value and key in local and not local[key]:
            value_ops.append({
                "opType": "formula",
                "rowFrom": key[0], "rowTo": key[0],
                "colFrom": key[1], "colTo": key[1],
                "formula": "",
            })

    _emit_progress(progress_callback, 38, "正在写入数据和公式", f"{ws.title} · {len(value_ops)} 个变更")
    _update_ops(file_id, worksheet_id, value_ops)
    # Formatting is idempotent and is deliberately refreshed on every sync.
    # This repairs a sheet if an earlier attempt stopped after writing values.
    _emit_progress(progress_callback, 58, "正在更新表格格式", ws.title)
    _update_ops(file_id, worksheet_id, _format_ops_for_sheet(ws, max_row, max_col))
    _emit_progress(progress_callback, 76, "正在整理工作表", ws.title)
    if created:
        merge_ops = []
        for merged in ws.merged_cells.ranges:
            merge_ops.append({
                "opType": "merge",
                "rowFrom": merged.min_row - 1, "rowTo": merged.max_row - 1,
                "colFrom": merged.min_col - 1, "colTo": merged.max_col - 1,
                "type": "MergeCenter",
            })
        _update_ops(file_id, worksheet_id, merge_ops)
        # Kdocs does not expose exact arbitrary width/height assignment here;
        # auto-fit keeps appended sheets readable while cell formatting remains faithful.
        if max_row and max_col:
            from openpyxl.utils import get_column_letter
            _api(
                "sheet",
                "auto-fit",
                {
                    "file_id": file_id,
                    "worksheet_id": worksheet_id,
                    "range": f"A1:{get_column_letter(max_col)}{max_row}",
                    "fit_type": "columns",
                },
            )

    if value_ops:
        # Independent post-write read and comparison.
        _emit_progress(progress_callback, 86, "正在回读校验", ws.title)
        after = _read_region(file_id, worksheet_id, max_row, max_col)
        mismatches = []
        for operation in value_ops:
            key = (operation["rowFrom"], operation["colFrom"])
            expected = _normalise(operation["formula"])
            if after.get(key, "") != expected:
                mismatches.append(key)
                if len(mismatches) >= 5:
                    break
        if mismatches:
            raise KdocsSyncError(f"工作表“{ws.title}”写入后校验不一致，请重试。")
    _emit_progress(progress_callback, 100, "班级表已完成", ws.title)
    return {"name": ws.title, "changed_cells": len(value_ops), "created": created}


def _sync_existing(
    path: Path,
    cloud_key: str,
    binding: dict,
    progress_callback: ProgressCallback | None = None,
) -> dict:
    file_id = str(binding["file_id"])
    _emit_progress(progress_callback, 8, "正在读取云表", "获取现有班级列表")
    infos = _sheet_infos(file_id)
    by_name = {str(item.get("sheetName")): item for item in infos}
    workbook = load_workbook(path, data_only=False, read_only=False)
    summaries = []
    try:
        visible_sheets = [ws for ws in workbook.worksheets if ws.sheet_state == "visible"]
        sheet_total = len(visible_sheets)
        _emit_progress(progress_callback, 14, "准备同步班级", f"共 {sheet_total} 个工作表", sheet_total=sheet_total)
        for sheet_index, ws in enumerate(visible_sheets, start=1):
            # Generated workbooks contain a hidden _values sheet for local
            # downstream calculations. It is not referenced by visible cloud
            # sheets and would overwrite another major's helper data.
            info = by_name.get(ws.title)
            # A previous attempt may have created the worksheet and then failed
            # before writing its contents. Resume an empty remote sheet as a new
            # one so values, formatting and merged cells are all restored.
            created = info is None or bool(info.get("isEmpty"))
            if created:
                if info is None:
                    _emit_progress(
                        progress_callback,
                        14 + ((sheet_index - 1) / max(sheet_total, 1)) * 76,
                        "正在新增班级表",
                        ws.title,
                        current_sheet=ws.title,
                        sheet_index=sheet_index,
                        sheet_total=sheet_total,
                    )
                    info = _add_sheet(file_id, ws.title)
                    by_name[ws.title] = info

            def sheet_progress(percent, stage, detail="", **_extra):
                overall = 14 + ((sheet_index - 1 + percent / 100) / max(sheet_total, 1)) * 76
                _emit_progress(
                    progress_callback,
                    overall,
                    stage,
                    detail,
                    current_sheet=ws.title,
                    sheet_index=sheet_index,
                    sheet_total=sheet_total,
                )

            summaries.append(_sync_sheet(file_id, ws, info, created, sheet_progress))
    finally:
        workbook.close()

    ordering = _reorder_workbook(
        file_id,
        progress_callback,
        progress_start=91,
        progress_end=97,
    )
    _emit_progress(progress_callback, 98, "正在整理同步结果", "保存共享链接和更新时间")
    link = binding.get("link_url") or _get_link(file_id)
    config = _load_config()
    config["bindings"][cloud_key].update({
        "link_url": link,
        "updated_at": dt.datetime.now().isoformat(timespec="seconds"),
        "last_local_file": path.name,
    })
    _save_config(config)
    _emit_progress(progress_callback, 97, "正在完成云端校验", f"已处理 {len(summaries)} 个工作表")
    return {
        "success": True,
        "created": False,
        "file_id": file_id,
        "link_url": link,
        "name": binding.get("name") or path.name,
        "changed_cells": sum(item["changed_cells"] for item in summaries),
        "created_sheets": [item["name"] for item in summaries if item["created"]],
        "sheets": summaries,
        "moved_sheets": ordering["moved_sheets"],
        "sheet_order": ordering["sheet_order"],
        "message": "已更新同一份金山文档，共享链接保持不变。",
    }


def reorder_bound_workbook(
    cloud_key: str,
    progress_callback: ProgressCallback | None = None,
) -> dict:
    """Manually restore the canonical worksheet order for one bound workbook."""
    if not _SYNC_LOCK.acquire(blocking=False):
        return {"success": False, "busy": True, "error": "另一个同步或整理任务正在进行，请等待完成后再试。"}
    try:
        _emit_progress(progress_callback, 1, "正在启动顺序整理", "检查金山文档账号")
        if not cloud_key or len(cloud_key) > 80:
            raise KdocsSyncError("云表标识无效。")
        status = auth_status()
        if not status.get("authenticated"):
            return {"success": False, "needs_login": True, "error": "请先连接金山文档账号。"}
        binding = _binding(cloud_key)
        if not binding or not binding.get("file_id"):
            raise KdocsSyncError("这类结果还没有绑定学院云表。")
        ordering = _reorder_workbook(
            str(binding["file_id"]),
            progress_callback,
            progress_start=5,
            progress_end=98,
        )
        return {
            "success": True,
            "file_id": str(binding["file_id"]),
            "link_url": binding.get("link_url") or _get_link(str(binding["file_id"])),
            "name": binding.get("name") or "学院共享表.xlsx",
            **ordering,
        }
    except (KdocsSyncError, OSError, ValueError) as exc:
        _emit_progress(progress_callback, 100, "整理失败", str(exc))
        return {"success": False, "needs_login": _is_auth_error(exc), "error": str(exc)}
    finally:
        _SYNC_LOCK.release()


def sync_workbook(
    local_path: str,
    cloud_key: str,
    progress_callback: ProgressCallback | None = None,
    force_create: bool = False,
) -> dict:
    if not _SYNC_LOCK.acquire(blocking=False):
        return {"success": False, "busy": True, "error": "另一个同步任务正在进行，请等待完成后再试。"}
    try:
        _emit_progress(progress_callback, 1, "正在启动同步", "检查本地文件")
        path = Path(local_path).expanduser().resolve()
        if not path.is_file():
            raise KdocsSyncError("找不到软件生成的 Excel 文件。")
        if path.suffix.lower() not in {".xlsx", ".xls"}:
            raise KdocsSyncError("当前试用版仅支持同步 Excel 表格。")
        if not cloud_key or len(cloud_key) > 80:
            raise KdocsSyncError("云表标识无效。")
        _emit_progress(progress_callback, 3, "正在检查账号", "验证金山文档访问权限")
        status = auth_status()
        if not status.get("authenticated"):
            return {"success": False, "needs_login": True, "error": "请先连接金山文档账号。"}
        _emit_progress(progress_callback, 6, "正在确认云表", path.name)
        binding = None if force_create else _binding(cloud_key)
        if not binding:
            result = _publish_new(path, cloud_key, progress_callback)
        else:
            result = _sync_existing(path, cloud_key, binding, progress_callback)
        if result.get("success"):
            _emit_progress(progress_callback, 100, "同步完成", result.get("name") or path.name)
        return result
    except (KdocsSyncError, OSError, ValueError) as exc:
        _emit_progress(progress_callback, 100, "同步失败", str(exc))
        return {"success": False, "needs_login": _is_auth_error(exc), "error": str(exc)}
    finally:
        _SYNC_LOCK.release()
