"""
Excel workbook writer using openpyxl for formatted multi-sheet output.

Supports both regular values and Excel formulas (strings starting with '=').
Formula cells are written natively and will calculate when opened in Excel.
"""

import os
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def unique_path(filepath: str) -> str:
    """Return a unique file path. If filepath exists, append (1), (2), etc.

    Example:
        '学分绩点.xlsx' → '学分绩点(1).xlsx' (if exists)
        '学分绩点(1).xlsx' → '学分绩点(2).xlsx' (if also exists)
    """
    if not os.path.exists(filepath):
        return filepath
    base, ext = os.path.splitext(filepath)
    # If the filename already ends with (N), start from that number
    counter = 1
    while os.path.exists(f"{base}({counter}){ext}"):
        counter += 1
    return f"{base}({counter}){ext}"



# Style constants — SimSun (宋体) 10pt, plain white, black text (v2.3)
HEADER_FONT = Font(name='SimSun', size=10, bold=True)
DATA_FONT = Font(name='SimSun', size=10)
FORMULA_FONT = Font(name='SimSun', size=10)
THIN_BORDER = Border(
    left=Side(style='thin', color='B0B0B0'),
    right=Side(style='thin', color='B0B0B0'),
    top=Side(style='thin', color='B0B0B0'),
    bottom=Side(style='thin', color='B0B0B0'),
)
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
LEFT_ALIGN = Alignment(horizontal='center', vertical='center')  # All centered


def write_multi_sheet_xlsx(
    output_path: str,
    sheets_data: dict,
    title: str = '',
    add_credit_row: bool = False,
    credit_data: dict = None,
    auto_filter: bool = True,
    freeze_row: int = 1,
):
    """Write a multi-sheet .xlsx file with consistent formatting.

    Args:
        output_path: Full path for the output .xlsx file
        sheets_data: Dict of {sheet_name: [list of dicts]} — each dict is one row
        title: Optional title text (placed in row 1 above headers)
        add_credit_row: Whether to add a credit/学分 row (for Module A)
        credit_data: Dict of {sheet_name: [credit values]} for credit row
        auto_filter: Enable auto-filter on header row
        freeze_row: Row number to freeze panes below (0-indexed)
    """
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    for sheet_name, rows in sheets_data.items():
        if not rows:
            continue

        ws = wb.create_sheet(title=str(sheet_name)[:31])  # Excel 31-char limit

        # Get column headers from first row
        headers = list(rows[0].keys()) if rows else []
        num_cols = len(headers)

        row_offset = 0

        # Write title row if provided
        if title:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
            cell = ws.cell(row=1, column=1, value=title)
            cell.font = Font(name='SimSun', size=14, bold=True, color='333333')
            cell.alignment = CENTER_ALIGN
            row_offset = 1

        # Write header row
        header_row = row_offset + 1
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER

        # Write credit row if needed
        if add_credit_row and credit_data and sheet_name in credit_data:
            credit_row = header_row + 1
            credits = credit_data[sheet_name]
            for col_idx in range(1, num_cols + 1):
                cell = ws.cell(row=credit_row, column=col_idx)
                if col_idx <= len(credits):
                    cell.value = credits[col_idx - 1]
                cell.font = Font(name='SimSun', size=10, italic=True)
                cell.alignment = CENTER_ALIGN
                cell.border = THIN_BORDER
            data_start_row = credit_row + 1
        else:
            data_start_row = header_row + 1

        # Write data rows
        for row_idx, row_data in enumerate(rows):
            excel_row = data_start_row + row_idx

            for col_idx, header in enumerate(headers, 1):
                value = row_data.get(header)
                cell = ws.cell(row=excel_row, column=col_idx, value=value)
                cell.border = THIN_BORDER

                # Student ID columns: force text format (MUST be before numeric check)
                is_id_col = header in ('学号', 'student_id', 'id', '行政班级',
                                       '班级', '学生行政班级', '专业', '专业组')

                # Check if value is an Excel formula
                is_formula = isinstance(value, str) and value.startswith('=')
                if is_formula:
                    cell.font = FORMULA_FONT
                    cell.alignment = CENTER_ALIGN
                    cell.number_format = '0.00'
                elif is_id_col and value is not None:
                    # ID/text columns: always text format, no decimals
                    cell.font = DATA_FONT
                    cell.alignment = CENTER_ALIGN
                    cell.number_format = '@'
                    # Ensure value is string (strip .0 from float student IDs)
                    if isinstance(value, float) and value == int(value):
                        value = str(int(value))
                    cell.value = str(value)
                elif isinstance(value, (int, float)):
                    cell.font = DATA_FONT
                    cell.alignment = CENTER_ALIGN
                    cell.number_format = '0.00'
                else:
                    cell.font = DATA_FONT
                    cell.alignment = LEFT_ALIGN

        # Auto-fit column widths (approximate)
        for col_idx in range(1, num_cols + 1):
            max_width = 0
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx,
                                    min_row=1, max_row=ws.max_row):
                for cell in row:
                    if cell.value:
                        # Chinese characters are ~2x width
                        val = str(cell.value)
                        width = sum(2 if ord(c) > 127 else 1 for c in val)
                        max_width = max(max_width, width)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_width + 4, 40)

        # Freeze panes
        freeze_cell = f'A{data_start_row}'
        ws.freeze_panes = freeze_cell

        # Auto-filter
        if auto_filter:
            ws.auto_filter.ref = f'A{header_row}:{get_column_letter(num_cols)}{header_row}'

    # Save
    wb.save(output_path)
    wb.close()


def write_ranking_xlsx(
    output_path: str,
    ranking_data: dict,
    columns: list = None,
):
    """Write a ranking .xlsx file with per-program-grade sheets.

    Args:
        output_path: Full path for output file
        ranking_data: Dict of {program_grade_key: [ranked student dicts]}
        columns: Column names for the output
    """
    if columns is None:
        columns = ['学号', '姓名', '学分绩点', '排名', '百分比']

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for group_name, students in ranking_data.items():
        if not students:
            continue

        ws = wb.create_sheet(title=str(group_name)[:31])
        num_cols = len(columns)

        # Write header
        for col_idx, header in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER

        # Write data
        for row_idx, student in enumerate(students):
            for col_idx, col_name in enumerate(columns, 1):
                value = student.get(col_name, '')
                # Force text format for ID columns — strip float .0
                if col_name == '学号' and value:
                    if isinstance(value, float) and value == int(value):
                        value = str(int(value))
                    else:
                        value = str(value)
                    cell = ws.cell(row=row_idx + 2, column=col_idx, value=value)
                    cell.number_format = '@'
                else:
                    cell = ws.cell(row=row_idx + 2, column=col_idx, value=value)
                cell.font = DATA_FONT
                cell.border = THIN_BORDER
                if isinstance(value, float):
                    if col_name == '百分比':
                        cell.number_format = '0.00%'
                    else:
                        cell.number_format = '0.00'
                cell.alignment = CENTER_ALIGN if col_idx > 2 else LEFT_ALIGN

        # Column widths
        col_widths = [14, 12, 14, 12, 12]
        for col_idx, width in enumerate(col_widths[:num_cols], 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Freeze header
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f'A1:{get_column_letter(num_cols)}1'

    wb.save(output_path)
    wb.close()


# ============================================================
# Formula helper utilities
# ============================================================

def col_letter(idx: int) -> str:
    """Convert 0-based column index to Excel column letter(s).
    0 -> A, 1 -> B, ..., 25 -> Z, 26 -> AA, etc.
    """
    return get_column_letter(idx + 1)


def cell_ref(row: int, col_idx: int, fixed_col: bool = False,
             fixed_row: bool = False) -> str:
    """Build an Excel cell reference like 'D5' or '$D$5'.

    Args:
        row: 1-based row number
        col_idx: 0-based column index
        fixed_col: If True, use absolute column reference ($D)
        fixed_row: If True, use absolute row reference ($5)

    Returns:
        Cell reference string, e.g. 'D5', '$D5', 'D$5', '$D$5'
    """
    col = col_letter(col_idx)
    if fixed_col:
        col = f'${col}'
    row_str = f'${row}' if fixed_row else str(row)
    return f'{col}{row_str}'


def sum_formula(row_start: int, row_end: int, col_idx: int) -> str:
    """Build a SUM formula for a column range.
    Example: sum_formula(5, 10, 3) -> '=SUM(D5:D10)'
    """
    col = col_letter(col_idx)
    return f'=SUM({col}{row_start}:{col}{row_end})'


def sum_range_formula(row: int, col_start: int, col_end: int) -> str:
    """Build a SUM formula for a row range.
    Example: sum_range_formula(5, 3, 10) -> '=SUM(D5:K5)'
    """
    c1 = col_letter(col_start)
    c2 = col_letter(col_end)
    return f'=SUM({c1}{row}:{c2}{row})'


def weighted_formula(components: list, row: int) -> str:
    """Build a weighted sum formula.
    Args:
        components: List of (col_idx, weight) tuples
        row: Row number
    Returns:
        Formula string like '=D5*0.6+E5*0.3+F5*0.1'
    """
    parts = []
    for col_idx, weight in components:
        ref = cell_ref(row, col_idx)
        parts.append(f'{ref}*{weight}')
    return '=' + '+'.join(parts)


def if_formula(condition_ref: str, true_val, false_val) -> str:
    """Build an IF formula.
    Example: if_formula('D5<>""', cell_ref(5,3), 0) -> '=IF(D5<>"",D5,0)'
    """
    return f'=IF({condition_ref},{true_val},{false_val})'


def is_formula(value) -> bool:
    """Check if a value is an Excel formula string."""
    return isinstance(value, str) and value.startswith('=')


def write_values_sheet(wb, values_data: list, score_column: str):
    """Write a hidden _values sheet with pre-computed scores.

    Downstream modules (Module D) read this sheet instead of evaluating
    Excel formulas, ensuring reliable numeric reads regardless of whether
    the formula cache is populated.

    Args:
        wb: openpyxl Workbook (class sheets already written)
        values_data: List of dicts with keys 学号, 姓名, 班级, <score_column>
        score_column: Name of the score column (e.g. '学分绩点', '德育分', '拓展分')
    """
    ws = wb.create_sheet(title='_values')
    headers = ['学号', '姓名', '班级', score_column]
    header_font = Font(name='SimSun', size=10, bold=True)
    data_font = Font(name='SimSun', size=10)
    thin_border = Border(
        left=Side(style='thin', color='B0B0B0'),
        right=Side(style='thin', color='B0B0B0'),
        top=Side(style='thin', color='B0B0B0'),
        bottom=Side(style='thin', color='B0B0B0'),
    )
    center_align = Alignment(horizontal='center', vertical='center')

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    for ri, row_data in enumerate(values_data):
        for ci, h in enumerate(headers, 1):
            val = row_data.get(h, '')
            # Ensure student ID is stored as text (no decimals)
            if h == '学号' and val:
                if isinstance(val, float) and val == int(val):
                    val = str(int(val))
                else:
                    val = str(val)
            cell = ws.cell(row=ri + 2, column=ci, value=val)
            cell.font = data_font
            cell.alignment = center_align
            if h == '学号':
                cell.number_format = '@'
            cell.border = thin_border
            if h == '学号':
                cell.number_format = '@'

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 12
    ws.sheet_state = 'hidden'
