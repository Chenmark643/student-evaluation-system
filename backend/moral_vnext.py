"""Configurable continuation workflow for partially completed moral-score workbooks.

The legacy Module B remains available for from-scratch calculation.  This module
handles the complementary workflow: restore an existing score before clipping,
apply newly uploaded add/deduct items, review unmatched students, and export an
auditable workbook whose totals remain Excel formulas.
"""

from __future__ import annotations

import os
import re
from collections import defaultdict
from copy import copy
from dataclasses import dataclass, field

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from backend.utils.excel_writer import unique_path, write_values_sheet
from backend.utils.class_utils import class_matches_program, parse_class_name


MORAL_REFERENCE_HEADERS = [
    "学号", "姓名", "基础分", "评议分", "晚寝负责人",
    "早晚自习出勤", "课堂出勤", "出勤总",
    "宿舍卫生", "教室卫生", "卫生总",
    "团课出勤", "青年大学习", "通报批评", "违纪情况",
    "德育分",
]

MORAL_REFERENCE_ITEM_ALIASES = {
    "评议": "评议分",
    "评议分": "评议分",
    "评议奖励": "评议分",
    "晚寝负责人": "晚寝负责人",
    "早晚自习": "早晚自习出勤",
    "早晚自习出勤": "早晚自习出勤",
    "课堂出勤": "课堂出勤",
    "宿舍卫生": "宿舍卫生",
    "教室卫生": "教室卫生",
    "团课": "团课出勤",
    "团课出勤": "团课出勤",
    "青年大学习": "青年大学习",
    "通报批评": "通报批评",
    "违纪情况": "违纪情况",
}


def _moral_reference_layout(items: list[dict]) -> tuple[list[str], dict[str, list[dict]]]:
    """Map dynamic projects into the established moral-score table columns."""
    item_groups: dict[str, list[dict]] = defaultdict(list)
    custom_headers: list[str] = []
    reserved = set(MORAL_REFERENCE_HEADERS)
    for index, item in enumerate(items):
        name = str(item.get("name") or f"项目{index + 1}").strip()
        normalized = re.sub(r"\s+", "", name)
        header = MORAL_REFERENCE_ITEM_ALIASES.get(normalized)
        if not header:
            header = name
            if header in reserved or header in custom_headers:
                direction = "加分" if item.get("direction") == "add" else "扣分"
                header = f"{name}（{direction}）"
            suffix = 2
            candidate = header
            while candidate in reserved or candidate in custom_headers:
                candidate = f"{header}{suffix}"
                suffix += 1
            header = candidate
            custom_headers.append(header)
        item_groups[header].append(item)
    headers = MORAL_REFERENCE_HEADERS[:-1] + custom_headers + [MORAL_REFERENCE_HEADERS[-1]]
    return headers, item_groups


def _apply_moral_reference_style(sheet, max_row: int, max_col: int) -> None:
    """Reproduce the supplied 2025-2026 moral workbook's plain table format."""
    thin_black = Side(style="thin", color="000000")
    border = Border(left=thin_black, right=thin_black, top=thin_black, bottom=thin_black)
    white = PatternFill("solid", fgColor="FFFFFF")
    header_font = Font(name="宋体", size=11, bold=False, color="000000")
    data_font = Font(name="宋体", size=10, bold=False, color="000000")
    center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.fill = white
            cell.border = border
            cell.alignment = center_wrap
            cell.font = header_font if cell.row == 1 else data_font
    sheet.row_dimensions[1].height = 28
    for row_index in range(2, max_row + 1):
        sheet.row_dimensions[row_index].height = 26
    for col_index in range(1, max_col + 1):
        sheet.column_dimensions[get_column_letter(col_index)].width = 8.09
    sheet.freeze_panes = None
    sheet.auto_filter.ref = None
    sheet.sheet_view.showGridLines = False


@dataclass
class StudentRecord:
    key: str
    student_id: str = ""
    name: str = ""
    class_name: str = ""
    existing_raw: float | None = None
    existing_display: float | None = None
    existing_deduct: float | None = None
    existing_add: float | None = None
    existing_sheet: str = ""
    existing_row: int = 0
    item_values: dict[str, float] = field(default_factory=dict)
    source_rows: list[str] = field(default_factory=list)


def _clean_text(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return re.sub(r"\s+", "", text)


def _number(value) -> float | None:
    if value is None or str(value).strip() == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _excel_number(value, fallback: float = 0.0) -> str:
    """Return a locale-independent numeric literal suitable for an Excel formula."""
    number = _number(value)
    if number is None:
        number = fallback
    if abs(number) < 1e-12:
        number = 0.0
    return f"{number:.15g}"


def _cell(row: tuple, index) -> object:
    return row[index] if isinstance(index, int) and 0 <= index < len(row) else None


def _identity(student_id: str, class_name: str, name: str) -> str:
    if student_id:
        return f"id:{student_id}"
    return f"person:{_clean_text(class_name).casefold()}|{_clean_text(name).casefold()}"


def _iter_mapped_rows(path: str, mappings: dict):
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        for sheet_name, mapping in (mappings or {}).items():
            if not mapping or not mapping.get("enabled") or sheet_name not in workbook.sheetnames:
                continue
            worksheet = workbook[sheet_name]
            header_row = int(mapping.get("header_row") or 0) + 1
            for excel_row, row in enumerate(
                worksheet.iter_rows(min_row=header_row + 1, values_only=True),
                start=header_row + 1,
            ):
                action = (mapping.get("row_actions") or {}).get(str(excel_row), {})
                if action.get("action") == "exclude":
                    continue
                yield sheet_name, excel_row, tuple(row), mapping, action
    finally:
        workbook.close()


def _read_existing(config: dict, base_score: float) -> tuple[dict[str, StudentRecord], list[dict]]:
    students: dict[str, StudentRecord] = {}
    issues: list[dict] = []
    existing = config.get("existing") or {}
    path = existing.get("path", "")
    if not path or not os.path.isfile(path):
        raise ValueError("请选择有效的已有德育文件")
    major_filter = str(config.get("major_filter") or "").strip()

    for sheet, excel_row, row, mapping, action in _iter_mapped_rows(path, existing.get("mappings") or {}):
        if _row_is_outside_major(mapping, row, sheet, major_filter):
            continue
        student_id = _clean_text(_cell(row, mapping.get("id_col")))
        name = _clean_text(_cell(row, mapping.get("name_col")))
        class_name = _clean_text(_cell(row, mapping.get("class_col"))) or _clean_text(sheet)
        if not name and not student_id:
            continue
        key = _identity(student_id, class_name, name)
        display = _number(_cell(row, mapping.get("score_col")))
        raw = _number(_cell(row, mapping.get("raw_score_col")))
        total_deduct = _number(_cell(row, mapping.get("deduction_col")))
        total_add = _number(_cell(row, mapping.get("addition_col")))
        if raw is None and (total_deduct is not None or total_add is not None):
            raw = base_score - (total_deduct or 0.0) + (total_add or 0.0)
        if action.get("action") == "replace":
            display = _number(action.get("value"))
        if display is None and raw is not None:
            display = raw
        record = students.get(key) or StudentRecord(key, student_id, name, class_name)
        if record.existing_display is not None:
            issues.append({
                "level": "warn", "type": "duplicate_existing", "student_key": key,
                "student": name or student_id, "class_name": class_name,
                "message": "已有德育中出现重复学生，已使用最后一条记录",
                "source": f"{os.path.basename(path)} / {sheet} / 第{excel_row}行",
            })
        record.student_id = student_id or record.student_id
        record.name = name or record.name
        record.class_name = class_name or record.class_name
        record.existing_raw = raw
        record.existing_display = display
        record.existing_deduct = total_deduct
        record.existing_add = total_add
        record.existing_sheet = sheet
        record.existing_row = excel_row
        record.source_rows.append(f"{os.path.basename(path)} / {sheet} / 第{excel_row}行")
        students[key] = record
    if not students:
        raise ValueError("已有德育文件没有读取到学生，请检查工作表和列映射")
    return students, issues


def _match_key(students: dict[str, StudentRecord], student_id: str, class_name: str, name: str) -> str:
    direct = _identity(student_id, class_name, name)
    if direct in students:
        return direct
    target_name = _clean_text(name).casefold()
    target_class = _clean_text(class_name).casefold()
    candidates = [
        key for key, record in students.items()
        if target_name and _clean_text(record.name).casefold() == target_name
        and (not target_class or _clean_text(record.class_name).casefold() == target_class)
    ]
    return candidates[0] if len(candidates) == 1 else direct


def _edit_distance(left: str, right: str) -> int:
    """Small Unicode-safe Levenshtein distance for human-reviewed name hints."""
    left, right = _clean_text(left).casefold(), _clean_text(right).casefold()
    if left == right:
        return 0
    if not left:
        return len(right)
    if not right:
        return len(left)
    previous = list(range(len(right) + 1))
    for row_index, left_char in enumerate(left, 1):
        current = [row_index]
        for col_index, right_char in enumerate(right, 1):
            current.append(min(
                current[-1] + 1,
                previous[col_index] + 1,
                previous[col_index - 1] + (left_char != right_char),
            ))
        previous = current
    return previous[-1]


def _name_suggestions(
    students: dict[str, StudentRecord], name: str, class_name: str, limit: int = 5
) -> list[dict]:
    """Suggest one-character name corrections; never apply them automatically."""
    target_name = _clean_text(name)
    target_class = _clean_text(class_name).casefold()
    if not target_name:
        return []
    suggestions = []
    for record in students.values():
        if target_class and _clean_text(record.class_name).casefold() != target_class:
            continue
        distance = _edit_distance(target_name, record.name)
        if distance != 1:
            continue
        suggestions.append({
            "key": record.key, "student_id": record.student_id, "name": record.name,
            "class_name": record.class_name, "distance": distance,
        })
    suggestions.sort(key=lambda entry: (entry["distance"], entry["name"], entry["student_id"]))
    return suggestions[:limit]


def _row_is_outside_major(mapping: dict, row: tuple, sheet: str, major_filter: str) -> bool:
    """Return True only when a source row can be proven to belong elsewhere."""
    if not major_filter:
        return False
    explicit_class = _clean_text(_cell(row, mapping.get("class_col")))
    if explicit_class:
        return not class_matches_program(explicit_class, major_filter)
    parsed_sheet = parse_class_name(_clean_text(sheet))
    if parsed_sheet.get("grade"):
        return not class_matches_program(sheet, major_filter)
    return False


def _apply_items(config: dict, students: dict[str, StudentRecord], issues: list[dict]) -> None:
    for item_index, item in enumerate(config.get("items") or []):
        item_id = str(item.get("id") or f"item-{item_index + 1}")
        item_name = str(item.get("name") or f"项目{item_index + 1}").strip()
        direction = "add" if item.get("direction") == "add" else "deduct"
        value_mode = item.get("value_mode") or "signed"
        for source_index, source in enumerate(item.get("sources") or []):
            path = source.get("path", "")
            if not path or not os.path.isfile(path):
                issues.append({"level": "block", "type": "missing_file", "student": "",
                               "class_name": "", "message": f"{item_name}的来源文件不存在", "source": path})
                continue
            for sheet, excel_row, row, mapping, action in _iter_mapped_rows(path, source.get("mappings") or {}):
                student_id = _clean_text(_cell(row, mapping.get("id_col")))
                name = _clean_text(_cell(row, mapping.get("name_col")))
                class_name = _clean_text(_cell(row, mapping.get("class_col"))) or _clean_text(sheet)
                if not name and not student_id:
                    continue
                fresh_mode = config.get("mode") == "fresh"
                if _row_is_outside_major(
                    mapping, row, sheet, str(config.get("major_filter") or "")
                ):
                    continue
                raw_value = _number(_cell(row, mapping.get("score_col")))
                if action.get("action") == "replace":
                    raw_value = _number(action.get("value"))
                if raw_value is None:
                    if source.get("standard_template") and _cell(row, mapping.get("score_col")) in (None, ""):
                        continue
                    issues.append({
                        "level": "warn", "type": "invalid_item_value", "student": name or student_id,
                        "class_name": class_name, "message": f"{item_name}分数不是数字，已忽略",
                        "source": f"{os.path.basename(path)} / {sheet} / 第{excel_row}行",
                    })
                    continue
                matched_key = action.get("student_key") if action.get("action") == "match" else ""
                key = matched_key or _match_key(students, student_id, class_name, name)
                if key not in students:
                    suggestions = _name_suggestions(students, name, class_name) if fresh_mode else []
                    suggested = suggestions[0] if suggestions and (
                        len(suggestions) == 1 or suggestions[0]["distance"] < suggestions[1]["distance"]
                    ) else None
                    students[key] = StudentRecord(key, student_id, name, class_name)
                    issues.append({
                        "level": "block", "type": "missing_roster_student" if fresh_mode else "missing_existing_student", "student_key": key,
                        "student": name or student_id, "class_name": class_name,
                        "message": ("项目材料中的学生不在花名册，请修正映射或花名册"
                                    if fresh_mode else "新材料中的学生不在已有德育总表，需要补充已有原始分/显示分或排除"),
                        "source": f"{os.path.basename(path)} / {sheet} / 第{excel_row}行",
                        "item_id": item_id, "source_index": source_index,
                        "source_path": path, "sheet_name": sheet, "excel_row": excel_row,
                        "standard_template": bool(source.get("standard_template")),
                        "suggestions": suggestions,
                        "suggested_student_key": suggested.get("key") if suggested else "",
                    })
                if value_mode == "signed":
                    conflict = (direction == "deduct" and raw_value > 0) or (direction == "add" and raw_value < 0)
                    if conflict:
                        issues.append({
                            "level": "warn", "type": "sign_conflict", "student_key": key,
                            "student": name or student_id, "class_name": class_name,
                            "message": f"{item_name}的正负号与“{'加分' if direction == 'add' else '扣分'}”方向不一致，按绝对值计入",
                            "source": f"{os.path.basename(path)} / {sheet} / 第{excel_row}行",
                        })
                amount = abs(raw_value)
                record = students[key]
                record.item_values[item_id] = record.item_values.get(item_id, 0.0) + amount
                record.source_rows.append(f"{item_name}: {os.path.basename(path)} / {sheet} / 第{excel_row}行")


def _apply_manual_item_values(config: dict, students: dict[str, StudentRecord], issues: list[dict]) -> None:
    """Apply user-entered batch values after uploaded material values."""
    for item_index, item in enumerate(config.get("items") or []):
        item_id = str(item.get("id") or f"item-{item_index + 1}")
        item_name = str(item.get("name") or f"项目{item_index + 1}").strip()
        for key, raw_value in (item.get("manual_values") or {}).items():
            amount = _number(raw_value)
            if amount is None:
                issues.append({
                    "level": "warn", "type": "invalid_manual_value", "student_key": key,
                    "student": "", "class_name": "", "message": f"{item_name}的批量录入值不是数字，已忽略",
                    "source": "批量录入",
                })
                continue
            if key not in students:
                if config.get("major_filter"):
                    continue
                issues.append({
                    "level": "block", "type": "missing_manual_student", "student_key": key,
                    "student": "", "class_name": "", "message": f"{item_name}的批量录入学生不在当前名单",
                    "source": "批量录入",
                })
                continue
            students[key].item_values[item_id] = students[key].item_values.get(item_id, 0.0) + abs(amount)
            students[key].source_rows.append(f"{item_name}: 批量录入 {abs(amount):g}分")


def _apply_overrides(students: dict[str, StudentRecord], overrides: dict, issues: list[dict]) -> None:
    excluded = set()
    for key, override in (overrides or {}).items():
        if key not in students:
            continue
        if override.get("exclude"):
            excluded.add(key)
            continue
        raw = _number(override.get("raw"))
        display = _number(override.get("display"))
        if raw is not None:
            students[key].existing_raw = raw
        if display is not None:
            students[key].existing_display = display
    for key in excluded:
        students.pop(key, None)
    issues[:] = [issue for issue in issues if not (
        issue.get("type") == "missing_existing_student"
        and (issue.get("student_key") in excluded or (
            issue.get("student_key") in students
            and (
                students[issue.get("student_key")].existing_raw is not None
                or students[issue.get("student_key")].existing_display is not None
            )
        ))
    )]


def build_moral_preview(config: dict) -> dict:
    scoring = config.get("scoring") or {}
    base = float(scoring.get("base", 115))
    minimum = float(scoring.get("min", 0))
    maximum = float(scoring.get("max", 115))
    if minimum > maximum:
        raise ValueError("最低分不能大于最高分")
    students, issues = _read_existing(config, base)
    _apply_items(config, students, issues)
    _apply_manual_item_values(config, students, issues)
    _apply_overrides(students, config.get("overrides") or {}, issues)

    basis = "display" if scoring.get("continuation_basis") == "display" else "raw"
    rows = []
    for record in students.values():
        add_total = sum(
            record.item_values.get(str(item.get("id")), 0.0)
            for item in config.get("items") or [] if item.get("direction") == "add"
        )
        deduct_total = sum(
            record.item_values.get(str(item.get("id")), 0.0)
            for item in config.get("items") or [] if item.get("direction") != "add"
        )
        raw_basis = record.existing_raw
        display_basis = record.existing_display
        selected_basis = raw_basis if basis == "raw" else display_basis
        final = None if selected_basis is None else min(maximum, max(minimum, selected_basis + add_total - deduct_total))
        rows.append({
            "key": record.key, "student_id": record.student_id, "name": record.name,
            "class_name": record.class_name, "existing_raw": raw_basis,
            "existing_display": display_basis, "add_total": round(add_total, 4),
            "existing_deduct": record.existing_deduct, "existing_add": record.existing_add,
            "deduct_total": round(deduct_total, 4), "final": None if final is None else round(final, 4),
            "existing_sheet": record.existing_sheet, "existing_row": record.existing_row,
            "items": dict(record.item_values), "sources": list(record.source_rows),
        })
        if basis == "raw" and raw_basis is None:
            message = (
                "已有表没有原始分；如确认不需要保留超限缓冲，可用显示分补齐原始分"
                if display_basis is not None
                else "已有表没有可用的原始分或显示分，请补充分数或明确排除"
            )
            issues.append({
                "level": "block", "type": "raw_basis_unavailable", "student_key": record.key,
                "student": record.name or record.student_id, "class_name": record.class_name,
                "message": message,
                "source": "已有德育映射",
            })
        if basis == "display" and display_basis is None:
            issues.append({
                "level": "block", "type": "display_basis_unavailable", "student_key": record.key,
                "student": record.name or record.student_id, "class_name": record.class_name,
                "message": "该学生缺少已有显示分",
                "source": "已有德育映射",
            })

    difference_count = 0
    for record in students.values():
        if record.existing_raw is None or record.existing_display is None:
            continue
        add_total = sum(record.item_values.get(str(item.get("id")), 0.0)
                        for item in config.get("items") or [] if item.get("direction") == "add")
        deduct_total = sum(record.item_values.get(str(item.get("id")), 0.0)
                           for item in config.get("items") or [] if item.get("direction") != "add")
        raw_final = min(maximum, max(minimum, record.existing_raw + add_total - deduct_total))
        display_final = min(maximum, max(minimum, record.existing_display + add_total - deduct_total))
        if abs(raw_final - display_final) > 1e-9:
            difference_count += 1

    return {
        "success": True,
        "needs_review": any(issue.get("level") == "block" for issue in issues),
        "students": sorted(rows, key=lambda row: (row["class_name"], row["student_id"], row["name"])),
        "issues": issues,
        "summary": {
            "student_count": len(rows),
            "block_count": sum(1 for issue in issues if issue.get("level") == "block"),
            "warning_count": sum(1 for issue in issues if issue.get("level") == "warn"),
            "basis_difference_count": difference_count,
        },
    }


def export_moral_vnext(config: dict, preview: dict | None = None) -> dict:
    preview = preview or build_moral_preview(config)
    if preview.get("needs_review"):
        return preview
    output_dir = config.get("output_dir") or ""
    if not output_dir:
        raise ValueError("请选择输出目录")
    os.makedirs(output_dir, exist_ok=True)
    output_path = unique_path(os.path.join(output_dir, "德育分-补齐.xlsx"))
    scoring = config.get("scoring") or {}
    basis = "display" if scoring.get("continuation_basis") == "display" else "raw"
    items = config.get("items") or []

    existing = config.get("existing") or {}
    workbook = openpyxl.load_workbook(existing.get("path"), data_only=False)
    settings_name = "_德育新增设置"
    if settings_name in workbook.sheetnames:
        del workbook[settings_name]
    settings = workbook.create_sheet(settings_name)
    settings_rows = [
        ["德育续算设置", "当前值"],
        ["基础分", float(scoring.get("base", 115))],
        ["最低分", float(scoring.get("min", 0))],
        ["最高分", float(scoring.get("max", 115))],
        ["续算基准", basis],
        ["续算说明", "保留未截断原始分" if basis == "raw" else "按已有显示分重新起算"],
    ]
    for row in settings_rows:
        settings.append(row)
    settings.column_dimensions["A"].width = 22
    settings.column_dimensions["B"].width = 34

    values_data = []
    preview_by_location = {
        (row.get("existing_sheet"), int(row.get("existing_row") or 0)): row
        for row in preview.get("students") or [] if row.get("existing_sheet") and row.get("existing_row")
    }
    settings.append([])
    settings.append(["学生键", "原总扣分", "原总加分", "原显示分", "原始分"])
    for row in preview.get("students") or []:
        settings.append([
            row.get("key", ""), row.get("existing_deduct"), row.get("existing_add"),
            row.get("existing_display"), row.get("existing_raw"),
        ])
    settings.sheet_state = "hidden"

    touched_sheets = set()
    for sheet_name, mapping in (existing.get("mappings") or {}).items():
        if not mapping.get("enabled") or sheet_name not in workbook.sheetnames:
            continue
        sheet_locations = [
            (location, student) for location, student in preview_by_location.items()
            if location[0] == sheet_name
        ]
        if not sheet_locations:
            continue
        sheet = workbook[sheet_name]
        header_row = int(mapping.get("header_row") or 0) + 1
        original_last_col = max((cell.column for cell in sheet[header_row] if cell.value not in (None, "")), default=sheet.max_column)
        raw_col_original = mapping.get("raw_score_col") + 1 if isinstance(mapping.get("raw_score_col"), int) else None
        score_col_original = mapping.get("score_col") + 1 if isinstance(mapping.get("score_col"), int) else None
        deduction_col_original = mapping.get("deduction_col") + 1 if isinstance(mapping.get("deduction_col"), int) else None
        addition_col_original = mapping.get("addition_col") + 1 if isinstance(mapping.get("addition_col"), int) else None
        deduct_items = [item for item in items if item.get("direction") != "add"]
        add_items = [item for item in items if item.get("direction") == "add"]

        deduct_insert_at = deduction_col_original or addition_col_original or score_col_original or (original_last_col + 1)
        if deduct_items:
            sheet.insert_cols(deduct_insert_at, len(deduct_items))
        insertions = [(deduct_insert_at, len(deduct_items))] if deduct_items else []

        def shifted(original_col):
            if not isinstance(original_col, int):
                return None
            return original_col + sum(count for position, count in insertions if position <= original_col)

        add_insert_at = shifted(addition_col_original or score_col_original or (original_last_col + 1))
        if add_items:
            sheet.insert_cols(add_insert_at, len(add_items))
            insertions.append((addition_col_original or score_col_original or (original_last_col + 1), len(add_items)))

        def shifted_final(original_col):
            if not isinstance(original_col, int):
                return None
            return original_col + sum(count for position, count in insertions if position <= original_col)

        deduct_item_cols = list(range(deduct_insert_at, deduct_insert_at + len(deduct_items)))
        add_item_cols = list(range(add_insert_at, add_insert_at + len(add_items)))
        item_col_by_id = {
            **{str(item.get("id")): col for item, col in zip(deduct_items, deduct_item_cols)},
            **{str(item.get("id")): col for item, col in zip(add_items, add_item_cols)},
        }
        deduction_col = shifted_final(deduction_col_original)
        addition_col = shifted_final(addition_col_original)
        score_col = shifted_final(score_col_original)
        raw_col = shifted_final(raw_col_original)
        style_col = score_col or addition_col or deduction_col or shifted_final(original_last_col)

        for item in items:
            col = item_col_by_id[str(item.get("id"))]
            direction_text = "加分" if item.get("direction") == "add" else "扣分"
            header = sheet.cell(header_row, col, f"{item.get('name') or '未命名项目'}（{direction_text}）")
            source_header = sheet.cell(header_row, style_col)
            if source_header.has_style:
                header._style = copy(source_header._style)
            header.alignment = copy(source_header.alignment)
            source_width = sheet.column_dimensions[get_column_letter(style_col)].width
            sheet.column_dimensions[get_column_letter(col)].width = source_width or 8.09

        identity_cols = {
            value + 1 for value in (mapping.get("id_col"), mapping.get("name_col"), mapping.get("class_col"))
            if isinstance(value, int)
        }
        original_deduct_components = [
            col for col in range(1, deduction_col_original or 1)
            if col not in identity_cols and col != raw_col_original
            and sheet.cell(header_row, shifted_final(col)).value not in (None, "")
        ]
        original_add_components = [
            col for col in range((deduction_col_original or 0) + 1, addition_col_original or ((deduction_col_original or 0) + 1))
            if col not in identity_cols and sheet.cell(header_row, shifted_final(col)).value not in (None, "")
        ]

        for (_, row_index), student in sheet_locations:
            for item in items:
                col = item_col_by_id[str(item.get("id"))]
                item_cell = sheet.cell(row_index, col, student.get("items", {}).get(str(item.get("id")), 0.0))
                source_cell = sheet.cell(row_index, style_col)
                if source_cell.has_style:
                    item_cell._style = copy(source_cell._style)
            add_refs = [f"{get_column_letter(col)}{row_index}" for col in add_item_cols]
            deduct_refs = [f"{get_column_letter(col)}{row_index}" for col in deduct_item_cols]
            add_expr = f"SUM({','.join(add_refs)})" if add_refs else "0"
            deduct_expr = f"SUM({','.join(deduct_refs)})" if deduct_refs else "0"
            if deduction_col:
                component_refs = [f"{get_column_letter(shifted_final(col))}{row_index}" for col in original_deduct_components] + deduct_refs
                fallback = f"{_excel_number(student.get('existing_deduct'))}+{deduct_expr}"
                sheet.cell(row_index, deduction_col, f"=SUM({','.join(component_refs)})" if component_refs else f"={fallback}")
            if addition_col:
                component_refs = [f"{get_column_letter(shifted_final(col))}{row_index}" for col in original_add_components] + add_refs
                fallback = f"{_excel_number(student.get('existing_add'))}+{add_expr}"
                sheet.cell(row_index, addition_col, f"=SUM({','.join(component_refs)})" if component_refs else f"={fallback}")

            if basis == "raw" and raw_col and raw_col != score_col:
                basis_value = f"{get_column_letter(raw_col)}{row_index}"
            elif basis == "raw":
                basis_value = _excel_number(student.get("existing_raw"))
            else:
                basis_value = _excel_number(student.get("existing_display"))
            basis_expr = f"{basis_value}+{add_expr}-{deduct_expr}"
            if score_col:
                minimum = _excel_number(scoring.get("min"), 0)
                maximum = _excel_number(scoring.get("max"), 115)
                sheet.cell(row_index, score_col, f"=MIN({maximum},MAX({minimum},{basis_expr}))")
            values_data.append({
                "学号": student.get("student_id", ""), "姓名": student.get("name", ""),
                "班级": student.get("class_name", ""), "德育分": student.get("final", 0),
            })
        touched_sheets.add(sheet_name)

    if "异常与审计" in workbook.sheetnames:
        del workbook["异常与审计"]
    audit = workbook.create_sheet("异常与审计")
    audit.append(["级别", "类型", "班级", "学生", "说明", "来源"])
    for issue in preview.get("issues") or []:
        audit.append([
            issue.get("level", ""), issue.get("type", ""), issue.get("class_name", ""),
            issue.get("student", ""), issue.get("message", ""), issue.get("source", ""),
        ])
    audit.freeze_panes = "A2"
    for col, width in enumerate([10, 24, 16, 14, 52, 48], 1):
        audit.column_dimensions[get_column_letter(col)].width = width
    if "_values" in workbook.sheetnames:
        del workbook["_values"]
    if values_data:
        write_values_sheet(workbook, values_data, "德育分")
    workbook.save(output_path)
    workbook.close()
    return {
        "success": True, "needs_review": False, "output": output_path,
        "student_count": len(values_data), "class_count": len(touched_sheets),
        "summary": preview.get("summary") or {},
    }


def process_moral_vnext(config: dict) -> dict:
    preview = build_moral_preview(config or {})
    if preview.get("needs_review"):
        return preview
    return export_moral_vnext(config or {}, preview)


def build_moral_fresh_preview(config: dict) -> dict:
    """Build a from-scratch preview using a roster and dynamic add/deduct items."""
    from backend.module_b_moral import _load_roster

    scoring = config.get("scoring") or {}
    base = float(scoring.get("base", 80))
    minimum = float(scoring.get("min", 0))
    maximum = float(scoring.get("max", 115))
    if minimum > maximum:
        raise ValueError("最低分不能大于最高分")
    roster_path = config.get("roster_path") or ""
    if not roster_path or not os.path.isfile(roster_path):
        raise ValueError("请选择有效的花名册文件")
    roster = _load_roster(roster_path)
    major_filter = str(config.get("major_filter") or "").strip()
    if major_filter:
        roster = {
            student_id: info for student_id, info in roster.items()
            if class_matches_program(info.get("class", ""), major_filter)
        }
    if not roster:
        if major_filter:
            raise ValueError(f"花名册中没有找到当前专业“{major_filter}”的学生，请检查专业名称或班级列")
        raise ValueError("花名册中没有读取到学生")
    students = {}
    for student_id, info in roster.items():
        name = _clean_text(info.get("name"))
        class_name = _clean_text(info.get("class"))
        key = _identity(_clean_text(student_id), class_name, name)
        students[key] = StudentRecord(
            key=key, student_id=_clean_text(student_id), name=name, class_name=class_name,
            existing_raw=base, existing_display=base,
        )
    issues = []
    fresh_config = dict(config)
    fresh_config["mode"] = "fresh"
    _apply_items(fresh_config, students, issues)
    _apply_manual_item_values(fresh_config, students, issues)
    rows = []
    for record in students.values():
        add_total = sum(
            record.item_values.get(str(item.get("id")), 0.0)
            for item in config.get("items") or [] if item.get("direction") == "add"
        )
        deduct_total = sum(
            record.item_values.get(str(item.get("id")), 0.0)
            for item in config.get("items") or [] if item.get("direction") != "add"
        )
        final = min(maximum, max(minimum, base + add_total - deduct_total))
        rows.append({
            "key": record.key, "student_id": record.student_id, "name": record.name,
            "class_name": record.class_name, "items": dict(record.item_values),
            "add_total": round(add_total, 4), "deduct_total": round(deduct_total, 4),
            "final": round(final, 4), "sources": list(record.source_rows),
        })
    return {
        "success": True,
        "needs_review": any(issue.get("level") == "block" for issue in issues),
        "students": sorted(rows, key=lambda row: (row["class_name"], row["student_id"], row["name"])),
        "issues": issues,
        "summary": {
            "student_count": len(rows),
            "block_count": sum(1 for issue in issues if issue.get("level") == "block"),
            "warning_count": sum(1 for issue in issues if issue.get("level") == "warn"),
        },
    }


def export_moral_fresh(config: dict, preview: dict | None = None) -> dict:
    preview = preview or build_moral_fresh_preview(config)
    if preview.get("needs_review"):
        return preview
    output_dir = config.get("output_dir") or ""
    if not output_dir:
        raise ValueError("请选择输出目录")
    os.makedirs(output_dir, exist_ok=True)
    output_path = unique_path(os.path.join(output_dir, "德育分-从零建立.xlsx"))
    scoring = config.get("scoring") or {}
    items = config.get("items") or []
    workbook = openpyxl.Workbook()
    settings = workbook.active
    settings.title = "计分设置"
    for row in [
        ["德育计分设置", "当前值"],
        ["基础分", float(scoring.get("base", 80))],
        ["最低分", float(scoring.get("min", 0))],
        ["最高分", float(scoring.get("max", 115))],
    ]:
        settings.append(row)
    settings.column_dimensions["A"].width = 22
    settings.column_dimensions["B"].width = 20

    grouped = defaultdict(list)
    for row in preview.get("students") or []:
        grouped[row.get("class_name") or "未分班"].append(row)
    values_data = []
    headers, item_groups = _moral_reference_layout(items)
    first_class_sheet = None
    for class_name, class_rows in grouped.items():
        sheet = workbook.create_sheet(str(class_name)[:31])
        if first_class_sheet is None:
            first_class_sheet = sheet
        sheet.append(headers)
        header_cols = {header: index for index, header in enumerate(headers, 1)}
        scoring_cols = [
            index for index, header in enumerate(headers, 1)
            if header not in {"学号", "姓名", "出勤总", "卫生总", "德育分"}
        ]
        for row_index, student in enumerate(class_rows, 2):
            signed_values = {}
            for header, grouped_items in item_groups.items():
                signed_values[header] = sum(
                    student.get("items", {}).get(str(item.get("id")), 0.0)
                    * (1 if item.get("direction") == "add" else -1)
                    for item in grouped_items
                )
            for col_index, header in enumerate(headers, 1):
                cell = sheet.cell(row_index, col_index)
                if header == "学号":
                    cell.value = student.get("student_id", "")
                    cell.number_format = "@"
                elif header == "姓名":
                    cell.value = student.get("name", "")
                elif header == "基础分":
                    cell.value = float(scoring.get("base", 80))
                elif header == "出勤总":
                    left = get_column_letter(header_cols["早晚自习出勤"])
                    right = get_column_letter(header_cols["课堂出勤"])
                    cell.value = f"=SUM({left}{row_index}:{right}{row_index})"
                elif header == "卫生总":
                    left = get_column_letter(header_cols["宿舍卫生"])
                    right = get_column_letter(header_cols["教室卫生"])
                    cell.value = f"=SUM({left}{row_index}:{right}{row_index})"
                elif header == "德育分":
                    refs = ",".join(f"{get_column_letter(col)}{row_index}" for col in scoring_cols)
                    minimum = _excel_number(scoring.get("min"), 0)
                    maximum = _excel_number(scoring.get("max"), 115)
                    cell.value = (
                        f"=MIN({maximum},MAX({minimum},SUM({refs})))"
                    )
                else:
                    value = signed_values.get(header, 0.0)
                    cell.value = value if abs(value) > 1e-12 else None
            values_data.append({"学号": student.get("student_id", ""), "姓名": student.get("name", ""), "班级": class_name, "德育分": student.get("final", 0)})
        _apply_moral_reference_style(sheet, sheet.max_row, len(headers))
        sheet.page_setup.orientation = sheet.ORIENTATION_PORTRAIT
        sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.page_margins.left = 0.75
        sheet.page_margins.right = 0.75
        sheet.page_margins.top = 1.0
        sheet.page_margins.bottom = 1.0
    if values_data:
        write_values_sheet(workbook, values_data, "德育分")
    settings.sheet_state = "hidden"
    if first_class_sheet is not None:
        workbook.active = workbook.index(first_class_sheet)
    workbook.save(output_path)
    workbook.close()
    return {
        "success": True, "needs_review": False, "output": output_path,
        "student_count": len(values_data), "class_count": len(grouped),
        "summary": preview.get("summary") or {},
    }


def process_moral_fresh(config: dict) -> dict:
    preview = build_moral_fresh_preview(config or {})
    if preview.get("needs_review"):
        return preview
    return export_moral_fresh(config or {}, preview)


def list_moral_students(config: dict) -> dict:
    """Return stable student keys for the batch-entry picker."""
    mode = config.get("mode") or "continue"
    if mode == "fresh":
        from backend.module_b_moral import _load_roster

        roster_path = config.get("roster_path") or ""
        if not roster_path or not os.path.isfile(roster_path):
            raise ValueError("请先选择有效的花名册文件")
        roster = _load_roster(roster_path)
        major_filter = str(config.get("major_filter") or "").strip()
        if major_filter:
            roster = {
                student_id: info for student_id, info in roster.items()
                if class_matches_program(info.get("class", ""), major_filter)
            }
        students = []
        for student_id, info in roster.items():
            name = _clean_text(info.get("name"))
            class_name = _clean_text(info.get("class"))
            students.append({
                "key": _identity(_clean_text(student_id), class_name, name),
                "student_id": _clean_text(student_id), "name": name, "class_name": class_name,
            })
    else:
        scoring = config.get("scoring") or {}
        records, _ = _read_existing(config, float(scoring.get("base", 115)))
        students = [{
            "key": record.key, "student_id": record.student_id,
            "name": record.name, "class_name": record.class_name,
        } for record in records.values()]
    return {
        "success": True,
        "students": sorted(students, key=lambda row: (row["class_name"], row["student_id"], row["name"])),
    }
