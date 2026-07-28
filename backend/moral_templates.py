"""Standard one-project moral-score templates and upload validation."""

from __future__ import annotations

import os
import re
import shutil
from collections import Counter
from pathlib import Path

import openpyxl

from config import BASE_DIR


MORAL_PROJECT_TEMPLATES = (
    {"key": "evaluation", "code": "DY-PY-01", "name": "评议分", "filename": "德育项目模板-评议分.xlsx"},
    {"key": "night_manager", "code": "DY-WQ-01", "name": "晚寝负责人", "filename": "德育项目模板-晚寝负责人.xlsx"},
    {"key": "self_study", "code": "DY-ZX-01", "name": "早晚自习出勤", "filename": "德育项目模板-早晚自习出勤.xlsx"},
    {"key": "class_attendance", "code": "DY-KT-01", "name": "课堂出勤", "filename": "德育项目模板-课堂出勤.xlsx"},
    {"key": "dorm_hygiene", "code": "DY-SS-01", "name": "宿舍卫生", "filename": "德育项目模板-宿舍卫生.xlsx"},
    {"key": "classroom_hygiene", "code": "DY-JS-01", "name": "教室卫生", "filename": "德育项目模板-教室卫生.xlsx"},
    {"key": "league_class", "code": "DY-TK-01", "name": "团课出勤", "filename": "德育项目模板-团课出勤.xlsx"},
    {"key": "youth_study", "code": "DY-QN-01", "name": "青年大学习", "filename": "德育项目模板-青年大学习.xlsx"},
    {"key": "criticism", "code": "DY-TB-01", "name": "通报批评", "filename": "德育项目模板-通报批评.xlsx"},
    {"key": "discipline", "code": "DY-WJ-01", "name": "违纪情况", "filename": "德育项目模板-违纪情况.xlsx"},
)

_PROJECT_BY_KEY = {item["key"]: item for item in MORAL_PROJECT_TEMPLATES}
_PROJECT_BY_CODE = {item["code"]: item for item in MORAL_PROJECT_TEMPLATES}
_PROJECT_BY_NAME = {item["name"]: item for item in MORAL_PROJECT_TEMPLATES}
_DATA_SHEET = "数据录入"
_HEADER_ROW = 4


def _template_directory() -> Path:
    bundled = Path(BASE_DIR) / "moral_templates"
    if bundled.is_dir():
        return bundled
    return Path(BASE_DIR) / "outputs" / "moral-project-templates"


def list_moral_project_templates() -> list[dict]:
    directory = _template_directory()
    return [
        {
            **item,
            "available": (directory / item["filename"]).is_file(),
            "add_header": f"{item['name']}加分",
            "deduct_header": f"{item['name']}扣分",
        }
        for item in MORAL_PROJECT_TEMPLATES
    ]


def _clean(value) -> str:
    return re.sub(r"\s+", "", str(value or "")).strip()


def _number(value):
    if value is None or value == "" or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text or text.startswith("="):
        return None
    try:
        return float(text)
    except (TypeError, ValueError):
        return None


def _project_from_sheet(sheet) -> dict | None:
    code = str(sheet["B2"].value or "").strip().upper()
    if code in _PROJECT_BY_CODE:
        return _PROJECT_BY_CODE[code]
    project_name = _clean(sheet["D2"].value)
    if project_name in _PROJECT_BY_NAME:
        return _PROJECT_BY_NAME[project_name]
    headers = [_clean(sheet.cell(_HEADER_ROW, col).value) for col in range(1, 6)]
    for project in MORAL_PROJECT_TEMPLATES:
        if headers == ["班级", "姓名", f"{project['name']}加分", f"{project['name']}扣分", "备注"]:
            return project
    return None


def analyze_moral_project_template(path: str) -> dict:
    result = {"success": False, "path": path, "filename": os.path.basename(path or "")}
    if not path or not os.path.isfile(path):
        return {**result, "error": "文件不存在"}
    if Path(path).suffix.lower() != ".xlsx":
        return {**result, "error": "标准项目模板只支持 .xlsx 文件"}
    if os.path.getsize(path) > 20 * 1024 * 1024:
        return {**result, "error": "文件超过20MB，请删除无关图片或工作表后重试"}

    try:
        workbook = openpyxl.load_workbook(path, data_only=False, read_only=True)
    except Exception as exc:
        return {**result, "error": f"无法打开工作簿：{exc}"}
    try:
        if _DATA_SHEET not in workbook.sheetnames:
            return {**result, "error": "缺少“数据录入”工作表，请使用软件提供的标准模板"}
        sheet = workbook[_DATA_SHEET]
        project = _project_from_sheet(sheet)
        if not project:
            return {**result, "error": "无法识别模板项目，模板编号或固定表头已被修改"}
        expected = ["班级", "姓名", f"{project['name']}加分", f"{project['name']}扣分", "备注"]
        actual = [_clean(sheet.cell(_HEADER_ROW, col).value) for col in range(1, 6)]
        if actual != expected:
            return {
                **result,
                "error": f"第{_HEADER_ROW}行表头不完整，应为：{'、'.join(expected)}",
                "project_key": project["key"],
                "project_name": project["name"],
            }

        errors, warnings, records = [], [], []
        identities = Counter()
        add_count = deduct_count = 0
        for excel_row, row in enumerate(
            sheet.iter_rows(min_row=_HEADER_ROW + 1, max_col=5, values_only=True),
            start=_HEADER_ROW + 1,
        ):
            class_name, name = str(row[0] or "").strip(), str(row[1] or "").strip()
            add_raw, deduct_raw = row[2], row[3]
            if not any(value not in (None, "") for value in row):
                continue
            if not class_name and not name and add_raw in (None, "") and deduct_raw in (None, ""):
                continue
            if not class_name or not name:
                errors.append(f"第{excel_row}行：班级和姓名必须同时填写")
                continue
            add_value, deduct_value = _number(add_raw), _number(deduct_raw)
            if add_raw not in (None, "") and add_value is None:
                errors.append(f"第{excel_row}行：加分必须是非负数字")
                continue
            if deduct_raw not in (None, "") and deduct_value is None:
                errors.append(f"第{excel_row}行：扣分必须是非负数字")
                continue
            add_value, deduct_value = add_value or 0.0, deduct_value or 0.0
            if add_value < 0 or deduct_value < 0:
                errors.append(f"第{excel_row}行：加分、扣分均填写正数，不要输入负号")
                continue
            if add_value > 0 and deduct_value > 0:
                errors.append(f"第{excel_row}行：同一行不能同时填写加分和扣分，请拆成两行")
                continue
            if add_value == 0 and deduct_value == 0:
                warnings.append(f"第{excel_row}行：加分和扣分均为空或0，运行时将忽略")
                continue
            identity = (_clean(class_name).casefold(), _clean(name).casefold())
            identities[identity] += 1
            add_count += int(add_value > 0)
            deduct_count += int(deduct_value > 0)
            records.append(excel_row)
        duplicate_count = sum(count - 1 for count in identities.values() if count > 1)
        if duplicate_count:
            warnings.append(f"发现{duplicate_count}条重复姓名记录，系统会按班级+姓名自动累计")
        if not records and not errors:
            warnings.append("模板中尚无有效计分记录")
        return {
            **result,
            "success": not errors,
            "error": "；".join(errors[:8]) if errors else "",
            "errors": errors,
            "warnings": warnings,
            "project_key": project["key"],
            "project_name": project["name"],
            "template_code": project["code"],
            "sheet_name": _DATA_SHEET,
            "header_row": _HEADER_ROW - 1,
            "row_count": len(records),
            "add_count": add_count,
            "deduct_count": deduct_count,
            "duplicate_count": duplicate_count,
        }
    finally:
        workbook.close()


def analyze_moral_project_templates(paths: list[str]) -> dict:
    files = [analyze_moral_project_template(str(path)) for path in (paths or [])]
    return {
        "success": True,
        "files": files,
        "valid_count": sum(1 for item in files if item.get("success")),
        "error_count": sum(1 for item in files if not item.get("success")),
    }


def copy_moral_project_templates(project_key: str, output_dir: str) -> dict:
    if not output_dir or not os.path.isdir(output_dir):
        return {"success": False, "error": "请选择有效的保存目录"}
    projects = MORAL_PROJECT_TEMPLATES if project_key == "all" else (
        (_PROJECT_BY_KEY[project_key],) if project_key in _PROJECT_BY_KEY else ()
    )
    if not projects:
        return {"success": False, "error": "找不到该德育项目模板"}
    source_dir = _template_directory()
    outputs = []
    for project in projects:
        source = source_dir / project["filename"]
        if not source.is_file():
            return {"success": False, "error": f"安装包中缺少模板：{project['name']}"}
        destination = Path(output_dir) / project["filename"]
        if destination.exists():
            stem, suffix, index = destination.stem, destination.suffix, 2
            while destination.exists():
                destination = Path(output_dir) / f"{stem} ({index}){suffix}"
                index += 1
        shutil.copy2(source, destination)
        outputs.append(str(destination))
    return {"success": True, "outputs": outputs, "count": len(outputs)}
