"""
Parser for absence statistics tables (旷课统计).
Handles interleaved format: class header row → data rows → class header row → ...
"""

import pandas as pd
from openpyxl import load_workbook


def parse_absence_statistics(filepath: str) -> dict:
    """Parse overall absence statistics file.

    Expected format:
    - Interleaved: class name in col 0, then student rows with NaN in col 0
    - Data rows: 姓名, 学号, 班级, 所缺课程, 任课教师, 缺课学时, 合计学时, 缺课日期

    Args:
        filepath: Path to the absence statistics .xlsx file

    Returns:
        Dict of {student_id: {'name': ..., 'class': ..., 'absent_hours': float}}
    """
    wb = load_workbook(filepath, data_only=True)
    results = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        current_class = ''
        in_data_section = False

        for row in rows:
            if not row:
                continue

            col0 = str(row[0]).strip() if row[0] else ''
            col1 = str(row[1]).strip() if len(row) > 1 and row[1] else ''

            # Detect class header row (col 0 has class name, col 1 is '姓名')
            if col1 == '姓名':
                in_data_section = True
                continue

            # Detect class section header (col 0 has class name pattern)
            if col0 and not col1:
                # Could be a class name - check if it looks like a class identifier
                if any(c.isdigit() for c in col0) or '级' in col0:
                    current_class = col0
                    in_data_section = True
                    continue

            # Data row
            if in_data_section and col1 and col1 not in ['姓名', '学号', '班级']:
                name = col1
                student_id = str(row[2]).strip() if len(row) > 2 and row[2] else ''
                class_name = str(row[3]).strip() if len(row) > 3 and row[3] else current_class

                # Absent hours (缺课学时) is typically col 6
                absent_hours = 0.0
                if len(row) > 6 and row[6]:
                    try:
                        absent_hours = float(row[6])
                    except (ValueError, TypeError):
                        pass

                if student_id and student_id not in results:
                    results[student_id] = {
                        'name': name,
                        'class': class_name,
                        'absent_hours': 0.0,
                    }
                elif student_id:
                    results[student_id]['absent_hours'] += absent_hours
                elif name:
                    # No student ID — use class+name as key
                    key = f"{class_name}|{name}"
                    if key not in results:
                        results[key] = {
                            'name': name,
                            'class': class_name,
                            'absent_hours': 0.0,
                            'student_id': '',
                        }
                    results[key]['absent_hours'] += absent_hours

    wb.close()
    return results
