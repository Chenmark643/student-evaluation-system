"""
Parser for discipline/inspection statistics (信息部秘书处纪检卫生统计表).
Handles evening self-study deductions and classroom hygiene scores.
"""

from openpyxl import load_workbook


def parse_discipline_stats(filepath: str) -> dict:
    """Parse discipline and hygiene statistics file.

    Contains two sheets:
    - Sheet 1 (晚自习): 班级, 姓名, 合计学时, 晚自习扣分数
      Deduction = 合计学时 * 2
    - Sheet 2 (宿舍卫生检查): 班级, 姓名, 个人分, 宿舍分, 宿舍卫生得分

    Args:
        filepath: Path to the discipline statistics .xlsx file

    Returns:
        Dict of {(class, name): {
            'evening_self_study_deduction': float,
            'classroom_hygiene': float,
            'dorm_hygiene_score': float,
        }}
    """
    wb = load_workbook(filepath, data_only=True)
    results = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if len(rows) < 2:
            continue

        # Find header row
        header_idx = None
        for i, row in enumerate(rows):
            if row and any('姓名' in str(c) for c in row if c):
                header_idx = i
                break
        if header_idx is None:
            continue

        headers = [str(c).strip() if c else '' for c in rows[header_idx]]

        # Determine sheet type
        is_evening_study = any('晚自习' in h or '扣分' in h for h in headers)
        is_hygiene = any('卫生' in h or '宿舍分' in h for h in headers)

        name_col = _find_col(headers, ['姓名'])
        class_col = _find_col(headers, ['班级'])

        if name_col is None:
            continue

        for row in rows[header_idx + 1:]:
            if not row or len(row) <= name_col:
                continue

            name = str(row[name_col]).strip() if row[name_col] else ''
            if not name or name in ['姓名', '班级', 'NaN', 'nan', '']:
                continue

            class_name = str(row[class_col]).strip() if class_col is not None and len(row) > class_col and row[class_col] else ''
            key = (class_name, name)

            if key not in results:
                results[key] = {
                    'class': class_name,
                    'name': name,
                    'evening_self_study_deduction': 0.0,
                    'classroom_hygiene': 0.0,
                    'dorm_hygiene_score': 0.0,
                }

            if is_evening_study:
                # 晚自习扣分数 column
                deduction_col = _find_col(headers, ['扣分数', '扣分'])
                if deduction_col is not None and len(row) > deduction_col:
                    try:
                        results[key]['evening_self_study_deduction'] = float(row[deduction_col]) if row[deduction_col] else 0.0
                    except (ValueError, TypeError):
                        pass
                # Also try to get from 合计学时 * 2
                hours_col = _find_col(headers, ['合计学时', '学时'])
                if hours_col is not None and len(row) > hours_col:
                    try:
                        hours = float(row[hours_col]) if row[hours_col] else 0.0
                        if hours > 0:
                            results[key]['evening_self_study_deduction'] = hours * 2.0
                    except (ValueError, TypeError):
                        pass

            elif is_hygiene:
                # 宿舍卫生得分 = 个人分 + 宿舍分
                personal_col = _find_col(headers, ['个人分'])
                dorm_col = _find_col(headers, ['宿舍分'])
                total_col = _find_col(headers, ['宿舍卫生得分', '卫生得分'])

                if total_col is not None and len(row) > total_col:
                    try:
                        results[key]['dorm_hygiene_score'] = float(row[total_col]) if row[total_col] else 0.0
                    except (ValueError, TypeError):
                        pass
                elif personal_col is not None and dorm_col is not None:
                    personal = 0.0
                    dorm = 0.0
                    try:
                        if len(row) > personal_col and row[personal_col]:
                            personal = float(row[personal_col])
                    except (ValueError, TypeError):
                        pass
                    try:
                        if len(row) > dorm_col and row[dorm_col]:
                            dorm = float(row[dorm_col])
                    except (ValueError, TypeError):
                        pass
                    results[key]['dorm_hygiene_score'] = personal + dorm

    wb.close()
    return results


def _find_col(headers: list, keywords: list) -> int:
    """Find column index matching any keyword."""
    for kw in keywords:
        for i, h in enumerate(headers):
            if kw in h:
                return i
    return None
