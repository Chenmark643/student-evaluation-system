"""
Parser for organization class (团课) absence statistics.
Handles format: class header, then (name, deduction) column pairs.
"""

from openpyxl import load_workbook


def parse_org_class_absence(filepath: str) -> dict:
    """Parse organization class absence statistics (团课旷课统计表).

    Expected format:
    - Class name as section header in col 0
    - Subsequent rows: student name in col 1/3, deduction score in col 2/4
    - Two columns of (name, deduction) pairs

    Args:
        filepath: Path to the organization class absence .xlsx file

    Returns:
        Dict of {(class, name): {'org_class_deduction': float}}
    """
    wb = load_workbook(filepath, data_only=True)
    results = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if len(rows) < 2:
            continue

        current_class = ''

        for row in rows:
            if not row:
                continue

            col0 = str(row[0]).strip() if row[0] else ''

            # Detect class header (has digits or specific class name pattern)
            if col0 and not any(
                str(row[i]).strip() if len(row) > i and row[i] else ''
                for i in [1, 3]
            ):
                # Could be a class name
                has_digit = any(c.isdigit() for c in col0)
                if has_digit or '班' in col0:
                    current_class = col0
                    continue

            # Data row: name in col1 + deduction in col2, name in col3 + deduction in col4
            for name_col, score_col in [(1, 2), (3, 4)]:
                if len(row) > max(name_col, score_col):
                    name = str(row[name_col]).strip() if row[name_col] else ''
                    if not name or name in ['姓名', 'NaN', 'nan', '']:
                        continue

                    try:
                        score = float(row[score_col]) if row[score_col] else 0.0
                    except (ValueError, TypeError):
                        score = 0.0

                    if score == 0.0:
                        continue

                    key = (current_class, name)
                    if key not in results:
                        results[key] = {
                            'class': current_class,
                            'name': name,
                            'org_class_deduction': 0.0,
                        }
                    results[key]['org_class_deduction'] += score

    wb.close()
    return results
