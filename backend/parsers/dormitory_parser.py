"""
Parser for dormitory hygiene tables (宿舍卫生表).
Handles the format: building, room, bed number, student name, class, score.
"""

import pandas as pd
from openpyxl import load_workbook


def parse_dormitory_hygiene(filepath: str) -> dict:
    """Parse dormitory hygiene table.

    Expected format:
    - Row 0: Class name header
    - Row 1: Column headers (公寓楼, 宿舍号, 床位号, 姓名, 班级, 得分情况)
    - Rows 2+: Student data with forward-filled building/room info

    Args:
        filepath: Path to the dormitory hygiene .xlsx file

    Returns:
        Dict of {(class_name, student_name): {'dormitory_score': float}}
    """
    wb = load_workbook(filepath, data_only=True)
    results = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if len(rows) < 3:
            continue

        # Find header row (contains '姓名' and '得分情况')
        header_idx = None
        for i, row in enumerate(rows):
            if row and any('姓名' in str(c) for c in row if c):
                header_idx = i
                break

        if header_idx is None:
            continue

        headers = [str(c).strip() if c else '' for c in rows[header_idx]]

        # Find column indices
        name_col = _find_column(headers, ['姓名'])
        score_col = _find_column(headers, ['得分情况', '得分', '分数'])
        class_col = _find_column(headers, ['班级'])

        if name_col is None or score_col is None:
            continue

        class_name = str(sheet_name).strip()

        for row in rows[header_idx + 1:]:
            if not row or len(row) <= max(name_col, score_col):
                continue

            name = str(row[name_col]).strip() if row[name_col] else ''
            if not name or name in ['姓名', '班级', 'NaN', 'nan']:
                continue

            # Get class from row if available, otherwise use sheet name
            if class_col is not None and len(row) > class_col and row[class_col]:
                row_class = str(row[class_col]).strip()
                if row_class and row_class != 'nan':
                    class_name = row_class

            try:
                score = float(row[score_col]) if row[score_col] else 20.0
            except (ValueError, TypeError):
                score = 20.0

            key = (class_name, name)
            if key not in results:
                results[key] = {'dormitory_score': score, 'class': class_name}
            else:
                results[key]['dormitory_score'] = max(
                    results[key]['dormitory_score'], score
                )

    wb.close()
    return results


def _find_column(headers: list, keywords: list) -> int:
    """Find column index matching any of the keywords."""
    for kw in keywords:
        for i, h in enumerate(headers):
            if kw in h:
                return i
    return None
