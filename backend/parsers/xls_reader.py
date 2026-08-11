from __future__ import annotations
"""
Excel file reader — supports .xls (via xlrd) and .xlsx (via openpyxl/pandas).
"""

import os
import pandas as pd
import numpy as np


def read_raw_xls(filepath: str) -> pd.DataFrame:
    """Read a legacy .xls or .xlsx file and return a DataFrame.

    Handles:
    - Legacy BIFF format (.xls) via xlrd
    - Modern .xlsx via openpyxl
    - GBK/GB2312 Chinese encoding
    - Interleaved headers and data rows

    Args:
        filepath: Path to the .xls or .xlsx file

    Returns:
        pandas DataFrame with the sheet data
    """
    ext = os.path.splitext(filepath)[1].lower()

    if ext == '.xls':
        return _read_xls_xlrd(filepath)
    else:
        return _read_xlsx_openpyxl(filepath)


def _read_xls_xlrd(filepath: str) -> pd.DataFrame:
    """Read .xls using xlrd (handles legacy BIFF format)."""
    import xlrd

    wb = xlrd.open_workbook(filepath, encoding_override='gbk')

    all_data = []
    for sheet_idx in range(wb.nsheets):
        sheet = wb.sheet_by_index(sheet_idx)
        if sheet.nrows < 2:
            continue

        # Convert sheet to list of lists
        data = []
        for row_idx in range(sheet.nrows):
            row = []
            for col_idx in range(sheet.ncols):
                cell = sheet.cell(row_idx, col_idx)
                if cell.ctype == xlrd.XL_CELL_EMPTY:
                    row.append(None)
                elif cell.ctype == xlrd.XL_CELL_NUMBER:
                    # Preserve integers
                    val = float(cell.value)
                    if val == int(val) and abs(val) < 1e10:
                        row.append(int(val))
                    else:
                        row.append(val)
                else:
                    row.append(str(cell.value).strip())
            data.append(row)

        # Find header row
        header_row_idx = _find_header_row(data)
        if header_row_idx is None:
            continue

        headers = [str(h) if h else '' for h in data[header_row_idx]]
        rows = data[header_row_idx + 1:]

        # Filter and pad rows
        valid_rows = []
        for row in rows:
            if not row or not row[0]:
                continue
            first_cell = str(row[0]) if row[0] else ''
            if any(kw in first_cell for kw in ['人', '不及格', '总计', '合计', '成绩信息']):
                continue
            # Pad to header length
            if len(row) < len(headers):
                row = list(row) + [None] * (len(headers) - len(row))
            valid_rows.append(row[:len(headers)])

        if valid_rows:
            df = pd.DataFrame(valid_rows, columns=headers)
            all_data.append(df)

    if not all_data:
        raise ValueError(f"No valid data found in {filepath}")

    return pd.concat(all_data, ignore_index=True)


def _read_xlsx_openpyxl(filepath: str) -> pd.DataFrame:
    """Read .xlsx using pandas/openpyxl, handling multi-sheet files."""
    xl = pd.ExcelFile(filepath, engine='openpyxl')

    all_data = []
    for sheet_name in xl.sheet_names:
        df = pd.read_excel(xl, sheet_name)
        if df.empty:
            continue

        # Check if pandas headers already contain valid column names
        headers = [str(h).strip() if h else '' for h in df.columns.tolist()]
        has_valid_headers = any(
            kw in str(headers[0]).lower() if headers else False
            for kw in ['学号', 'student', 'id', '序号']
        ) if headers else False

        data = df.values.tolist()

        # Only search for embedded header if pandas headers are NOT already valid
        if not has_valid_headers:
            header_row_idx = _find_header_row(data)
            if header_row_idx is not None and header_row_idx > 0:
                headers = [str(h) if h else '' for h in data[header_row_idx]]
                rows = data[header_row_idx + 1:]
                valid_rows = []
                for row in rows:
                    if not row or not row[0]:
                        continue
                    first_cell = str(row[0]) if row[0] else ''
                    if any(kw in first_cell for kw in ['人', '不及格', '总计', '合计']):
                        continue
                    if len(row) < len(headers):
                        row = list(row) + [None] * (len(headers) - len(row))
                    valid_rows.append(row[:len(headers)])
                if valid_rows:
                    df = pd.DataFrame(valid_rows, columns=headers)

        all_data.append(df)

    xl.close()

    if not all_data:
        raise ValueError(f"No valid data found in {filepath}")

    return pd.concat(all_data, ignore_index=True)


def _find_header_row(data: list) -> int:
    """Find the row index containing column headers.

    Headers typically contain '学号' and '姓名' in the first two columns.
    """
    for i, row in enumerate(data):
        if not row:
            continue
        col0 = str(row[0]).strip() if row[0] else ''
        col1 = str(row[1]).strip() if len(row) > 1 and row[1] else ''
        if '学号' in col0 and ('姓名' in col1 or '名称' in col1):
            return i
    # Fallback: return row 2 (common format: title row, empty row, header row)
    if len(data) >= 3:
        return 2
    return 0


def read_values_sheet(filepath: str) -> dict | None:
    """Read the hidden _values sheet from a module output file.

    The _values sheet stores pre-computed final scores keyed by student ID,
    so downstream modules don't need to evaluate Excel formulas.

    Args:
        filepath: Path to the .xlsx file

    Returns:
        Dict of {student_id: {name, class, score}} or None if not found.
    """
    import openpyxl
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    except Exception:
        return None

    if '_values' not in wb.sheetnames:
        wb.close()
        return None

    ws = wb['_values']
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        return None

    # First row is headers: 学号, 姓名, 班级, <score_column>
    headers = [str(h).strip() if h else '' for h in rows[0]]
    id_col = 0
    name_col = 1
    class_col = 2
    score_col = 3

    result = {}
    for row in rows[1:]:
        if not row or not row[id_col]:
            continue
        sid = str(row[id_col]).strip()
        # Clean float-format IDs (221352107.0 → 221352107)
        if sid.endswith('.0') and sid[:-2].isdigit():
            sid = sid[:-2]
        name = str(row[name_col]).strip() if len(row) > name_col and row[name_col] else ''
        cls = str(row[class_col]).strip() if len(row) > class_col and row[class_col] else ''
        try:
            score = float(row[score_col]) if len(row) > score_col and row[score_col] is not None else 0.0
        except (ValueError, TypeError):
            score = 0.0

        if sid:
            result[sid] = {'name': name, 'class': cls, 'score': score}

    return result if result else None


def read_xlsx_sheets(filepath: str, data_only: bool = False) -> dict:
    """Read all sheets from a .xlsx file.

    Args:
        filepath: Path to the .xlsx file
        data_only: If True, read cached formula values instead of formulas themselves

    Returns:
        Dict of {sheet_name: pd.DataFrame}
    """
    if data_only:
        # Use openpyxl directly with data_only=True for formula values
        import openpyxl
        wb = openpyxl.load_workbook(filepath, data_only=True)
        sheets = {}
        for sn in wb.sheetnames:
            ws = wb[sn]
            data = list(ws.values)
            if not data:
                sheets[sn] = pd.DataFrame()
            else:
                headers = data[0]
                rows = data[1:]
                df = pd.DataFrame(rows, columns=headers)
                sheets[sn] = df
        wb.close()
        return sheets
    else:
        xl = pd.ExcelFile(filepath, engine='openpyxl')
        sheets = {}
        for sheet_name in xl.sheet_names:
            df = pd.read_excel(xl, sheet_name)
            sheets[sheet_name] = df
        xl.close()
        return sheets
