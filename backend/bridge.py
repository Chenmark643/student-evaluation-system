"""
Bridge module - Central API surface for Eel frontend ↔ backend communication.

Every @eel.expose function here is callable from JavaScript via
    eel.function_name(args)(callback)

All long-running operations spawn threads to keep the Eel event loop responsive,
and report progress via eel.updateProgress().
"""

import os
import sys
import re
import json
import zipfile
import base64
import tempfile
import shutil
import subprocess
import threading
import uuid
import tkinter as tk
from tkinter import filedialog
from collections import defaultdict

import pandas as pd
import eel

# ── cross-platform helpers ───────────────────────────────────────────

_IS_MAC = sys.platform == 'darwin'
_IS_WIN = sys.platform == 'win32'


def _open_path_in_os(path: str):
    """Open a file or folder in the OS-native file manager / default app.

    Windows → Explorer (or associated app)
    macOS   → Finder (or associated app via `open`)
    Linux   → xdg-open
    """
    if _IS_MAC:
        subprocess.run(['open', path])
    elif _IS_WIN:
        os.startfile(path)  # type: ignore[attr-defined]
    else:
        subprocess.run(['xdg-open', path])

from backend.module_a_gpa import process_gpa
from backend.module_b_moral import process_moral_education, preview_file_sheets
from backend.moral_vnext import list_moral_students as list_moral_students_impl, process_moral_fresh, process_moral_vnext
from backend.moral_templates import (
    analyze_moral_project_templates as analyze_moral_project_templates_impl,
    copy_moral_project_templates as copy_moral_project_templates_impl,
    list_moral_project_templates as list_moral_project_templates_impl,
)
from backend.moral_cloud import prepare_moral_cloud_bundle as prepare_moral_cloud_bundle_impl
from backend.module_c_quality import (
    load_activity_mappings, record_activity, get_activity_suggestion,
    get_categories, get_grades_for_category, get_thresholds,
    get_thresholds_dict,
    calculate_quality_scores, export_quality_scores,
    add_activity_mapping, delete_activity_mapping,
    add_custom_threshold, remove_custom_threshold,
    add_custom_category, remove_custom_category,
)
from backend.module_d_comprehensive import process_comprehensive
from backend.utils.progress_reporter import ProgressReporter
from backend.import_studio import (
    analyze_import_file as analyze_import_file_impl,
    list_import_templates as list_import_templates_impl,
    save_import_template as save_import_template_impl,
)
from backend.quality_presets import (
    build_official_presets, calculate_activity_score, validate_manual_score,
)
from backend.kdocs_sync import (
    auth_status as kdocs_auth_status_impl,
    bind_workbook as kdocs_bind_workbook_impl,
    get_binding as kdocs_get_binding_impl,
    get_sync_overview as kdocs_get_sync_overview_impl,
    login as kdocs_login_impl,
    logout as kdocs_logout_impl,
    reorder_bound_workbook as kdocs_reorder_workbook_impl,
    sync_workbook as kdocs_sync_workbook_impl,
)


# ============================================================
# File dialog helpers for legacy Eel/dev mode.
# The desktop build overrides these through backend.api and uses pywebview's
# window-owned dialogs so they cannot appear behind the application.
# ============================================================

def _diag(msg):
    try:
        with open(os.path.join(tempfile.gettempdir(), 'app_diag.log'), 'a', encoding='utf-8') as f:
            import datetime as _dt
            f.write(f'[{_dt.datetime.now():%H:%M:%S}] {msg}\n')
    except Exception:
        pass

@eel.expose
def ping_diag():
    """Diagnostic: verify API bridge works."""
    _diag('ping_diag called — API bridge OK')
    return 'pong'


# ============================================================
# Kdocs cloud workbook integration
# ============================================================

_KDOCS_JOBS = {}
_KDOCS_JOBS_LOCK = threading.Lock()


def _update_kdocs_job(job_id: str, **updates) -> None:
    with _KDOCS_JOBS_LOCK:
        job = _KDOCS_JOBS.get(job_id)
        if job is not None:
            job.update(updates)

@eel.expose
def kdocs_auth_status() -> dict:
    """Return connection state without exposing credentials."""
    return kdocs_auth_status_impl()


@eel.expose
def kdocs_login() -> dict:
    """Open the browser OAuth flow and persist credentials in the OS keychain."""
    return kdocs_login_impl()


@eel.expose
def kdocs_logout() -> dict:
    """Disconnect the local WPS account by removing its saved credential."""
    return kdocs_logout_impl()


@eel.expose
def kdocs_get_binding(cloud_key: str) -> dict:
    """Return the saved cloud workbook link for one logical output."""
    return kdocs_get_binding_impl(cloud_key)


@eel.expose
def kdocs_get_sync_overview(cloud_key: str, major: str = "") -> dict:
    """Return current visible cloud sheets and last synchronization state."""
    return kdocs_get_sync_overview_impl(cloud_key, major)


@eel.expose
def kdocs_bind_workbook(cloud_key: str, link_url: str) -> dict:
    """Bind another operator to the existing college-wide cloud workbook."""
    return kdocs_bind_workbook_impl(cloud_key, link_url)


@eel.expose
def kdocs_sync_workbook(local_path: str, cloud_key: str) -> dict:
    """Publish a generated workbook or update its existing Kdocs counterpart."""
    return kdocs_sync_workbook_impl(local_path, cloud_key)


@eel.expose
def prepare_moral_cloud_bundle(local_paths: list) -> dict:
    """Combine mixed A/B moral outputs into visible per-class sheets for cloud sync."""
    try:
        return prepare_moral_cloud_bundle_impl(local_paths or [])
    except Exception as exc:
        return {"success": False, "error": str(exc)}


@eel.expose
def kdocs_start_sync_workbook(local_path: str, cloud_key: str, force_create: bool = False) -> dict:
    """Start synchronization in a background thread and return a progress id."""
    job_id = uuid.uuid4().hex
    with _KDOCS_JOBS_LOCK:
        finished = [key for key, value in _KDOCS_JOBS.items() if value.get("done")]
        for old_id in finished[:-20]:
            _KDOCS_JOBS.pop(old_id, None)
        _KDOCS_JOBS[job_id] = {
            "success": True,
            "job_id": job_id,
            "done": False,
            "status": "running",
            "percent": 0,
            "stage": "正在排队",
            "detail": "准备连接金山文档",
            "current_sheet": "",
            "sheet_index": 0,
            "sheet_total": 0,
        }

    def report(percent, stage, detail="", **extra):
        allowed = {key: extra[key] for key in ("current_sheet", "sheet_index", "sheet_total") if key in extra}
        _update_kdocs_job(
            job_id,
            percent=max(0, min(100, int(percent))),
            stage=str(stage or "正在同步"),
            detail=str(detail or ""),
            **allowed,
        )

    def worker():
        try:
            result = kdocs_sync_workbook_impl(
                local_path,
                cloud_key,
                progress_callback=report,
                force_create=bool(force_create),
            )
        except Exception as exc:
            result = {"success": False, "error": str(exc)}
        ok = bool(result.get("success"))
        _update_kdocs_job(
            job_id,
            done=True,
            status="success" if ok else "error",
            percent=100,
            stage="同步完成" if ok else "同步失败",
            detail=(result.get("name") or "云表已更新") if ok else result.get("error", "同步未完成"),
            result=result,
        )

    threading.Thread(target=worker, name=f"kdocs-sync-{job_id[:8]}", daemon=True).start()
    return {"success": True, "job_id": job_id}


@eel.expose
def kdocs_get_sync_progress(job_id: str) -> dict:
    """Read one background synchronization snapshot."""
    with _KDOCS_JOBS_LOCK:
        job = _KDOCS_JOBS.get(str(job_id))
        if job is None:
            return {"success": False, "error": "找不到该同步任务，请重新开始。"}
        return dict(job)


@eel.expose
def kdocs_start_reorder_workbook(cloud_key: str) -> dict:
    """Start deterministic worksheet ordering in a background job."""
    job_id = uuid.uuid4().hex
    with _KDOCS_JOBS_LOCK:
        _KDOCS_JOBS[job_id] = {
            "success": True,
            "job_id": job_id,
            "done": False,
            "status": "running",
            "percent": 0,
            "stage": "正在排队",
            "detail": "准备整理学院云表",
            "current_sheet": "",
            "sheet_index": 0,
            "sheet_total": 0,
        }

    def report(percent, stage, detail="", **extra):
        allowed = {key: extra[key] for key in ("current_sheet", "sheet_index", "sheet_total") if key in extra}
        _update_kdocs_job(
            job_id,
            percent=max(0, min(100, int(percent))),
            stage=str(stage or "正在整理"),
            detail=str(detail or ""),
            **allowed,
        )

    def worker():
        try:
            result = kdocs_reorder_workbook_impl(cloud_key, progress_callback=report)
        except Exception as exc:
            result = {"success": False, "error": str(exc)}
        ok = bool(result.get("success"))
        _update_kdocs_job(
            job_id,
            done=True,
            status="success" if ok else "error",
            percent=100,
            stage="顺序整理完成" if ok else "顺序整理失败",
            detail=(f"已检查 {result.get('sheet_count', 0)} 张工作表" if ok else result.get("error", "整理未完成")),
            result=result,
        )

    threading.Thread(target=worker, name=f"kdocs-reorder-{job_id[:8]}", daemon=True).start()
    return {"success": True, "job_id": job_id}


@eel.expose
def open_web_link(url: str) -> dict:
    """Open a validated HTTP(S) link in the user's default browser."""
    if not isinstance(url, str) or not re.match(r'^https?://', url, re.I):
        return {'success': False, 'error': '链接格式无效'}
    try:
        _open_path_in_os(url)
        return {'success': True}
    except Exception as exc:
        return {'success': False, 'error': str(exc)}


def _legacy_dialog(dialog, *, title, file_types=None):
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    try:
        return dialog(title=title, filetypes=file_types or [('所有文件', '*.*')])
    finally:
        root.destroy()

@eel.expose
def select_file(file_types: list = None, title: str = '选择文件') -> str:
    _diag(f'select_file called: title={title}')
    return _legacy_dialog(filedialog.askopenfilename, title=title, file_types=file_types) or ''

@eel.expose
def select_directory(title: str = '选择输出目录') -> str:
    _diag(f'select_directory called')
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    try:
        return filedialog.askdirectory(title=title) or ''
    finally:
        root.destroy()

@eel.expose
def select_files(file_types: list = None, title: str = '选择文件') -> list:
    _diag(f'select_files called')
    paths = _legacy_dialog(filedialog.askopenfilenames, title=title, file_types=file_types)
    return list(paths or [])


@eel.expose
def list_moral_project_templates() -> list:
    """Return the fixed one-project template catalogue."""
    return list_moral_project_templates_impl()


@eel.expose
def analyze_moral_project_templates(paths: list) -> dict:
    """Validate multiple standard project templates without changing them."""
    try:
        return analyze_moral_project_templates_impl(paths or [])
    except Exception as exc:
        return {"success": False, "error": str(exc), "files": []}


@eel.expose
def copy_moral_project_templates(project_key: str, output_dir: str) -> dict:
    """Copy one or all bundled project templates to a selected folder."""
    try:
        return copy_moral_project_templates_impl(project_key or "", output_dir or "")
    except Exception as exc:
        return {"success": False, "error": str(exc)}


# ============================================================
# Module A: GPA Calculation
# ============================================================

@eel.expose
def run_module_a(input_path: str, output_dir: str,
                 column_mappings: dict = None, major_filter: str = '') -> dict:
    """Run GPA calculation (Module A) — single file.

    Args:
        input_path: Path to raw .xls grade file
        output_dir: Directory for output files

    Returns:
        Result dict with output paths and statistics.
    """
    try:
        result = process_gpa(input_path, output_dir, column_mappings=column_mappings,
                             major_filter=major_filter)
        return result
    except Exception as e:
        return {'success': False, 'error': str(e)}


@eel.expose
def run_module_a_batch(input_paths: list, output_dir: str,
                       column_mappings: dict = None, major_filter: str = '') -> dict:
    """Run GPA calculation (Module A) — batch mode.

    Args:
        input_paths: List of paths to raw .xls/.xlsx grade files
        output_dir: Directory for output files

    Returns:
        Result dict with output paths and statistics.
    """
    try:
        from backend.module_a_gpa import process_gpa_batch
        result = process_gpa_batch(input_paths, output_dir,
                                   column_mappings=column_mappings or {},
                                   major_filter=major_filter)
        return result
    except Exception as e:
        return {'success': False, 'error': str(e)}


@eel.expose
def run_module_a_async(input_path: str, output_dir: str):
    """Run GPA calculation in a background thread with progress updates."""
    def _run():
        try:
            process_gpa(input_path, output_dir)
        except Exception as e:
            eel.onModuleError('A', str(e))()

    thread = threading.Thread(target=_run, daemon=True)
    thread.start()


# ============================================================
# Module B: Moral Education
# ============================================================

@eel.expose
def preview_moral_file(filepath: str) -> dict:
    """Preview Excel file sheets and headers for column mapping.

    Returns:
        {sheet_name: {headers: [str], sample_rows: [[str]]}}
    """
    try:
        return preview_file_sheets(filepath)
    except Exception as e:
        return {'error': str(e)}


@eel.expose
def analyze_import_file(filepath: str, module_type: str) -> dict:
    return analyze_import_file_impl(filepath, module_type)


@eel.expose
def analyze_gpa_course_structure(filepath: str, sheet_mapping: dict = None) -> dict:
    try:
        from backend.gpa_course_audit import analyze_gpa_course_structure as analyze
        return analyze(filepath, sheet_mapping or {})
    except Exception as exc:
        return {'success': False, 'error': str(exc)}


@eel.expose
def audit_toolbox_applicants(config: dict) -> dict:
    try:
        from backend.toolbox_audit import audit_applicants
        return audit_applicants(config or {})
    except Exception as exc:
        return {'success': False, 'error': str(exc)}

@eel.expose
def prepare_award_roster(file_path: str) -> dict:
    try:
        from backend.award_eligibility import prepare_roster
        return prepare_roster(file_path)
    except Exception as exc:
        return {'success':False,'error':str(exc)}

@eel.expose
def audit_award_candidates(config: dict) -> dict:
    try:
        from backend.award_eligibility import audit_candidates
        return audit_candidates(config)
    except Exception as exc:
        return {'success':False,'error':str(exc)}


@eel.expose
def export_toolbox_audit(students: list, output_path: str) -> dict:
    try:
        from backend.toolbox_audit import export_audit_report
        return export_audit_report(students or [], output_path)
    except Exception as exc:
        return {'success': False, 'error': str(exc)}


@eel.expose
def save_import_template(name: str, module_type: str,
                         fingerprint: str, mappings: dict) -> dict:
    try:
        record = save_import_template_impl(name, module_type, fingerprint, mappings or {})
        return {'success': True, 'template': record}
    except Exception as exc:
        return {'success': False, 'error': str(exc)}


@eel.expose
def list_import_templates(module_type: str = '') -> list:
    return list_import_templates_impl(module_type)


@eel.expose
def run_module_b(roster_path: str,
                 absence_files: list, class_absence_files: list,
                 dormitory_files: list, classroom_files: list,
                 org_class_files: list,
                 review_scores: dict,
                 output_dir: str,
                 column_mappings: dict = None,
                 manual_scores: dict = None,
                 selected_columns: list = None,
                 grade_filter: str = 'all', major_filter: str = '') -> dict:
    """Run moral education calculation (Module B).

    Args:
        roster_path: Path to GPA output (学分绩点.xlsx) for roster
        absence_files: List of 早晚自习 absence file paths
        class_absence_files: List of 课堂出勤 absence file paths
        dormitory_files: List of 宿舍卫生 file paths
        classroom_files: List of 教室卫生 file paths
        org_class_files: List of 团课出勤 file paths
        review_scores: Dict of {student_id: 评议分}
        output_dir: Output directory
        column_mappings: Interactive column mappings
        manual_scores: Dict of {student_id: {field: value}} for manual fields
        selected_columns: List of column headers to include in output
        grade_filter: Grade filter ('all', '24级', etc.)

    Returns:
        Result dict.
    """
    try:
        result = process_moral_education(
            roster_path=roster_path,
            absence_files=[f for f in (absence_files or []) if f],
            class_absence_files=[f for f in (class_absence_files or []) if f],
            dormitory_files=[f for f in (dormitory_files or []) if f],
            classroom_files=[f for f in (classroom_files or []) if f],
            org_class_files=[f for f in (org_class_files or []) if f],
            review_scores=review_scores or {},
            output_dir=output_dir,
            column_mappings=column_mappings or {},
            manual_scores=manual_scores or {},
            selected_columns=selected_columns,
            grade_filter=grade_filter,
            major_filter=major_filter,
        )
        return result
    except Exception as e:
        return {'success': False, 'error': str(e)}


@eel.expose
def run_moral_vnext(config: dict) -> dict:
    """Continue a partially completed moral workbook with configurable items."""
    try:
        return process_moral_vnext(config or {})
    except Exception as exc:
        return {'success': False, 'error': str(exc)}


@eel.expose
def run_moral_fresh(config: dict) -> dict:
    """Create moral scores from a roster with configurable add/deduct items."""
    try:
        return process_moral_fresh(config or {})
    except Exception as exc:
        return {'success': False, 'error': str(exc)}


@eel.expose
def list_moral_students(config: dict) -> dict:
    """List students for project-level batch score entry."""
    try:
        return list_moral_students_impl(config or {})
    except Exception as exc:
        return {'success': False, 'error': str(exc)}


# ============================================================
# Module C: Quality Development
# ============================================================

@eel.expose
def get_quality_categories() -> list:
    """Get list of quality development categories."""
    return get_categories()


@eel.expose
def get_quality_grades(category: str) -> list:
    """Get available grade options for a category."""
    return get_grades_for_category(category)


@eel.expose
def get_quality_thresholds() -> dict:
    """Get default threshold settings."""
    return get_thresholds()


@eel.expose
def get_official_quality_presets() -> list:
    """Return the immutable rule catalog used by the scoring drawer."""
    return build_official_presets()


@eel.expose
def preview_quality_activity_score(base_score: float, count: int = 1,
                                   contribution: float = 1.0,
                                   related: bool = False) -> dict:
    return calculate_activity_score(base_score, count, contribution, related)


@eel.expose
def validate_quality_manual_score(value: float, score_range=None) -> dict:
    return validate_manual_score(value, score_range)


@eel.expose
def get_activity_suggestions(activity_name: str) -> dict:
    """Get auto-fill suggestion for an activity name.

    Returns:
        Dict with category, grade, score if found, else empty dict.
    """
    result = get_activity_suggestion(activity_name)
    return result if result else {}


@eel.expose
def save_activity_mapping(activity_name: str, category: str,
                          grade: str, score: float):
    """Save an activity mapping for future auto-fill."""
    record_activity(activity_name, category, grade, score)


@eel.expose
def save_all_activity_mappings(mappings: dict):
    """Save all activity mappings (full replace)."""
    from backend.module_c_quality import save_activity_mappings
    save_activity_mappings(mappings)


@eel.expose
def delete_activity_mapping(activity_name: str):
    """Delete a single activity mapping."""
    return delete_activity_mapping(activity_name)


@eel.expose
def add_new_activity_mapping(name: str, category: str, grade: str, score: float) -> dict:
    """Add a new activity mapping from scratch."""
    return add_activity_mapping(name, category, grade, score)


@eel.expose
def add_custom_threshold_category(name: str, max_score: float, categories: list,
                                    mode: str = 'sum') -> list:
    """Add a custom threshold linked to specific bonus categories.

    Args:
        mode: 'sum' (求和封顶) or 'max_item' (取最高分)
    """
    return add_custom_threshold(name, max_score, categories, mode)


@eel.expose
def remove_custom_threshold_category(name: str) -> list:
    """Remove a custom threshold by name."""
    return remove_custom_threshold(name)


@eel.expose
def get_all_thresholds() -> list:
    """Get all thresholds including custom ones (list format)."""
    return get_thresholds()


@eel.expose
def add_quality_category(name: str) -> list:
    """Add a custom quality category."""
    return add_custom_category(name)


@eel.expose
def remove_quality_category(name: str) -> list:
    """Remove a custom quality category."""
    return remove_custom_category(name)


@eel.expose
def load_activity_mappings_json() -> dict:
    """Load all saved activity mappings."""
    return load_activity_mappings()


@eel.expose
def read_roster_for_quality(roster_path: str) -> dict:
    """Read student roster from GPA file for quality module.

    Returns: {student_id: {name, class}}
    """
    try:
        from backend.module_b_moral import _load_roster
        return _load_roster(roster_path)
    except Exception as e:
        return {}


@eel.expose
def export_quality_with_roster(roster: dict, quality_data: dict,
                                output_path: str, thresholds=None,
                                major_filter: str = '') -> dict:
    """Export quality scores with merged cells and formulas."""
    try:
        from backend.module_c_quality import export_quality_merged
        from backend.utils.class_utils import class_matches_program
        if major_filter:
            roster = {sid: info for sid, info in (roster or {}).items()
                      if class_matches_program(info.get('class', info.get('班级', '')), major_filter)}
            quality_data = {sid: value for sid, value in (quality_data or {}).items()
                            if sid in roster}
        # Normalize thresholds: if dict from JS, convert to list format
        if isinstance(thresholds, dict):
            th_list = []
            for name, val in thresholds.items():
                if isinstance(val, dict):
                    th_list.append({'name': name, 'max': val.get('max', 0),
                                    'categories': val.get('categories', [name]),
                                    'mode': val.get('mode', 'sum')})
                else:
                    th_list.append({'name': name, 'max': float(val),
                                    'categories': [name]})
            thresholds = th_list
        return export_quality_merged(roster, quality_data, output_path, thresholds)
    except Exception as e:
        return {'success': False, 'error': str(e)}


@eel.expose
def export_quality_scores_json(class_data: dict, output_path: str,
                                thresholds=None, major_filter: str = '') -> dict:
    """Export quality development scores."""
    try:
        from backend.utils.class_utils import class_matches_program
        if major_filter:
            class_data = {name: value for name, value in (class_data or {}).items()
                          if class_matches_program(name, major_filter)}
        result = export_quality_scores(class_data, output_path, thresholds)
        return result
    except Exception as e:
        return {'success': False, 'error': str(e)}


# ============================================================
# Module D: Comprehensive Evaluation
# ============================================================

@eel.expose
def run_module_d(gpa_path: str, moral_path: str,
                 quality_path: str, output_dir: str,
                 has_sports: bool = False,
                 sports_programs: list = None,
                 column_mappings: dict = None,
                 grade_filter: str = 'all', major_filter: str = '') -> dict:
    """Run comprehensive evaluation (Module D).

    Args:
        gpa_path: Path to Module A output (学分绩点.xlsx)
        moral_path: Path to Module B output (德育分.xlsx)
        quality_path: Path to Module C output (素拓.xlsx)
        output_dir: Output directory
        has_sports: Whether to use sports formula
        sports_programs: List of program-grade keys with sports
        column_mappings: Dict of {filepath: {sheet: {id_col, name_col, class_col, score_col}}}
        grade_filter: 'all' or 'XX级' to filter output

    Returns:
        Result dict.
    """
    try:
        result = process_comprehensive(
            gpa_path=gpa_path,
            moral_path=moral_path,
            quality_path=quality_path,
            output_dir=output_dir,
            has_sports=has_sports,
            sports_programs=sports_programs or [],
            column_mappings=column_mappings or {},
            grade_filter=grade_filter,
            major_filter=major_filter,
        )
        return result
    except Exception as e:
        return {'success': False, 'error': str(e)}


# ============================================================
# V3.0: Semester Comparison
# ============================================================
@eel.expose
def compare_semesters(file1: str, file2: str) -> dict:
    """Compare two semester comprehensive evaluation files."""
    try:
        from backend.parsers.xls_reader import read_xlsx_sheets
        sheets1 = read_xlsx_sheets(file1)
        sheets2 = read_xlsx_sheets(file2)

        students1 = {}
        students2 = {}

        for sn, df in sheets1.items():
            if df.empty or len(df.columns) < 3: continue
            ncols = len(df.columns)
            for idx, row in df.iterrows():
                sid = str(row.iloc[0]).strip() if not pd.isna(row.iloc[0]) else ''
                if not sid or len(sid) < 6: continue
                try:
                    score = float(row.iloc[ncols - 1]) if not pd.isna(row.iloc[ncols - 1]) else 0.0
                except (ValueError, TypeError):
                    score = 0.0
                name = str(row.iloc[1]).strip() if len(df.columns) > 1 and not pd.isna(row.iloc[1]) else ''
                students1[sid] = {'name': name, 'score': round(score, 2)}

        for sn, df in sheets2.items():
            if df.empty or len(df.columns) < 3: continue
            ncols = len(df.columns)
            for idx, row in df.iterrows():
                sid = str(row.iloc[0]).strip() if not pd.isna(row.iloc[0]) else ''
                if not sid or len(sid) < 6: continue
                try:
                    score = float(row.iloc[ncols - 1]) if not pd.isna(row.iloc[ncols - 1]) else 0.0
                except (ValueError, TypeError):
                    score = 0.0
                name = str(row.iloc[1]).strip() if len(df.columns) > 1 and not pd.isna(row.iloc[1]) else ''
                students2[sid] = {'name': name, 'score': round(score, 2)}

        # Rank students in each file
        sorted1 = sorted(students1.items(), key=lambda x: x[1]['score'], reverse=True)
        sorted2 = sorted(students2.items(), key=lambda x: x[1]['score'], reverse=True)
        rank1 = {sid: i+1 for i, (sid, _) in enumerate(sorted1)}
        rank2 = {sid: i+1 for i, (sid, _) in enumerate(sorted2)}

        all_sids = set(list(students1.keys()) + list(students2.keys()))
        data = []
        for sid in all_sids:
            s1 = students1.get(sid, {})
            s2 = students2.get(sid, {})
            data.append({
                'id': sid,
                'name': s2.get('name', s1.get('name', '')),
                'previous': s1.get('score', 0),
                'current': s2.get('score', 0),
                'rank_prev': rank1.get(sid, 0),
                'rank_curr': rank2.get(sid, 0),
            })

        return {'success': True, 'data': data}
    except Exception as e:
        return {'success': False, 'error': str(e)}


# ============================================================
# V3.0: Preview Export — write edited data to Excel
# ============================================================
@eel.expose
def export_preview_data(headers: list, rows: list, output_path: str) -> dict:
    """Export preview-edited data to Excel file."""
    from backend.utils.excel_writer import unique_path as _up
    output_path = _up(output_path)
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '预览导出'

        header_font = Font(name='SimSun', size=10, bold=True)
        data_font = Font(name='SimSun', size=10)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal='center', vertical='center')

        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border

        for ri, row in enumerate(rows):
            for ci, val in enumerate(row):
                cell = ws.cell(row=ri + 2, column=ci + 1, value=val if val != '' else None)
                cell.font = data_font
                cell.alignment = center_align
                cell.border = thin_border
                if ci == 0:  # Student ID column → text format
                    cell.number_format = '@'

        for ci in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(ci)].width = 14

        os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else '.', exist_ok=True)
        wb.save(output_path)
        wb.close()
        return {'success': True, 'output': output_path}
    except Exception as e:
        return {'success': False, 'error': str(e)}


# ============================================================
# V6.0: One-click Export All
# ============================================================
@eel.expose
def export_all_modules(gpa_path: str, moral_path: str, quality_path: str,
                       output_dir: str, has_sports: bool = False) -> dict:
    """Run all four modules in sequence: GPA → Moral → Quality → Comprehensive.
    Returns summary with all output paths."""
    results = {}
    try:
        # Module A: GPA
        from backend.module_a_gpa import process_gpa_batch
        gpa_files = [gpa_path] if gpa_path and os.path.exists(gpa_path) else []
        if gpa_files:
            gpa_result = process_gpa_batch(gpa_files, output_dir)
            results['gpa'] = gpa_result

        # Module B: Moral (uses GPA output as roster — use actual path from Module A)
        gpa_output = results.get('gpa', {}).get('output1', os.path.join(output_dir, '学分绩点.xlsx'))
        if os.path.exists(gpa_output):
            moral_result = process_moral_education(
                roster_path=gpa_output,
                absence_files=[], class_absence_files=[], dormitory_files=[],
                classroom_files=[], org_class_files=[],
                review_scores={}, output_dir=output_dir, column_mappings={},
                manual_scores={}, selected_columns=None, grade_filter='all')
            results['moral'] = moral_result

        # Module C: Quality
        quality_output = os.path.join(output_dir, '素拓分.xlsx')
        if quality_path and os.path.exists(quality_path):
            results['quality'] = {'success': True, 'output': quality_path, 'note': '手动处理模块'}

        # Module D: Comprehensive — use actual paths from Module A and B
        moral_output = results.get('moral', {}).get('output', os.path.join(output_dir, '德育分.xlsx'))
        quality_file = quality_path if quality_path and os.path.exists(quality_path) else quality_output.replace('素拓分', '素拓分')
        if os.path.exists(gpa_output) and os.path.exists(moral_output):
            comp_result = process_comprehensive(
                gpa_path=gpa_output, moral_path=moral_output,
                quality_path=quality_file if os.path.exists(quality_file) else quality_output,
                output_dir=output_dir, has_sports=has_sports)
            results['comprehensive'] = comp_result

        return {'success': True, 'results': results,
                'summary': f'完成：GPA→德育→素拓→综测，输出到 {output_dir}'}
    except Exception as e:
        return {'success': False, 'error': str(e)}
@eel.expose
def load_counselor_data(filepath: str) -> dict:
    """Load comprehensive evaluation data from a specific file.
    Detects columns by header keywords: 绩点, 德育, 素拓/拓展, 综测."""
    try:
        from backend.parsers.xls_reader import read_xlsx_sheets

        if not filepath or not os.path.exists(filepath):
            return {'success': False, 'error': '文件不存在'}

        sheets = read_xlsx_sheets(filepath)
        data = []

        for sn, df in sheets.items():
            if df.empty or len(df.columns) < 3:
                continue
            ncols = len(df.columns)
            headers = [str(c).strip() for c in df.columns]

            for idx, row in df.iterrows():
                sid = str(row.iloc[0]).strip() if not pd.isna(row.iloc[0]) else ''
                if not sid or len(sid) < 6:
                    continue
                name = str(row.iloc[1]).strip() if len(df.columns) > 1 and not pd.isna(row.iloc[1]) else ''

                result = {'id': sid, 'name': name, 'class': str(sn)}
                for ci, h in enumerate(headers):
                    try:
                        val = float(row.iloc[ci]) if not pd.isna(row.iloc[ci]) else None
                    except (ValueError, TypeError):
                        val = None
                    if val is None:
                        continue
                    hl = h.lower()
                    if '绩点' in hl or 'gpa' in hl:
                        result['gpa'] = round(val, 4)
                    elif '德育' in hl or 'moral' in hl:
                        result['moral'] = round(val, 1)
                    elif '素拓' in hl or '拓展' in hl or 'quality' in hl:
                        result['quality'] = round(val, 1)
                    elif '综测' in hl or '综合' in hl or 'comprehensive' in hl:
                        result['comp'] = round(val, 2)

                data.append(result)

        return {'success': True, 'data': data, 'file': os.path.basename(filepath)}
    except Exception as e:
        return {'success': False, 'error': str(e)}


@eel.expose
def compare_with_current(prev_file: str, current_data: list) -> dict:
    """Compare previous semester file with current in-memory data."""
    try:
        from backend.parsers.xls_reader import read_xlsx_sheets

        sheets = read_xlsx_sheets(prev_file)
        prev_data = {}
        for sn, df in sheets.items():
            if df.empty or len(df.columns) < 3:
                continue
            ncols = len(df.columns)
            for idx, row in df.iterrows():
                sid = str(row.iloc[0]).strip() if not pd.isna(row.iloc[0]) else ''
                if not sid or len(sid) < 6:
                    continue
                name = str(row.iloc[1]).strip() if len(df.columns) > 1 and not pd.isna(row.iloc[1]) else ''
                try:
                    score = float(row.iloc[ncols - 1]) if not pd.isna(row.iloc[ncols - 1]) else 0.0
                except (ValueError, TypeError):
                    score = 0.0
                prev_data[sid] = {'name': name, 'score': round(score, 2)}

        data = []
        for cur in (current_data or []):
            sid = cur.get('id', '')
            prev = prev_data.get(sid, {})
            data.append({
                'id': sid,
                'name': cur.get('name', ''),
                'previous': prev.get('score', 0),
                'current': cur.get('comp', 0),
            })
        return {'success': True, 'data': data}
    except Exception as e:
        return {'success': False, 'error': str(e)}


# ============================================================
# Utility
# ============================================================

@eel.expose
def get_app_version() -> str:
    """Get application version."""
    from config import APP_VERSION
    return APP_VERSION


# ============================================================
# V8.2: Local File Update System
# ============================================================

@eel.expose
def verify_new_exe(filepath: str) -> dict:
    """Verify that a file looks like a valid new version exe.

    Returns:
        {valid, version, error}
    """
    try:
        if not filepath or not os.path.exists(filepath):
            return {'valid': False, 'error': '文件不存在'}
        if not filepath.lower().endswith('.exe'):
            return {'valid': False, 'error': '不是exe文件'}
        # Check file size (must be > 50MB to be valid)
        size = os.path.getsize(filepath)
        if size < 50 * 1024 * 1024:
            return {'valid': False, 'error': f'文件太小({size/1024/1024:.0f}MB)，可能不完整'}
        return {'valid': True, 'size': size, 'filename': os.path.basename(filepath)}
    except Exception as e:
        return {'valid': False, 'error': str(e)}


@eel.expose
def install_local_update(new_exe_path: str) -> bool:
    """Install update from a local exe file.
    1. Write updater script next to current exe
    2. Launch updater (waits for app to close, then replaces)
    3. App exits

    Returns True if updater launched successfully.
    """
    try:
        import sys as _sys

        if getattr(_sys, 'frozen', False):
            current_path = _sys.executable
        else:
            current_path = _sys.executable

        current_dir = os.path.dirname(current_path)
        old_name = os.path.basename(current_path)

        updater_path = os.path.join(current_dir, '_updater.bat')
        with open(updater_path, 'w', encoding='gbk') as f:
            f.write(f'''@echo off
chcp 65001 >nul
title 更新学生综合测评系统...
echo.
echo   正在更新学生综合测评系统...
echo   请勿关闭此窗口
echo.
:wait
timeout /t 2 /nobreak >nul
tasklist /FI "IMAGENAME eq {old_name}" 2>NUL | find /I "{old_name}" >NUL
if %errorlevel% == 0 goto wait
echo   正在替换...
move /Y "{new_exe_path}" "{current_path}"
if exist "{current_path}" (
    echo   更新完成！即将启动...
    start "" "{current_path}"
) else (
    echo   替换失败！请手动操作
    pause
)
del "%~f0" 2>nul
''')

        subprocess.Popen(
            ['cmd', '/c', 'start', '更新', updater_path],
            shell=True, creationflags=subprocess.CREATE_NEW_CONSOLE
        )
        return True
    except Exception as e:
        import traceback
        traceback.print_exc()
        return False


@eel.expose
def open_file_explorer(path: str):
    """Open the file explorer at the given path (cross-platform)."""
    if os.path.exists(path):
        _open_path_in_os(path)
    elif os.path.exists(os.path.dirname(path)):
        _open_path_in_os(os.path.dirname(path))


# ============================================================
# V8.1: Student Course Scores Lookup
# ============================================================

@eel.expose
def get_student_course_scores(filepath: str, student_id: str) -> dict:
    """Get individual course scores for a specific student from a raw grade file.

    Args:
        filepath: Path to raw grade .xlsx file
        student_id: Student ID to look up

    Returns:
        {success, student: {id, name, class, course_scores: {course: score}, failed_courses: [...]}}
    """
    try:
        from backend.course_analyzer import _extract_students_from_xlsx
        students = _extract_students_from_xlsx(filepath)
        for s in students:
            if s['id'] == student_id:
                return {'success': True, 'student': s}
        return {'success': False, 'error': f'未找到学生 {student_id}'}
    except Exception as e:
        return {'success': False, 'error': str(e)}


@eel.expose
def get_student_report_card(student_data: dict, course_data: dict = None,
                             semester: str = '', class_avg: dict = None,
                             trend_data: list = None, counselor_notes: str = '') -> str:
    """Generate a formatted student report card Excel file.

    Returns path to generated .xlsx file.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '个人成绩报告'

        # Styles
        title_font = Font(name='SimSun', size=16, bold=True)
        header_font = Font(name='SimSun', size=11, bold=True)
        data_font = Font(name='SimSun', size=10)
        red_font = Font(name='SimSun', size=10, color='FF0000')
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        center = Alignment(horizontal='center', vertical='center')
        red_fill = PatternFill(start_color='FFE0E0', end_color='FFE0E0', fill_type='solid')

        row = 1

        # Title
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws.cell(row=row, column=1, value='学生个人成绩报告').font = title_font
        ws.cell(row=row, column=1).alignment = center
        row += 2

        # Student info
        for label, val in [('姓名', student_data.get('name', '')), ('学号', student_data.get('id', '')),
                           ('班级', student_data.get('class', '')), ('学期', semester)]:
            ws.cell(row=row, column=1, value=label).font = header_font
            ws.cell(row=row, column=2, value=str(val)).font = data_font
            row += 1
        row += 1

        # Scores section
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws.cell(row=row, column=1, value='成绩概览').font = header_font
        row += 1
        for label, val in [('学分绩点', student_data.get('gpa', '—')), ('德育分', student_data.get('moral', '—')),
                           ('素拓分', student_data.get('quality', '—')), ('综测成绩', student_data.get('comp', '—')),
                           ('班级排名', str(student_data.get('rank', '—')))]:
            ws.cell(row=row, column=1, value=label).font = data_font
            ws.cell(row=row, column=2, value=str(val)).font = data_font
            row += 1
        row += 1

        # Class comparison
        if class_avg:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            ws.cell(row=row, column=1, value='与班级平均对比').font = header_font
            row += 1
            for label, val in [('绩点(班级平均)', f"{student_data.get('gpa','—')} / {class_avg.get('gpa','—')}"),
                               ('德育(班级平均)', f"{student_data.get('moral','—')} / {class_avg.get('moral','—')}"),
                               ('综测(班级平均)', f"{student_data.get('comp','—')} / {class_avg.get('comp','—')}")]:
                ws.cell(row=row, column=1, value=label).font = data_font
                ws.cell(row=row, column=2, value=str(val)).font = data_font
                row += 1
            row += 1

        # Course scores
        if course_data and course_data.get('course_scores'):
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            ws.cell(row=row, column=1, value='各科成绩').font = header_font
            row += 1
            ws.cell(row=row, column=1, value='课程名称').font = header_font
            ws.cell(row=row, column=2, value='分数').font = header_font
            ws.cell(row=row, column=3, value='是否及格').font = header_font
            row += 1
            failed = set(course_data.get('failed_courses', []))
            for cn, score in course_data.get('course_scores', {}).items():
                is_pass = cn not in failed
                c1 = ws.cell(row=row, column=1, value=cn)
                c2 = ws.cell(row=row, column=2, value=score)
                c3 = ws.cell(row=row, column=3, value='✓ 及格' if is_pass else '✗ 不及格')
                c1.font = data_font; c2.font = data_font; c3.font = data_font
                if not is_pass:
                    c1.font = red_font; c2.font = red_font; c3.font = red_font
                    c1.fill = red_fill; c2.fill = red_fill; c3.fill = red_fill
                row += 1
            row += 1

        # Trend
        if trend_data:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            ws.cell(row=row, column=1, value='学期趋势').font = header_font
            row += 1
            for td in trend_data:
                ws.cell(row=row, column=1, value=str(td.get('semester', ''))).font = data_font
                ws.cell(row=row, column=2, value=str(td.get('comp', '—'))).font = data_font
                row += 1
            row += 1

        # Counselor notes
        if counselor_notes:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            ws.cell(row=row, column=1, value='辅导员评语').font = header_font
            row += 1
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            ws.cell(row=row, column=1, value=counselor_notes).font = data_font
            row += 2

        # Footer
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws.cell(row=row, column=1, value='顿河学院团委秘书处').font = Font(name='SimSun', size=9, italic=True)
        ws.cell(row=row, column=1).alignment = center

        # Column widths
        for ci, w in enumerate([25, 15, 15, 15], 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

        import os
        output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..')
        output_path = os.path.join(output_dir, f'个人报告_{student_data.get("name","")}_{student_data.get("id","")}.xlsx')
        wb.save(output_path)
        wb.close()
        return output_path
    except Exception as e:
        import traceback
        return f'ERROR: {str(e)}\n{traceback.format_exc()}'


# ============================================================
# V6.2: Course Analysis & Multi-Semester
# ============================================================

@eel.expose
def analyze_semester_courses_api(file_path: str, grade_filter: str = '',
                                  major_filter: str = '') -> dict:
    """Analyze single semester grades for course-level insights."""
    from backend.course_analyzer import analyze_semester_courses
    return analyze_semester_courses(
        file_path,
        grade_filter=grade_filter or None,
        major_filter=major_filter or None,
    )


@eel.expose
def generate_parent_notice(student_data: dict, semester: str,
                            class_avg: dict = None) -> str:
    """Generate a formatted parent notice for a single student.

    Args:
        student_data: {name, id, class, gpa, moral, quality, comp, rank, ...}
        semester: e.g. '2025-2026-1'
        class_avg: {gpa, moral, comp} class averages for comparison

    Returns:
        Formatted text notice (can be printed or exported).
    """
    try:
        prompt = f"""你是一位高校辅导员。请为以下学生撰写一份《学业情况告知书》，语气温和但严肃，适合发给家长。

学生信息：
- 姓名：{student_data.get('name', '')}
- 学号：{student_data.get('id', '')}
- 班级：{student_data.get('class', '')}
- 学期：{semester}
- 学分绩点：{student_data.get('gpa', '—')}（班级平均：{class_avg.get('gpa', '—') if class_avg else '—'}）
- 德育分：{student_data.get('moral', '—')}（班级平均：{class_avg.get('moral', '—') if class_avg else '—'}）
- 素拓分：{student_data.get('quality', '—')}
- 综测成绩：{student_data.get('comp', '—')}（班级平均：{class_avg.get('comp', '—') if class_avg else '—'}）
- 班级排名：第{student_data.get('rank', '—')}名
- 挂科情况：{student_data.get('failed_courses', '无')}

要求：
1. 200-300字
2. 包含：成绩汇报、与班级平均对比、优势与不足、改进建议
3. 末尾署名「顿河学院团委秘书处」
4. 开头直接写您好，不要用尊敬的家长开头
"""
        return _generate_simple_notice(student_data, semester, class_avg)
    except Exception:
        return _generate_simple_notice(student_data, semester, class_avg)


def _generate_simple_notice(student_data: dict, semester: str,
                             class_avg: dict = None) -> str:
    """Fallback simple notice without AI."""
    name = student_data.get('name', '')
    cls = student_data.get('class', '')
    gpa = student_data.get('gpa', '—')
    moral = student_data.get('moral', '—')
    comp = student_data.get('comp', '—')
    rank = student_data.get('rank', '—')

    notice = f"""学业情况告知书

您好！

现将{name}同学（{cls}）在{semester}学期的学业情况汇报如下：

学分绩点：{gpa}
德育分：{moral}
综测成绩：{comp}
班级排名：第{rank}名"""

    if class_avg:
        notice += f"""
班级平均：绩点{class_avg.get('gpa', '—')} / 德育{class_avg.get('moral', '—')} / 综测{class_avg.get('comp', '—')}"""

    notice += """

请家长与学校共同努力，关注学生的学业发展与全面成长。

顿河学院团委秘书处"""
    return notice


@eel.expose
def batch_generate_notices(student_list: list, semester: str,
                            class_averages: dict = None) -> list:
    """Generate parent notices for multiple students.

    Returns list of {id, name, notice_text}.
    """
    notices = []
    for s in student_list:
        sid = s.get('id', '')
        cls = s.get('class', '')
        ca = class_averages.get(cls, None) if class_averages else None
        notice = _generate_simple_notice(s, semester, ca)
        notices.append({'id': sid, 'name': s.get('name', ''), 'notice': notice})
    return notices


@eel.expose
def save_conversation_records(records: list) -> bool:
    """Save conversation records to persistent storage."""
    import json
    try:
        path = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                            '..', 'data', 'conversation_records.json')
        os.makedirs(os.path.dirname(path), exist_ok=True)
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(records, f, ensure_ascii=False, indent=2)
        return True
    except Exception:
        return False


@eel.expose
def load_conversation_records() -> list:
    """Load conversation records from persistent storage."""
    import json
    try:
        path = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                            '..', 'data', 'conversation_records.json')
        if os.path.exists(path):
            with open(path, 'r', encoding='utf-8') as f:
                return json.load(f)
    except Exception:
        pass
    return []


@eel.expose
def smart_detect_file_info(file_path: str) -> dict:
    """Smart detect file info from filename and headers.

    Returns {semester, grade, major, file_type} inferred from filename.
    """
    basename = os.path.basename(file_path)
    info = {'file_type': 'unknown', 'semester': '', 'grade': '', 'major': ''}

    # Detect semester from filename patterns
    sem_patterns = [
        (r'(\d{4}-\d{4}-\d{1,2})', 1),  # 2025-2026-1
        (r'(\d{4}-\d{4})', 1),            # 2025-2026
        (r'(\d{2}-\d{2}-\d{1,2})', 1),    # 25-26-1
    ]
    for pat, grp in sem_patterns:
        m = re.search(pat, basename)
        if m:
            info['semester'] = m.group(grp)
            break

    # Detect grade
    grade_m = re.search(r'(\d{2})级', basename)
    if grade_m:
        info['grade'] = grade_m.group(1)
    else:
        # Try to find 4-digit year pattern
        year_m = re.search(r'20(\d{2})级', basename)
        if year_m:
            info['grade'] = year_m.group(1)

    # Detect major
    major_keywords = ['顿河交', '顿河土', '顿河信', '国电']
    for mk in major_keywords:
        if mk in basename:
            info['major'] = mk
            info['file_type'] = 'grade'
            break

    # Detect file type by content keywords
    type_keywords = {
        'grade': ['成绩', '绩点', 'gpa', '学分'],
        'moral': ['德育', '卫生', '旷课', '出勤', '宿舍', '团课'],
        'quality': ['素拓', '拓展', '加分', '活动'],
        'comprehensive': ['综测', '综合', '排名'],
    }
    for ftype, kws in type_keywords.items():
        for kw in kws:
            if kw in basename.lower():
                info['file_type'] = ftype
                break
        if info['file_type'] != 'unknown':
            break

    return info


# ============================================================
# V6.2: Deep Analytics Dashboard
# ============================================================

@eel.expose
def get_deep_analytics(data_json: str) -> dict:
    """Generate deep analytics from current student data.

    Args:
        data_json: JSON string of [{id, name, class, gpa, moral, quality, comp}, ...]

    Returns comprehensive analytics.
    """
    import json
    try:
        data = json.loads(data_json) if isinstance(data_json, str) else data_json
        if not data:
            return {'success': False, 'error': '无数据'}

        n = len(data)

        # GPA stats
        gpas = [d.get('gpa', 0) or 0 for d in data]
        gpas_valid = [g for g in gpas if g > 0]
        gpa_avg = sum(gpas_valid) / len(gpas_valid) if gpas_valid else 0
        gpa_std = (sum((g - gpa_avg) ** 2 for g in gpas_valid) / len(gpas_valid)) ** 0.5 if gpas_valid else 0

        # Comp stats
        comps = [d.get('comp', 0) or 0 for d in data]
        comps_valid = [c for c in comps if c > 0]
        comp_avg = sum(comps_valid) / len(comps_valid) if comps_valid else 0
        comp_std = (sum((c - comp_avg) ** 2 for c in comps_valid) / len(comps_valid)) ** 0.5 if comps_valid else 0

        # Moral stats
        morals = [d.get('moral', 0) or 0 for d in data]
        morals_valid = [m for m in morals if m > 0]
        moral_avg = sum(morals_valid) / len(morals_valid) if morals_valid else 0

        # Class grouping
        classes = defaultdict(list)
        for d in data:
            cls = d.get('class', '未知')
            classes[cls].append(d)

        # Grade grouping
        grades = defaultdict(list)
        for d in data:
            cls = str(d.get('class', ''))
            m = re.match(r'.*?(\d{2})\d{1,2}$', cls)
            grade = m.group(1) + '级' if m else '未知'
            grades[grade].append(d)

        # Distribution analysis
        comp_bins = [0, 30, 50, 60, 70, 80, 90, 100]
        comp_dist = []
        for i in range(len(comp_bins) - 1):
            lo, hi = comp_bins[i], comp_bins[i+1]
            cnt = sum(1 for c in comps_valid if lo <= c < hi)
            comp_dist.append({'range': f'{lo}-{hi}', 'count': cnt, 'pct': round(cnt/n*100, 1) if n else 0})

        # Top/Bottom performers
        sorted_by_comp = sorted(data, key=lambda d: d.get('comp', 0) or 0, reverse=True)
        top10 = [{'name': d.get('name', ''), 'class': d.get('class', ''),
                  'comp': d.get('comp', 0), 'gpa': d.get('gpa', 0)}
                 for d in sorted_by_comp[:10]]
        bottom10 = [{'name': d.get('name', ''), 'class': d.get('class', ''),
                     'comp': d.get('comp', 0), 'gpa': d.get('gpa', 0)}
                    for d in sorted_by_comp[-10:]]

        # Class comparison
        class_comparison = []
        for cls, sts in classes.items():
            c_comps = [(d.get('comp', 0) or 0) for d in sts]
            c_gpas = [(d.get('gpa', 0) or 0) for d in sts if (d.get('gpa', 0) or 0) > 0]
            class_comparison.append({
                'class': cls, 'count': len(sts),
                'comp_avg': round(sum(c_comps)/len(c_comps), 2) if c_comps else 0,
                'gpa_avg': round(sum(c_gpas)/len(c_gpas), 2) if c_gpas else 0,
                'fail_count': sum(1 for c in c_comps if c < 60),
            })
        class_comparison.sort(key=lambda x: x['comp_avg'], reverse=True)

        # Grade comparison
        grade_comparison = []
        for grd, sts in grades.items():
            g_comps = [(d.get('comp', 0) or 0) for d in sts]
            grade_comparison.append({
                'grade': grd, 'count': len(sts),
                'comp_avg': round(sum(g_comps)/len(g_comps), 2) if g_comps else 0,
            })
        grade_comparison.sort(key=lambda x: x['comp_avg'], reverse=True)

        return {
            'success': True,
            'summary': {
                'total': n,
                'gpa_avg': round(gpa_avg, 2),
                'gpa_std': round(gpa_std, 2),
                'comp_avg': round(comp_avg, 2),
                'comp_std': round(comp_std, 2),
                'moral_avg': round(moral_avg, 1),
            },
            'comp_distribution': comp_dist,
            'top10': top10,
            'bottom10': bottom10,
            'class_comparison': class_comparison,
            'grade_comparison': grade_comparison,
        }
    except Exception as e:
        return {'success': False, 'error': str(e)}


# ============================================================
# V7.1: Semi-finished quality file parsing (辅助人工素拓汇总)
# ============================================================

@eel.expose
def parse_semi_quality_file(file_path: str, roster: dict = None) -> dict:
    """Parse a semi-finished quality score Excel file.

    The file has scores already filled in per student but no totals.
    This function reads the file, identifies score columns, sums them
    per student, and returns the aggregated results.
    """
    import pandas as pd
    import openpyxl

    if roster is None:
        roster = {}

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        return {'success': False, 'error': f'无法打开文件: {e}'}

    result_students = []
    all_headers = []
    score_columns = []

    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue

        # Find header row
        header_idx = 0
        for i, row in enumerate(rows[:10]):
            non_empty = sum(1 for c in row if c is not None and str(c).strip())
            if non_empty >= 3:
                header_idx = i
                break

        headers = [str(c).strip() if c is not None else '' for c in rows[header_idx]]
        all_headers = headers

        # Identify columns
        id_col = None
        name_col = None
        score_cols = []

        for ci, h in enumerate(headers):
            if '学号' in h and id_col is None:
                id_col = ci
            elif '姓名' in h and name_col is None:
                name_col = ci
            elif any(kw in h for kw in ['分', '加分', 'score']):
                score_cols.append(ci)

        # If no explicit score columns, use numeric columns after name
        if not score_cols:
            for ci in range(2, len(headers)):
                if ci != id_col and ci != name_col:
                    num_count = 0
                    total_check = 0
                    for row in rows[header_idx + 1:header_idx + 11]:
                        if ci < len(row) and row[ci] is not None:
                            try:
                                float(row[ci])
                                num_count += 1
                            except (ValueError, TypeError):
                                pass
                            total_check += 1
                    if num_count > 0 and num_count / max(total_check, 1) > 0.3:
                        score_cols.append(ci)

        score_columns = [headers[ci] for ci in score_cols if ci < len(headers)]

        # Parse student rows
        for row in rows[header_idx + 1:]:
            if not row or all(c is None for c in row):
                continue

            sid = str(row[id_col]).strip() if id_col is not None and id_col < len(row) and row[id_col] is not None else ''
            name = str(row[name_col]).strip() if name_col is not None and name_col < len(row) and row[name_col] is not None else ''

            if sid.endswith('.0') and sid[:-2].isdigit():
                sid = sid[:-2]
            sid = sid.strip()

            if not sid or sid in ('nan', 'None', '', '学号'):
                continue

            # Get name/class from roster if available
            if sid in roster:
                name = roster[sid].get('name', name)
                cls = roster[sid].get('class', str(sn))
            else:
                cls = str(sn)

            # Collect scores
            items = []
            total = 0.0
            for ci in score_cols:
                if ci < len(row) and row[ci] is not None:
                    try:
                        score = float(row[ci])
                        if score != 0:
                            hdr = headers[ci] if ci < len(headers) else f'列{ci}'
                            items.append({'name': hdr, 'score': score})
                            total += score
                    except (ValueError, TypeError):
                        pass

            if sid:
                result_students.append({
                    'id': sid,
                    'name': name,
                    'class': cls,
                    'items': items,
                    'total': round(total, 2),
                })

    wb.close()

    if not result_students:
        return {'success': False, 'error': '未识别到任何学生数据，请检查文件格式（需包含学号列和加分列）'}

    return {
        'success': True,
        'student_count': len(result_students),
        'students': result_students,
        'headers': all_headers,
        'score_columns': score_columns,
    }


# ============================================================
# V7.1: Debug moral file parsing
# ============================================================

@eel.expose
def debug_moral_file(file_path: str) -> dict:
    """Debug tool: show what the moral parser sees in an absence/hygiene file.

    Returns the raw rows, detected format, and which students would match.
    """
    import openpyxl

    result = {'success': True, 'sheets': [], 'total_rows': 0, 'matched': 0}

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        return {'success': False, 'error': f'无法打开文件: {e}'}

    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        sheet_info = {
            'name': sn,
            'row_count': len(rows),
            'first_rows': [],
            'format': 'unknown',
        }

        # Show first 15 rows
        for i, row in enumerate(rows[:15]):
            cells = [str(c).strip() if c is not None else '' for c in row]
            sheet_info['first_rows'].append(cells)

        # Detect format
        has_class = False
        has_student_id = False
        has_course = False
        has_hours = False
        has_deduction = False

        for row in rows[:10]:
            row_text = ' '.join([str(c) if c else '' for c in row])
            if '学号' in row_text:
                has_student_id = True
            if '所缺课程' in row_text or '任课教师' in row_text:
                has_course = True
            if '合计学时' in row_text or '缺课学时' in row_text:
                has_hours = True
            if '扣分' in row_text:
                has_deduction = True
            if '班级' in row_text:
                has_class = True

        if has_course and has_hours:
            sheet_info['format'] = 'summary (旷课统计汇总)'
        elif has_deduction:
            sheet_info['format'] = 'discipline (信息部扣分格式)'
        elif has_class and has_hours:
            sheet_info['format'] = 'discipline'
        else:
            sheet_info['format'] = 'simple/generic'

        sheet_info['detected'] = {
            'has_student_id': has_student_id,
            'has_class': has_class,
            'has_course': has_course,
            'has_hours': has_hours,
            'has_deduction': has_deduction,
        }

        result['sheets'].append(sheet_info)
        result['total_rows'] += len(rows)

    wb.close()
    return result


@eel.expose
def debug_roster_file(file_path: str) -> dict:
    """Debug tool: show what _load_roster extracts from a GPA/roster file."""
    from backend.module_b_moral import _load_roster
    roster = _load_roster(file_path)
    if not roster:
        return {'success': False, 'error': '未提取到任何学生，请检查文件是否为学分绩点输出表。'}
    # Show sample + stats
    samples = []
    classes = set()
    for i, (sid, info) in enumerate(roster.items()):
        classes.add(info.get('class', '?'))
        if i < 20:
            samples.append({'id': sid, 'name': info.get('name', ''), 'class': info.get('class', '')})
    return {
        'success': True,
        'student_count': len(roster),
        'class_count': len(classes),
        'classes': sorted(classes)[:30],
        'samples': samples,
    }


# ============================================================
# V9.0: Smart Material Import (智能材料导入)
# ============================================================

# Smart name detection patterns
_CLASS_PATTERNS = [
    re.compile(r'(顿河[交土信]\d{2}\d{1,2})'),
    re.compile(r'(国[电商]\d{2}\d{1,2})'),
]
_STUDENT_ID_PATTERN = re.compile(r'(\d{6,12})')
_NON_NAME_WORDS = {
    '材料', '证明', '证书', '加分', '素拓', '综测', '图片', '照片',
    '文件', '附件', '压缩包', '资料', '佐证', '汇总', '成绩', '表格',
    '文档', '截图', '扫描', '复印件', '原件', 'new', '新建文件夹',
}
_IMAGE_EXTS = {'.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp', '.tiff'}
_PDF_EXTS = {'.pdf'}
_TEXT_EXTS = {'.txt', '.md', '.py', '.js', '.html', '.css', '.json', '.xml',
              '.csv', '.log', '.yaml', '.yml', '.ini', '.cfg', '.sh', '.bat'}


def _safe_filename(name: str) -> str:
    """Remove characters unsafe for Windows filenames."""
    return re.sub(r'[<>:\"/\\|?*]', '_', name).strip()


def _smart_detect_class_student(name: str) -> dict:
    """Try to detect class name and student name from a directory/file name.

    Returns:
        {type: 'class'|'student'|'unknown', class_name, student_name, student_id}
    """
    result = {'type': 'unknown', 'class_name': '', 'student_name': '', 'student_id': ''}

    # 1. Try class patterns first
    for pat in _CLASS_PATTERNS:
        m = pat.search(name)
        if m:
            result['type'] = 'class'
            result['class_name'] = m.group(1)
            return result

    # 2. Extract student ID
    sid_m = _STUDENT_ID_PATTERN.search(name)
    if sid_m:
        result['student_id'] = sid_m.group(1)

    # 3. Try to extract Chinese name (2-4 chars, not non-name words)
    # Remove digits and common separators first
    clean = re.sub(r'[\d\s_\-\.（）()\[\]【】]+', '', name)
    # Find Chinese character sequences
    cn_matches = re.findall(r'([一-鿿]{2,4})', clean)
    for cn in cn_matches:
        if cn not in _NON_NAME_WORDS:
            result['student_name'] = cn
            result['type'] = 'student'
            break

    # 4. If no Chinese name found but has ID, it's a student
    if result['student_id'] and result['type'] == 'unknown':
        result['type'] = 'student'

    # 5. Fallback: treat the whole cleaned name as student name
    if result['type'] == 'unknown' and len(clean) >= 2:
        if clean not in _NON_NAME_WORDS:
            result['student_name'] = clean
            result['type'] = 'student'

    return result


def _get_class_name_from_zip(zip_path: str) -> str:
    """Guess class name from zip filename."""
    basename = os.path.splitext(os.path.basename(zip_path))[0]
    detected = _smart_detect_class_student(basename)
    if detected['class_name']:
        return detected['class_name']
    # Remove common suffixes
    for suffix in ['素拓', '加分', '材料', '压缩包', '_', '-']:
        basename = basename.replace(suffix, ' ')
    basename = basename.strip()
    return basename if basename else '未识别班级'


def _get_7z_path() -> str:
    """Get path to the bundled 7z.exe (works in dev and PyInstaller mode)."""
    import sys as _sys
    if getattr(_sys, 'frozen', False):
        base = _sys._MEIPASS
    else:
        base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base, 'tools', '7z.exe')


def _extract_archive(archive_path: str, dest_dir: str) -> bool:
    """Extract a ZIP/RAR/7z archive to dest_dir.

    ZIP  → built-in zipfile (fast, always available)
    RAR/7z → bundled 7z.exe (no external dependencies needed)
    """
    ext = os.path.splitext(archive_path)[1].lower()

    if ext == '.zip':
        with zipfile.ZipFile(archive_path, 'r') as zf:
            zf.extractall(dest_dir)
        return True

    # RAR, 7z, and everything else → bundled 7z.exe
    sz = _get_7z_path()
    if os.path.isfile(sz):
        try:
            subprocess.run(
                [sz, 'x', archive_path, f'-o{dest_dir}', '-y'],
                check=True, capture_output=True, timeout=300,
                creationflags=0x08000000 if sys.platform == 'win32' else 0,
            )
            return True
        except subprocess.CalledProcessError as e:
            stderr = e.stderr.decode('utf-8', errors='replace') if e.stderr else ''
            raise RuntimeError(f'解压失败({os.path.basename(archive_path)}): {stderr[:200]}')
        except Exception as e:
            raise RuntimeError(f'解压失败({os.path.basename(archive_path)}): {e}')
    else:
        # Fallback: try patoolib (requires system WinRAR/7-Zip)
        try:
            import patoolib
            patoolib.extract_archive(archive_path, outdir=dest_dir, interactive=False)
            return True
        except ImportError:
            raise RuntimeError('需要安装patool库。请在终端运行: pip install patool')
        except Exception as e:
            raise RuntimeError(
                f'解压失败({os.path.basename(archive_path)}): {e}. '
                '请确保系统已安装WinRAR或7-Zip。')


def _flatten_single_child_folders(dir_path: str) -> int:
    """Aggressively flatten folder nesting.

    Rule: if ANY folder contains exactly ONE item and that item is a
    directory, merge the child's contents up and remove the wrapper.
    This runs recursively until the tree is fully clean.

    Before:  班级/张三/张三/证书.jpg  or  班级/顿河交241/张三/证书.jpg
    After:   班级/张三/证书.jpg

    Returns:
        Number of folders flattened.
    """
    total = 0
    if not os.path.isdir(dir_path):
        return total

    while True:
        flattened_this_pass = 0

        for entry in os.listdir(dir_path):
            entry_path = os.path.join(dir_path, entry)
            if not os.path.isdir(entry_path):
                continue

            # First, recursively flatten deeper levels
            total += _flatten_single_child_folders(entry_path)

            # Now check: does this folder contain exactly one item
            # that is itself a directory?
            try:
                contents = os.listdir(entry_path)
            except OSError:
                continue

            # Filter out hidden files
            visible = [c for c in contents if not c.startswith('.')
                       and c not in ('Thumbs.db', '__MACOSX')]

            if len(visible) == 1:
                child = visible[0]
                child_path = os.path.join(entry_path, child)
                if os.path.isdir(child_path):
                    # Merge child's contents up to entry_path
                    try:
                        for item in os.listdir(child_path):
                            src = os.path.join(child_path, item)
                            dst = os.path.join(entry_path, item)
                            if os.path.exists(dst):
                                if os.path.isdir(src) and os.path.isdir(dst):
                                    # Merge subdirectories
                                    for sub in os.listdir(src):
                                        shutil.move(
                                            os.path.join(src, sub),
                                            os.path.join(dst, sub))
                                    shutil.rmtree(src)
                                else:
                                    # File conflict: rename
                                    base, ext = os.path.splitext(item)
                                    counter = 1
                                    while os.path.exists(
                                        os.path.join(entry_path,
                                                     f'{base}_{counter}{ext}')):
                                        counter += 1
                                    shutil.move(
                                        src,
                                        os.path.join(entry_path,
                                                     f'{base}_{counter}{ext}'))
                            else:
                                shutil.move(src, dst)
                        # Remove the now-empty child folder
                        if os.path.isdir(child_path):
                            shutil.rmtree(child_path)
                        flattened_this_pass += 1
                        total += 1
                    except OSError:
                        pass

        if flattened_this_pass == 0:
            break

    return total


def _build_file_tree(dir_path: str, prefix: str = '') -> list:
    """Build a nested file/folder tree structure for a directory.

    Returns a list of nodes:
    [{name, type: 'file'|'dir', children?: [...], path: relative_path}]

    All paths use forward slashes for safe embedding in JS/HTML contexts.
    """
    if not os.path.isdir(dir_path):
        return []

    items = []
    try:
        entries = sorted(os.listdir(dir_path))
    except OSError:
        return []

    for entry in entries:
        if entry.startswith('.') or entry == 'Thumbs.db':
            continue
        full_path = os.path.join(dir_path, entry)
        rel = os.path.join(prefix, entry) if prefix else entry
        # Normalize to forward slashes — critical for safe JS embedding
        rel = rel.replace('\\', '/')

        if os.path.isdir(full_path):
            children = _build_file_tree(full_path, rel)
            items.append({
                'name': entry,
                'type': 'dir',
                'path': rel,
                'children': children,
            })
        else:
            items.append({
                'name': entry,
                'type': 'file',
                'path': rel,
                'ext': os.path.splitext(entry)[1].lower(),
            })

    return items


def _extract_nested_archives(root_dir: str) -> int:
    """Recursively find and extract ALL nested archives (zip/rar/7z) inside root_dir.

    Keeps scanning and extracting until no more archives remain
    (handles archives-within-archives).

    Each nested archive is extracted into a subfolder named after the archive
    (without extension), then the original archive is deleted.

    Returns:
        Total number of nested archives extracted.
    """
    total_count = 0
    archive_exts = {'.zip', '.rar', '.7z'}

    while True:
        found = False
        for dirpath, dirnames, filenames in os.walk(root_dir):
            # Skip hidden dirs
            dirnames[:] = [d for d in dirnames
                           if not d.startswith('.') and d != '__MACOSX']
            for fn in filenames:
                if fn.startswith('.') or fn == 'Thumbs.db':
                    continue
                ext = os.path.splitext(fn)[1].lower()
                if ext not in archive_exts:
                    continue

                archive_path = os.path.join(dirpath, fn)
                base = os.path.splitext(fn)[0]
                extract_dir = os.path.join(dirpath, base)

                # Avoid overwriting
                if os.path.exists(extract_dir):
                    extract_dir = os.path.join(dirpath, base + '_解压')

                try:
                    os.makedirs(extract_dir, exist_ok=True)
                    _extract_archive(archive_path, extract_dir)
                    os.remove(archive_path)
                    total_count += 1
                    found = True
                    # Break out of os.walk to restart scan
                    break
                except Exception:
                    pass

            if found:
                break  # Break outer os.walk, restart while loop

        if not found:
            break  # No more archives found, done

    return total_count


@eel.expose
def smart_unzip_materials(zip_paths: list, output_base_dir: str) -> dict:
    """Smart unzip class material archives with class/student name detection.

    Supports ZIP, RAR, and 7z formats.

    For each archive file:
    1. Extract to temp dir
    2. Scan folder names to detect class and student names
    3. Re-organize as: output_base_dir/班级名/学生名/files

    Args:
        zip_paths: List of paths to archive files
        output_base_dir: Target directory for organized output

    Returns:
        {success, classes: [{name, students: [{name, id, file_count, files}]}],
         errors, total_students, total_files, output_dir}
    """
    os.makedirs(output_base_dir, exist_ok=True)
    all_errors = []
    all_class_data = {}  # class_name -> {students: {student_key -> {name, id, files}}}

    for zp in (zip_paths or []):
        if not zp or not os.path.exists(zp):
            all_errors.append(f'文件不存在: {zp}')
            continue

        try:
            with tempfile.TemporaryDirectory() as tmpdir:
                # Extract archive (ZIP, RAR, or 7z)
                _extract_archive(zp, tmpdir)

                # Detect class name from zip filename
                class_name = _get_class_name_from_zip(zp)

                if class_name not in all_class_data:
                    all_class_data[class_name] = {}

                # ---- Find content root: step into single-directory wrappers ----
                content_dir = tmpdir
                entries = os.listdir(content_dir)
                # Filter out hidden files
                visible = [e for e in entries
                           if not e.startswith('.') and e != '__MACOSX']
                # If only one directory, step into it
                while (len(visible) == 1
                       and os.path.isdir(os.path.join(content_dir, visible[0]))):
                    content_dir = os.path.join(content_dir, visible[0])
                    visible = [e for e in os.listdir(content_dir)
                               if not e.startswith('.') and e != '__MACOSX']

                # ---- Now every item in content_dir is a student ----
                for entry in sorted(visible):
                    entry_path = os.path.join(content_dir, entry)
                    ext = os.path.splitext(entry)[1].lower()

                    # Case 1: Student zip/rar file → extract as student
                    if os.path.isfile(entry_path) and ext in ('.zip', '.rar', '.7z'):
                        student_name = os.path.splitext(entry)[0]
                        # Strip class prefix from student name
                        if class_name and student_name.startswith(class_name):
                            student_name = student_name[len(class_name):].strip()
                        student_name = student_name.strip() or entry
                        safe_student = _safe_filename(student_name)
                        student_dir = os.path.join(
                            output_base_dir, class_name, safe_student)
                        os.makedirs(student_dir, exist_ok=True)
                        try:
                            _extract_archive(entry_path, student_dir)
                        except Exception:
                            # If extraction fails, copy the file as-is
                            shutil.copy2(entry_path, student_dir)
                        # Remove the original archive from temp
                        try:
                            os.remove(entry_path)
                        except OSError:
                            pass

                    # Case 2: Directory → copy as student folder
                    elif os.path.isdir(entry_path):
                        student_name = entry.strip()
                        # Strip class prefix
                        if class_name and student_name.startswith(class_name):
                            student_name = student_name[len(class_name):].strip()
                        student_name = student_name or entry
                        safe_student = _safe_filename(student_name)
                        student_dir = os.path.join(
                            output_base_dir, class_name, safe_student)
                        if os.path.exists(student_dir):
                            # Merge: copy contents
                            for root, dirs, filenames in os.walk(entry_path):
                                for fn in filenames:
                                    src = os.path.join(root, fn)
                                    rel = os.path.relpath(src, entry_path)
                                    dst = os.path.join(student_dir, rel)
                                    os.makedirs(os.path.dirname(dst), exist_ok=True)
                                    if not os.path.exists(dst):
                                        shutil.copy2(src, dst)
                        else:
                            shutil.copytree(entry_path, student_dir)

                    # Case 3: Loose file → group under a "未分类" student
                    elif os.path.isfile(entry_path):
                        unknown_dir = os.path.join(
                            output_base_dir, class_name, '未分类')
                        os.makedirs(unknown_dir, exist_ok=True)
                        shutil.copy2(entry_path, unknown_dir)

        except zipfile.BadZipFile:
            all_errors.append(f'不是有效的压缩文件: {os.path.basename(zp)}')
        except RuntimeError as e:
            all_errors.append(str(e))
        except Exception as e:
            all_errors.append(f'处理 {os.path.basename(zp)} 时出错: {str(e)}')

    # V9.1: Extract nested archives & aggressive flattening
    nested_count = 0
    flatten_count = 0
    if all_class_data:
        for cls_name in all_class_data:
            cls_dir = os.path.join(output_base_dir, cls_name)
            if os.path.isdir(cls_dir):
                nested_count += _extract_nested_archives(cls_dir)
                # Run flatten twice — first pass handles deep nesting,
                # second pass cleans up any remaining wrappers
                flatten_count += _flatten_single_child_folders(cls_dir)
                flatten_count += _flatten_single_child_folders(cls_dir)

    # Re-scan: every immediate subdirectory of a class folder IS a student
    for cls_name in list(all_class_data.keys()):
        cls_dir = os.path.join(output_base_dir, cls_name)
        if not os.path.isdir(cls_dir):
            continue

        # Discover students from actual filesystem
        new_students = {}
        try:
            for entry in sorted(os.listdir(cls_dir)):
                entry_path = os.path.join(cls_dir, entry)
                if not os.path.isdir(entry_path):
                    continue
                if entry.startswith('.') or entry == '__MACOSX':
                    continue

                # Every subfolder is a student — try to extract name/id
                detected = _smart_detect_class_student(entry)
                student_name = detected['student_name'] or entry
                student_id = detected['student_id'] or ''
                student_key = student_id if student_id else entry

                # Build file tree
                file_tree = _build_file_tree(entry_path)
                files = []
                for root, dirs, filenames in os.walk(entry_path):
                    dirs[:] = [d for d in dirs
                               if not d.startswith('.') and d != '__MACOSX']
                    for fn in filenames:
                        if fn.startswith('.') or fn == 'Thumbs.db':
                            continue
                        rel = os.path.relpath(
                            os.path.join(root, fn),
                            entry_path).replace('\\', '/')
                        files.append(rel)

                new_students[student_key] = {
                    'name': student_name,
                    'id': student_id,
                    'dir_name': entry,
                    'key': student_key,
                    'file_count': len(files),
                    'files': sorted(files),
                    'file_tree': file_tree,
                }
        except OSError:
            pass

        # Merge with previously detected (keep old if still valid)
        old_students = all_class_data.get(cls_name, {})
        # Prefer new scan results (they reflect the actual filesystem)
        all_class_data[cls_name] = new_students if new_students else old_students

    # Build response
    classes_list = []
    total_students = 0
    total_files = 0

    for cls_name in sorted(all_class_data.keys()):
        students_dict = all_class_data[cls_name]
        students_list = []
        for sk, sdata in students_dict.items():
            students_list.append({
                'name': sdata['name'],
                'id': sdata['id'],
                'dir_name': sdata['dir_name'],
                'key': sk,
                'file_count': len(sdata['files']),
                'files': sdata['files'],
                'file_tree': sdata.get('file_tree', []),
            })
            total_files += len(sdata['files'])
        total_students += len(students_list)
        # Sort students by name
        students_list.sort(key=lambda s: s['name'])
        classes_list.append({
            'name': cls_name,
            'students': students_list,
        })

    return {
        'success': len(classes_list) > 0,
        'classes': classes_list,
        'errors': all_errors,
        'total_students': total_students,
        'total_files': total_files,
        'nested_extracted': nested_count,
        'folders_flattened': flatten_count,
        'output_dir': output_base_dir,
    }


@eel.expose
def get_material_directory_tree(base_dir: str) -> dict:
    """Read the organized material directory and return tree structure.

    Returns:
        {success, classes: [{name, students: [{name, id, key, file_count, files}]}]}
    """
    if not base_dir or not os.path.exists(base_dir):
        return {'success': False, 'error': '目录不存在', 'classes': []}

    classes_list = []
    try:
        entries = sorted(os.listdir(base_dir))
        for entry in entries:
            entry_path = os.path.join(base_dir, entry)
            if not os.path.isdir(entry_path):
                continue

            # entry = class folder
            students_list = []
            sub_entries = sorted(os.listdir(entry_path))
            for sub in sub_entries:
                sub_path = os.path.join(entry_path, sub)
                if not os.path.isdir(sub_path):
                    continue

                # sub = student folder
                # Flatten single-child nesting first
                _flatten_single_child_folders(sub_path)

                files = []
                for root, dirs, filenames in os.walk(sub_path):
                    dirs[:] = [d for d in dirs if not d.startswith('.') and d != '__MACOSX']
                    for fn in filenames:
                        if fn.startswith('.') or fn == 'Thumbs.db':
                            continue
                        full = os.path.join(root, fn)
                        rel = os.path.relpath(full, sub_path).replace('\\', '/')
                        files.append(rel)

                # Build file tree for nested display
                file_tree = _build_file_tree(sub_path)

                # Detect name/ID from folder name
                detected = _smart_detect_class_student(sub)
                student_name = detected['student_name'] or sub
                student_id = detected['student_id'] or ''

                students_list.append({
                    'name': student_name,
                    'id': student_id,
                    'key': student_id or student_name,
                    'dir_name': sub,
                    'file_count': len(files),
                    'files': sorted(files),
                    'file_tree': file_tree,
                })

            students_list.sort(key=lambda s: s['name'])
            classes_list.append({
                'name': entry,
                'students': students_list,
            })
    except Exception as e:
        return {'success': False, 'error': str(e), 'classes': []}

    return {
        'success': True,
        'classes': classes_list,
        'total_students': sum(len(c['students']) for c in classes_list),
        'total_files': sum(
            s['file_count'] for c in classes_list for s in c['students']),
    }


@eel.expose
def rename_material_item(base_dir: str, rel_path: str,
                         new_name: str, item_type: str) -> dict:
    """Rename a class folder or student folder.

    Args:
        base_dir: Base directory
        rel_path: Relative path from base_dir (e.g., '顿河交241/张三')
        new_name: New name (safe for filesystem)
        item_type: 'class' or 'student'

    Returns:
        {success, new_path}
    """
    old_path = os.path.join(base_dir, rel_path)
    if not os.path.exists(old_path):
        return {'success': False, 'error': '路径不存在'}

    new_name = _safe_filename(new_name)
    parent = os.path.dirname(old_path)
    new_path = os.path.join(parent, new_name)

    if os.path.exists(new_path) and old_path != new_path:
        return {'success': False, 'error': '目标名称已存在'}

    try:
        os.rename(old_path, new_path)
        # Return the new relative path
        new_rel = os.path.relpath(new_path, base_dir)
        return {'success': True, 'new_path': new_rel, 'new_name': new_name}
    except Exception as e:
        return {'success': False, 'error': str(e)}


@eel.expose
def delete_material_item(base_dir: str, rel_path: str) -> dict:
    """Delete a student folder or a single file.

    Returns:
        {success, message}
    """
    full_path = os.path.join(base_dir, rel_path)
    if not os.path.exists(full_path):
        return {'success': False, 'error': '路径不存在'}

    try:
        if os.path.isdir(full_path):
            shutil.rmtree(full_path)
        else:
            os.remove(full_path)
        return {'success': True, 'message': f'已删除: {os.path.basename(full_path)}'}
    except Exception as e:
        return {'success': False, 'error': str(e)}


@eel.expose
def add_files_to_student(base_dir: str, student_rel_path: str,
                         selected_paths: list = None) -> dict:
    """Open file picker and copy selected files into student folder.

    Returns:
        {success, added_files: [str]}
    """
    dest_dir = os.path.join(base_dir, student_rel_path)
    if not os.path.exists(dest_dir):
        return {'success': False, 'error': '学生目录不存在'}

    paths = selected_paths
    if paths is None:
        paths = select_files(
            title='选择要添加的文件',
            file_types=[('图片和文档', '*.jpg *.jpeg *.png *.gif *.bmp *.pdf *.doc *.docx'),
                        ('所有文件', '*.*')])

    if not paths:
        return {'success': False, 'error': '未选择文件', 'added_files': []}

    added = []
    for src in paths:
        fn = os.path.basename(src)
        dst = os.path.join(dest_dir, fn)
        # Avoid overwrite: add suffix if exists
        if os.path.exists(dst):
            base, ext = os.path.splitext(fn)
            counter = 1
            while os.path.exists(os.path.join(dest_dir, f'{base}_{counter}{ext}')):
                counter += 1
            dst = os.path.join(dest_dir, f'{base}_{counter}{ext}')
            fn = f'{base}_{counter}{ext}'
        shutil.copy2(src, dst)
        added.append(fn)

    return {'success': True, 'added_files': added}


@eel.expose
def add_student_manually(base_dir: str, class_name: str,
                          student_name: str, student_id: str = '') -> dict:
    """Manually add a student folder to a class in the material directory.

    For students who didn't upload a ZIP but still need scores.

    Args:
        base_dir: Material base directory
        class_name: Class folder name
        student_name: Student display name
        student_id: Optional student ID (学号)

    Returns:
        {success, class_name, student_name, student_id, dir_name}
    """
    if not base_dir or not os.path.exists(base_dir):
        return {'success': False, 'error': '输出目录不存在，请先解压或选择目录'}
    if not class_name or not student_name:
        return {'success': False, 'error': '班级和学生姓名不能为空'}

    class_dir = os.path.join(base_dir, class_name)
    os.makedirs(class_dir, exist_ok=True)

    # Build folder name: prefer ID if available, otherwise name
    safe_name = _safe_filename(student_name)
    if student_id:
        dir_name = f'{student_id}_{safe_name}'
    else:
        dir_name = safe_name

    student_dir = os.path.join(class_dir, dir_name)
    os.makedirs(student_dir, exist_ok=True)

    return {
        'success': True,
        'class_name': class_name,
        'student_name': student_name,
        'student_id': student_id,
        'dir_name': dir_name,
        'key': student_id or student_name,
    }


@eel.expose
def read_material_file(file_path: str) -> dict:
    """Read a material file for in-app preview.

    Images, PDFs return base64 data for inline preview.
    Other files return path for external opening.

    Returns:
        {success, type: 'image'|'pdf'|'text'|'other', data, mime_type, filename, size}
    """
    if not file_path or not os.path.exists(file_path):
        return {'success': False, 'error': '文件不存在'}

    ext = os.path.splitext(file_path)[1].lower()
    filename = os.path.basename(file_path)
    file_size = os.path.getsize(file_path)

    try:
        if ext in _IMAGE_EXTS:
            mime_map = {
                '.jpg': 'image/jpeg', '.jpeg': 'image/jpeg',
                '.png': 'image/png', '.gif': 'image/gif',
                '.bmp': 'image/bmp', '.webp': 'image/webp',
                '.tiff': 'image/tiff',
            }
            mime = mime_map.get(ext, 'image/png')
            with open(file_path, 'rb') as f:
                data = base64.b64encode(f.read()).decode('ascii')
            return {
                'success': True, 'type': 'image',
                'data': f'data:{mime};base64,{data}',
                'mime_type': mime, 'filename': filename, 'size': file_size,
            }
        elif ext in _PDF_EXTS:
            # V9.2: Read PDF as base64 for in-app <embed> preview
            with open(file_path, 'rb') as f:
                data = base64.b64encode(f.read()).decode('ascii')
            return {
                'success': True, 'type': 'pdf',
                'data': f'data:application/pdf;base64,{data}',
                'mime_type': 'application/pdf', 'filename': filename, 'size': file_size,
            }
        elif ext in _TEXT_EXTS:
            # Text/code files: show content directly
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    text = f.read()
            except UnicodeDecodeError:
                with open(file_path, 'r', encoding='gbk', errors='replace') as f:
                    text = f.read()
            return {
                'success': True, 'type': 'text',
                'data': text[:50000],  # Truncate very long files
                'mime_type': 'text/plain', 'filename': filename, 'size': file_size,
            }
        else:
            return {
                'success': True, 'type': 'other',
                'data': file_path, 'mime_type': 'application/octet-stream',
                'filename': filename, 'size': file_size,
            }
    except Exception as e:
        return {'success': False, 'error': str(e)}


@eel.expose
def open_file_externally(file_path: str):
    """Open a file with the system default application (cross-platform)."""
    if file_path and os.path.exists(file_path):
        _open_path_in_os(file_path)


@eel.expose
def save_quality_data_snapshot(base_dir: str, quality_data: dict) -> bool:
    """Save quality scoring data to disk (survives app restart/crash).

    Stored alongside material progress in the output directory.
    """
    try:
        snapshot_file = os.path.join(base_dir, '.quality_data_snapshot.json')
        # Convert to JSON-safe format
        with open(snapshot_file, 'w', encoding='utf-8') as f:
            json.dump(quality_data, f, ensure_ascii=False, indent=2)
        return True
    except Exception:
        return False


@eel.expose
def load_quality_data_snapshot(base_dir: str) -> dict:
    """Load previously saved quality scoring data."""
    try:
        snapshot_file = os.path.join(base_dir, '.quality_data_snapshot.json')
        if os.path.exists(snapshot_file):
            with open(snapshot_file, 'r', encoding='utf-8') as f:
                return json.load(f)
    except Exception:
        pass
    return {}


@eel.expose
def save_material_progress(base_dir: str, progress_data: dict) -> bool:
    """Save material review progress to persistent storage.

    progress_data: {student_key: 'pending'|'processing'|'done'}
    """
    try:
        progress_file = os.path.join(base_dir, '.material_progress.json')
        with open(progress_file, 'w', encoding='utf-8') as f:
            json.dump(progress_data, f, ensure_ascii=False, indent=2)
        return True
    except Exception:
        return False


@eel.expose
def load_material_progress(base_dir: str) -> dict:
    """Load saved material review progress."""
    try:
        progress_file = os.path.join(base_dir, '.material_progress.json')
        if os.path.exists(progress_file):
            with open(progress_file, 'r', encoding='utf-8') as f:
                return json.load(f)
    except Exception:
        pass
    return {}


# ============================================================
# V9.2: Manual quality progress save/restore (user-visible JSON)
# ============================================================

@eel.expose
def save_quality_progress_to_file(file_path: str, quality_data: dict) -> dict:
    """Save quality scoring progress to a user-chosen JSON file.

    The JSON contains:
        {student_key: [{activity, category, grade, score, _manual, _is_total}, ...]}
    """
    try:
        os.makedirs(os.path.dirname(file_path) if os.path.dirname(file_path) else '.', exist_ok=True)
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(quality_data, f, ensure_ascii=False, indent=2)
        return {'success': True, 'path': file_path,
                'student_count': len(quality_data),
                'total_items': sum(len(v) for v in quality_data.values())}
    except Exception as e:
        return {'success': False, 'error': str(e)}


@eel.expose
def load_quality_progress_from_file(file_path: str) -> dict:
    """Load quality scoring progress from a user-chosen JSON file.

    Returns:
        {success, data: {student_key: [...]}, student_count, total_items}
    """
    try:
        if not os.path.exists(file_path):
            return {'success': False, 'error': '文件不存在'}
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        if not isinstance(data, dict):
            return {'success': False, 'error': 'JSON格式不正确：需要对象格式'}
        return {'success': True, 'data': data,
                'student_count': len(data),
                'total_items': sum(len(v) for v in data.values() if isinstance(v, list))}
    except json.JSONDecodeError as e:
        return {'success': False, 'error': f'JSON解析失败: {e}'}
    except Exception as e:
        return {'success': False, 'error': str(e)}


# ============================================================
# V9.2: Smart directory scanning (handles any folder structure)
# ============================================================

def _is_likely_student_folder(name: str) -> bool:
    """Heuristic: does this folder name look like a student folder?"""
    detected = _smart_detect_class_student(name)
    if detected['type'] == 'student':
        return True
    if re.match(r'^\d{6,12}$', name.strip()):
        return True
    cn_chars = re.findall(r'[一-鿿]', name)
    if 2 <= len(cn_chars) <= 4:
        return True
    return False


def _is_likely_class_folder(subdirs: list) -> bool:
    """Does this set of subdirs look like it contains student folders?"""
    student_count = sum(1 for d in subdirs if _is_likely_student_folder(d))
    if student_count >= 2:
        return True
    if len(subdirs) > 0 and all(
        _is_likely_student_folder(d) or re.match(r'^\d{6,12}$', d.strip())
        for d in subdirs
    ):
        return True
    return False


@eel.expose
def smart_scan_directory(dir_path: str) -> dict:
    """Intelligently scan any directory and return a material tree.

    Handles three structure types automatically:
      Type A (base):    base/班级A/学生1/文件, 班级A/学生2/文件
                        → classes detected from subdirs
      Type B (class):   顿河交241/张三/文件, 顿河交241/李四/文件
                        → auto-wraps in one class named after the dir
      Type C (flat):    file1.jpg, 张三/, 李四.pdf
                        → creates one class, detects students by name

    Returns same format as get_material_directory_tree:
        {success, classes: [{name, students: [...]}], total_students, total_files}
    """
    if not dir_path or not os.path.isdir(dir_path):
        return {'success': False, 'error': '目录不存在', 'classes': []}

    try:
        entries = sorted(os.listdir(dir_path))
        visible = [e for e in entries
                   if not e.startswith('.') and e not in ('__MACOSX', 'Thumbs.db')]

        if not visible:
            return {'success': True, 'classes': [], 'total_students': 0, 'total_files': 0}

        # Separate subdirs and files
        subdirs = []
        loose_files = []
        for e in visible:
            full = os.path.join(dir_path, e)
            if os.path.isdir(full):
                try:
                    sub_contents = [s for s in os.listdir(full)
                                    if not s.startswith('.') and s not in ('__MACOSX', 'Thumbs.db')]
                except OSError:
                    sub_contents = []
                subdirs.append((e, full, sub_contents))
            else:
                loose_files.append(e)

        # Determine structure type
        subdir_names = [s[0] for s in subdirs]
        # Type A: subdirs contain more subdirs that look like students
        class_like = sum(1 for name, path, subs in subdirs
                        if _is_likely_class_folder(
                            [s for s in subs if os.path.isdir(os.path.join(path, s))]))
        # Type B: subdirs themselves look like students
        student_like = sum(1 for name, path, subs in subdirs
                          if _is_likely_student_folder(name))

        if class_like > 0 and class_like >= student_like:
            # Type A: Standard base → use existing logic
            return get_material_directory_tree(dir_path)

        elif student_like > 0:
            # Type B: Class → student, auto-wrap
            class_name = os.path.basename(dir_path.rstrip('/\\'))
            detected_cls = _smart_detect_class_student(class_name)
            if detected_cls.get('class_name'):
                class_name = detected_cls['class_name']

            students_list = _build_student_list_from_subdirs(dir_path, subdirs)
            classes_list = [{'name': class_name, 'students': students_list}]

        else:
            # Type C: Flat / unknown
            class_name = os.path.basename(dir_path.rstrip('/\\')) or '未分类'
            detected_cls = _smart_detect_class_student(class_name)
            if detected_cls.get('class_name'):
                class_name = detected_cls['class_name']

            students_dict = {}
            for name, full_path, subs in subdirs:
                detected = _smart_detect_class_student(name)
                sname = detected['student_name'] or name
                sid = detected['student_id'] or ''
                skey = sid or sname

                _flatten_single_child_folders(full_path)
                _flatten_single_child_folders(full_path)

                file_tree = _build_file_tree(full_path)
                all_files = []
                for root, dirs, filenames in os.walk(full_path):
                    dirs[:] = [d for d in dirs
                               if not d.startswith('.') and d != '__MACOSX']
                    for fn in filenames:
                        if fn.startswith('.') or fn == 'Thumbs.db':
                            continue
                        all_files.append(os.path.relpath(
                            os.path.join(root, fn), full_path).replace('\\', '/'))

                students_dict[skey] = {
                    'name': sname, 'id': sid, 'key': skey,
                    'dir_name': name, 'file_count': len(all_files),
                    'files': sorted(all_files), 'file_tree': file_tree,
                }

            if loose_files:
                unclass_dir = os.path.join(dir_path, '未分类材料')
                os.makedirs(unclass_dir, exist_ok=True)
                for fn in loose_files:
                    try:
                        shutil.copy2(os.path.join(dir_path, fn),
                                    os.path.join(unclass_dir, fn))
                    except OSError:
                        pass
                students_dict['未分类'] = {
                    'name': '未分类', 'id': '', 'key': '未分类',
                    'dir_name': '未分类材料', 'file_count': len(loose_files),
                    'files': sorted(loose_files),
                    'file_tree': _build_file_tree(unclass_dir),
                }

            students_list = sorted(students_dict.values(), key=lambda s: s['name'])
            classes_list = [{'name': class_name, 'students': students_list}]

        # Final flatten pass
        for cls in classes_list:
            for s in cls['students']:
                sd = os.path.join(dir_path, cls['name'], s['dir_name'])
                if os.path.isdir(sd):
                    _flatten_single_child_folders(sd)
                    _flatten_single_child_folders(sd)

        total_students = sum(len(c['students']) for c in classes_list)
        total_files = sum(
            s['file_count'] for c in classes_list for s in c['students'])

        return {
            'success': True,
            'classes': classes_list,
            'total_students': total_students,
            'total_files': total_files,
            'structure_type': (
                'base' if class_like > 0 and class_like >= student_like
                else 'class' if student_like > 0
                else 'flat'
            ),
        }

    except Exception as e:
        return {'success': False, 'error': str(e), 'classes': []}


def _build_student_list_from_subdirs(base_dir: str,
                                      subdirs: list) -> list:
    """Build student list from subdirectories of a class folder."""
    students_list = []
    for name, full_path, subs in subdirs:
        _flatten_single_child_folders(full_path)
        _flatten_single_child_folders(full_path)

        detected = _smart_detect_class_student(name)
        student_name = detected['student_name'] or name
        student_id = detected['student_id'] or ''
        student_key = student_id if student_id else name

        file_tree = _build_file_tree(full_path)
        all_files = []
        for root, dirs, filenames in os.walk(full_path):
            dirs[:] = [d for d in dirs
                       if not d.startswith('.') and d != '__MACOSX']
            for fn in filenames:
                if fn.startswith('.') or fn == 'Thumbs.db':
                    continue
                all_files.append(os.path.relpath(
                    os.path.join(root, fn), full_path).replace('\\', '/'))

        students_list.append({
            'name': student_name,
            'id': student_id,
            'key': student_key,
            'dir_name': name,
            'file_count': len(all_files),
            'files': sorted(all_files),
            'file_tree': file_tree,
        })
    students_list.sort(key=lambda s: s['name'])
    return students_list

