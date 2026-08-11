from __future__ import annotations
"""
Module A: GPA Calculation (学分绩点计算)

Reads raw .xls/.xlsx grade tables, splits by class, generates formatted output
with Excel formulas matching the reference format exactly.
"""

import os
import re
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from backend.parsers.xls_reader import read_raw_xls
from backend.parsers.course_header_parser import parse_course_header
from backend.utils.class_utils import parse_class_name, group_by_program_grade, class_matches_program
from backend.utils.rank_calculator import calculate_ranking
from backend.utils.progress_reporter import ProgressReporter
from backend.utils.excel_writer import write_values_sheet, unique_path
from config import PE_KEYWORDS, SCORE_MAPPING


def process_gpa(input_path: str, output_dir: str,
                progress: ProgressReporter = None,
                column_mappings: dict = None, major_filter: str = '') -> dict:
    """Process a single raw grade file. (legacy single-file mode)"""
    return process_gpa_batch([input_path], output_dir, progress, column_mappings, major_filter)


def process_gpa_batch(input_paths: list, output_dir: str,
                      progress: ProgressReporter = None,
                      column_mappings: dict = None,
                      major_filter: str = '') -> dict:
    """Process multiple raw grade files and merge into combined outputs.

    Output format matches the reference file exactly:
    - Row 1: Headers (学号, 姓名, 行政班级, 课程门数, [course names], 学分绩点)
    - Row 2: Credit row (总学分, totalCredits, empty, empty, [credit per course])
    - Rows 3+: Student data with raw scores preserved
    - Last column: Excel formula =AVERAGE(score×credit+...)/totalCredits
    """
    if progress is None:
        progress = ProgressReporter()
    column_mappings = column_mappings or {}

    # v2.3: Consistent styling — SimSun 10pt, centered, thin border, no fills
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal='center', vertical='center')
    header_font = Font(name='SimSun', size=10, bold=True)
    data_font = Font(name='SimSun', size=10)
    formula_font = Font(name='SimSun', size=10)

    total_files = len(input_paths)
    source_jobs = _gpa_source_jobs(input_paths, column_mappings)
    total_jobs = max(1, len(source_jobs))
    all_source_rows = []   # (class_name, raw_data_row_dict)
    all_course_headers = {}  # {canonical_header: {course_name, credit, is_pe}}

    # ---- Phase 1: Read all files ----
    for fi, (path, selected_sheet, file_mapping) in enumerate(source_jobs):
        pct = (fi / total_jobs) * 60
        progress.update(pct, f'读取 {fi+1}/{total_jobs}: {os.path.basename(path)} · {selected_sheet or "默认工作表"}')
        df = _read_gpa_source(path, selected_sheet, file_mapping)
        if df.empty:
            continue

        columns = list(df.columns)
        col_map = _identify_columns(columns, file_mapping)
        course_cols = col_map['course_cols']
        course_defs = col_map.get('course_defs', [])

        # Parse course headers
        for course_def in course_defs:
            hdr_str = str(course_def.get('name') or columns[course_def['score_col']])
            if hdr_str not in all_course_headers:
                parsed = parse_course_header(hdr_str)
                parsed['course_name'] = course_def.get('name') or parsed.get('course_name') or hdr_str
                if course_def.get('credit') is not None:
                    parsed['credit'] = float(course_def.get('credit') or 0)
                if course_def.get('is_pe') is not None:
                    parsed['is_pe'] = bool(course_def.get('is_pe'))
                all_course_headers[hdr_str] = parsed

        # Process each student row — keep RAW values
        for idx, row in df.iterrows():
            excel_row = int(file_mapping.get('header_row', 0)) + int(idx) + 2
            action = (file_mapping.get('row_actions', {}).get(str(excel_row), {})
                      if isinstance(file_mapping, dict) else {})
            if action.get('action') == 'exclude':
                continue
            sid = _safe_str(row.get(col_map['student_id_col'], ''))
            if not sid or sid == 'nan' or not re.fullmatch(r'\d{6,20}', sid):
                continue

            name = _safe_str(row.get(col_map['name_col'], ''))
            class_name = _safe_str(row.get(col_map['class_col'], ''))
            if major_filter and not class_matches_program(class_name, major_filter):
                continue
            course_count = row.get(col_map['course_count_col'], len(course_cols))

            # Build a row dict with raw values
            row_dict = {
                '学号': sid,
                '姓名': name,
                '课程门数': course_count,
                '总学分': 0.0,  # Will be computed later
            }

            detected_count = 0
            for course_def in course_defs:
                score_idx = course_def['score_col']
                value_idx = course_def.get('value_col')
                raw_primary = row.iloc[score_idx] if score_idx < len(row) else None
                raw_companion = row.iloc[value_idx] if isinstance(value_idx, int) and value_idx < len(row) else None
                raw_val = raw_companion if _parse_score(raw_companion) is not None else raw_primary
                hdr_str = str(course_def.get('name') or columns[score_idx])
                info = all_course_headers.get(hdr_str, {})
                course_name = info.get('course_name', hdr_str[:30])
                row_dict[course_name] = raw_val  # Preserve raw value
                if _parse_score(raw_val) is not None:
                    detected_count += 1

            row_dict['课程门数'] = detected_count

            all_source_rows.append((class_name, row_dict))

    progress.update(65, f'读取完成，共 {len(all_source_rows)} 名学生')

    if not all_source_rows:
        suffix = f'；当前专业“{major_filter}”在源文件中没有匹配班级' if major_filter else ''
        raise ValueError(f'未找到任何有效学生数据{suffix}')

    # ---- Phase 2: Split by class ----
    class_groups = {}
    for class_name, row_dict in all_source_rows:
        if class_name not in class_groups:
            class_groups[class_name] = []
        class_groups[class_name].append(row_dict)

    # Sort by student ID within each class
    for cn in class_groups:
        class_groups[cn].sort(key=lambda r: str(r.get('学号', '')))

    progress.update(70, f'已分 {len(class_groups)} 个班级')

    # ---- Phase 3: Determine unified column set ----
    # Collect all course names that appear in any row
    all_course_names = set()
    for cn, rows in class_groups.items():
        for r in rows:
            for k in r:
                if k not in ('学号', '姓名', '行政班级', '课程门数'):
                    all_course_names.add(k)

    # Sort course names to maintain consistent order
    # We order by the original header to match source file order
    info_cols = ['学号', '姓名', '课程门数', '总学分']
    gpa_col_name = '学分绩点'

    # Exclude keywords for non-course columns that might leak
    _non_course_names = {'总学分', '获得学分', '所得学分', '平均学分绩', '平均学分绩点',
                         '学分绩点', '学分', '平均分', '总分', '排名', '不及格门数'}

    # Build a canonical course order from the first file's headers
    course_order = []
    for hdr_str in all_course_headers:
        cn = all_course_headers[hdr_str].get('course_name', '')
        if cn and cn in all_course_names and cn not in _non_course_names:
            if cn not in course_order:
                course_order.append(cn)
    # Add any remaining course names not in the header map
    for cn in sorted(all_course_names):
        if cn not in course_order and cn not in _non_course_names:
            course_order.append(cn)

    progress.update(75, '正在生成 Excel...')

    # ---- Phase 4: Write .xlsx with formulas ----
    os.makedirs(output_dir, exist_ok=True)
    output1_path = unique_path(os.path.join(output_dir, '学分绩点.xlsx'))

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Collect students for ranking and _values sheet
    all_students_for_ranking = []
    values_data = []  # For hidden _values sheet: {学号, 姓名, 班级, 学分绩点}

    for class_name, rows in class_groups.items():
        ws = wb.create_sheet(title=str(class_name)[:31])

        # Determine which courses this class actually has
        class_courses = []
        for cn in course_order:
            for r in rows:
                if cn in r and r[cn] is not None and str(r[cn]).strip() not in ('', 'nan', 'None'):
                    if cn not in class_courses:
                        class_courses.append(cn)

        # If no data in course_order, use all courses that appear
        if not class_courses:
            class_courses = [c for c in course_order if any(
                c in r for r in rows
            )]

        # Detect which courses use five-level grading (have grade text like 优/良/中/及格/不及格)
        five_level_courses = set()
        for cn in class_courses:
            for r in rows:
                raw = r.get(cn)
                if raw is not None:
                    raw_str = str(raw).strip()
                    if raw_str in SCORE_MAPPING:
                        five_level_courses.add(cn)
                        break

        # Build interleaved column list: for five-level courses, add a conversion column
        # Example: [数学, 体育, 体育(转化), 英语]
        class_cols = []
        for cn in class_courses:
            class_cols.append(cn)
            if cn in five_level_courses:
                class_cols.append(cn + '(转化)')

        # Build header row
        all_headers = info_cols + class_cols + [gpa_col_name]
        num_cols = len(all_headers)

        # Pre-compute course info lookup
        course_info_map = {}  # {course_name: {credit, is_pe}}
        for cn in class_courses:
            for hdr_str, info in all_course_headers.items():
                if info.get('course_name', '') == cn:
                    course_info_map[cn] = {
                        'credit': info.get('credit', 0.0),
                        'is_pe': info.get('is_pe', False),
                    }
                    break
            if cn not in course_info_map:
                course_info_map[cn] = {'credit': 0.0, 'is_pe': False}

        # ---- Row 1: Headers ----
        for ci, hdr in enumerate(all_headers):
            cell = ws.cell(row=1, column=ci + 1, value=hdr)
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border

        # ---- Row 2: Credit row ----
        # - Primary columns: always show actual credit (PE included)
        # - Conversion columns: 0 (credit displayed on primary, not here)
        # GPA formula: (score * credit) sum / total_credits, PE excluded
        credit_values = {}     # {col_idx: value for row 2}
        gpa_terms = []         # [(score_col_idx, credit_col_idx)] for formula — only non-PE

        for ci, hdr in enumerate(all_headers):
            col_idx = ci + 1
            if hdr in info_cols or hdr == gpa_col_name:
                credit_values[col_idx] = None
                continue

            is_conversion = hdr.endswith('(转化)')
            base_course = hdr.replace('(转化)', '') if is_conversion else hdr
            cinfo = course_info_map.get(base_course, {'credit': 0.0, 'is_pe': False})

            if is_conversion:
                # Conversion column: credit = 0 (shown on primary)
                credit_values[col_idx] = 0
                if not cinfo['is_pe'] and base_course in five_level_courses:
                    # Non-PE five-level: formula uses conv_col{row} * primary_col$2
                    primary_idx = None
                    for pci, ph in enumerate(all_headers):
                        if ph == base_course:
                            primary_idx = pci + 1
                            break
                    if primary_idx:
                        gpa_terms.append((col_idx, primary_idx))
            else:
                # Primary column: show actual credit
                credit_values[col_idx] = cinfo['credit']
                if not cinfo['is_pe'] and base_course not in five_level_courses:
                    # Non-PE numeric: formula uses col{row} * col$2
                    gpa_terms.append((col_idx, col_idx))

        for col_idx, credit_val in credit_values.items():
            if credit_val is None:
                cell = ws.cell(row=2, column=col_idx, value='')
            else:
                cell = ws.cell(row=2, column=col_idx, value=credit_val)
                cell.font = Font(name='SimSun', size=10, italic=True)
            cell.alignment = center_align
            cell.border = thin_border

        # ---- Student rows (start at row 3) ----
        for ri, row_dict in enumerate(rows):
            excel_row = ri + 3

            gpa_numeric = 0.0
            total_credit_for_gpa = 0.0

            # Info columns
            for ci, hdr in enumerate(info_cols):
                if hdr == '课程门数':
                    try:
                        val = int(float(row_dict.get(hdr, 0)))
                    except (ValueError, TypeError):
                        val = len(class_courses)
                    cell = ws.cell(row=excel_row, column=ci + 1, value=val)
                elif hdr == '总学分':
                    total_cred = 0.0
                    for c_hdr in class_courses:
                        raw = row_dict.get(c_hdr)
                        sv = _parse_score(raw)
                        cinfo = course_info_map.get(c_hdr, {'credit': 0.0, 'is_pe': False})
                        if sv is not None and not cinfo['is_pe']:
                            total_cred += cinfo['credit']
                    cell = ws.cell(row=excel_row, column=ci + 1, value=total_cred)
                else:
                    val = row_dict.get(hdr, '')
                    cell = ws.cell(row=excel_row, column=ci + 1, value=val)
                cell.font = data_font
                cell.alignment = center_align
                cell.border = thin_border
                if hdr == '学号':
                    cell.number_format = '@'

            # Track which courses this student has scores for (to build formula)
            student_has_score = set()

            # Course columns
            for ci, hdr in enumerate(class_cols):
                col_idx = ci + len(info_cols) + 1
                is_conversion = hdr.endswith('(转化)')
                base_course = hdr.replace('(转化)', '') if is_conversion else hdr
                cinfo = course_info_map.get(base_course, {'credit': 0.0, 'is_pe': False})

                raw_val = row_dict.get(base_course)
                score_val = _parse_score(raw_val)
                raw_str = str(raw_val).strip() if raw_val is not None else ''

                if is_conversion:
                    # Only write value if there IS a score (avoid #VALUE! in formula)
                    if score_val is not None:
                        cell = ws.cell(row=excel_row, column=col_idx, value=score_val)
                    else:
                        cell = ws.cell(row=excel_row, column=col_idx)
                    cell.font = Font(name='SimSun', size=10, color='555555')
                else:
                    is_five_level = raw_str in SCORE_MAPPING
                    if is_five_level:
                        display_val = raw_str
                    elif score_val is not None:
                        display_val = score_val
                    else:
                        display_val = raw_val if raw_val not in ('', 'nan', 'None') else None
                    cell = ws.cell(row=excel_row, column=col_idx, value=display_val)
                    cell.font = data_font

                cell.alignment = center_align
                cell.border = thin_border

                # A five-level course has both an original grade column and a
                # numeric conversion column in the main table. Count the
                # underlying course once so ranking uses the same GPA as the
                # visible Excel formula.
                if (score_val is not None and not cinfo['is_pe'] and
                        cinfo['credit'] > 0 and base_course not in student_has_score):
                    total_credit_for_gpa += cinfo['credit']
                    gpa_numeric += score_val * cinfo['credit']
                    student_has_score.add(base_course)

            # GPA formula per student: only courses with actual scores
            # Five-level: score from conversion, credit from primary
            # Numeric: both from same column
            # PE / no-score courses: excluded
            gpa_col_idx = num_cols
            total_credits_letter = get_column_letter(len(info_cols))

            parts = []
            for score_col, credit_col in gpa_terms:
                # Find which course this term belongs to
                score_hdr = all_headers[score_col - 1]
                base = score_hdr.replace('(转化)', '') if score_hdr.endswith('(转化)') else score_hdr
                if base not in student_has_score:
                    continue  # skip — no score for this course
                sl = get_column_letter(score_col)
                cl = get_column_letter(credit_col)
                parts.append(f'{sl}{excel_row}*{cl}$2')

            if parts:
                formula_str = '=(' + '+'.join(parts) + f')/{total_credits_letter}{excel_row}'
            else:
                formula_str = '0'

            cell = ws.cell(row=excel_row, column=gpa_col_idx, value=formula_str)
            cell.font = formula_font
            cell.alignment = center_align
            cell.border = thin_border
            cell.number_format = '0.00'

            numeric_gpa = gpa_numeric / total_credit_for_gpa if total_credit_for_gpa > 0 else 0.0
            all_students_for_ranking.append({
                '学号': row_dict['学号'],
                '姓名': row_dict['姓名'],
                '学生行政班级': class_name,
                'class_name': class_name,
                '学分绩点': round(numeric_gpa, 6),
            })
            values_data.append({
                '学号': row_dict['学号'],
                '姓名': row_dict['姓名'],
                '班级': class_name,
                '学分绩点': round(numeric_gpa, 6),
            })

        # Column widths
        for ci in range(1, num_cols + 1):
            ws.column_dimensions[get_column_letter(ci)].width = 12

        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 16
        last_col = get_column_letter(num_cols)
        ws.column_dimensions[last_col].width = 14

        # Freeze header (rows 1-2 frozen, data starts at A3)
        ws.freeze_panes = 'A3'

    # --- Write hidden _values sheet for downstream Module D ---
    write_values_sheet(wb, values_data, '学分绩点')

    # Save GPA output
    wb.save(output1_path)
    wb.close()

    progress.update(85, '正在计算专业排名...')

    # ---- Phase 5: Ranking ----
    program_groups = group_by_program_grade(all_students_for_ranking)
    ranking_data = {}
    for pg_key, pg_students in program_groups.items():
        ranked = calculate_ranking(pg_students, '学分绩点', desc=True)
        ranking_data[pg_key] = ranked

    output2_path = unique_path(os.path.join(output_dir, '学分绩点专业排名百分比.xlsx'))

    wb2 = openpyxl.Workbook()
    wb2.remove(wb2.active)

    rank_columns = ['学号', '姓名', '学分绩点', '专业排名', '百分比']
    # Map keys: calculate_ranking sets '排名' and '百分比', but our header is '专业排名'
    rank_key_map = {'学号': '学号', '姓名': '姓名', '学分绩点': '学分绩点',
                    '专业排名': '排名', '百分比': '百分比'}
    for pg_key, ranked_students in ranking_data.items():
        ws = wb2.create_sheet(title=str(pg_key)[:31])
        for ci, hdr in enumerate(rank_columns, 1):
            cell = ws.cell(row=1, column=ci, value=hdr)
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border

        for ri, s in enumerate(ranked_students):
            for ci, col_name in enumerate(rank_columns, 1):
                val = s.get(rank_key_map.get(col_name, col_name), '')
                cell = ws.cell(row=ri + 2, column=ci, value=val)
                cell.font = data_font
                cell.alignment = center_align
                cell.border = thin_border
                if col_name == '百分比':
                    cell.number_format = '0.00%'
                elif col_name == '学分绩点':
                    cell.number_format = '0.00'

        for ci in range(1, 6):
            ws.column_dimensions[get_column_letter(ci)].width = [14, 10, 12, 10, 10][ci - 1]
        ws.freeze_panes = 'A2'

    wb2.save(output2_path)
    wb2.close()

    progress.done('学分绩点计算完成！')

    return {
        'success': True,
        'file_count': total_files,
        'student_count': len(all_students_for_ranking),
        'class_count': len(class_groups),
        'program_count': len(ranking_data),
        'output1': output1_path,
        'output2': output2_path,
    }


# ============================================================
# Helpers
# ============================================================

def _gpa_source_jobs(input_paths: list, column_mappings: dict) -> list:
    """Expand every enabled worksheet into a processing job."""
    jobs = []
    for path in input_paths:
        mapping = column_mappings.get(path, {}) if isinstance(column_mappings, dict) else {}
        if isinstance(mapping, dict) and any(key.endswith('_col') for key in mapping):
            jobs.append((path, None, mapping))
            continue
        enabled = [(name, value) for name, value in (mapping.items() if isinstance(mapping, dict) else [])
                   if isinstance(value, dict) and value.get('enabled', True)]
        if enabled:
            jobs.extend((path, name, value) for name, value in enabled)
        else:
            jobs.append((path, None, {}))
    return jobs

def _selected_mapping_for_file(column_mappings: dict, path: str):
    mapping = column_mappings.get(path, {}) if isinstance(column_mappings, dict) else {}
    if not isinstance(mapping, dict):
        return None, {}
    if any(key.endswith('_col') for key in mapping):
        return None, mapping
    for sheet_name, sheet_mapping in mapping.items():
        if isinstance(sheet_mapping, dict) and sheet_mapping.get('enabled', True):
            return sheet_name, sheet_mapping
    return None, {}


def _read_gpa_source(path: str, sheet_name, mapping: dict) -> pd.DataFrame:
    """Read the exact worksheet/header row confirmed in Import Studio."""
    if sheet_name is None and not mapping:
        return read_raw_xls(path)
    header_row = mapping.get('header_row', 0)
    if not isinstance(header_row, int) or header_row < 0:
        header_row = 0
    engine = 'xlrd' if os.path.splitext(path)[1].lower() == '.xls' else 'openpyxl'
    return pd.read_excel(path, sheet_name=sheet_name or 0,
                         header=header_row, engine=engine)


def _identify_columns(columns: list, mapping: dict = None) -> dict:
    """Identify column roles from header names.

    Standard source format (29 columns):
      Col 0: 学号
      Col 1: 姓名
      Col 2: 学生行政班级
      Col 3: 课程门数
      Col 4-20: Course columns (17 courses)
      Col 21: 不及格门数
      Col 22: 平均分
      Col 23: 总分
      Col 24: 总学分
      Col 25: 获得学分/所得学分
      Col 26: 平均学分绩
      Col 27: 平均学分绩点
      Col 28: 排名

    Returns dict with actual column name strings.
    """
    total = len(columns)

    mapping = mapping or {}
    result = {
        'student_id_col': columns[0] if total > 0 else None,
        'name_col': columns[1] if total > 1 else None,
        'class_col': columns[2] if total > 2 else None,
        'course_count_col': columns[3] if total > 3 else None,
        'course_cols': [],
        'course_defs': [],
    }

    # Course columns: from index 4 to (total - 8)
    # Last 8 columns are: 不及格门数, 平均分, 总分, 总学分, 获得学分, 平均学分绩, 平均学分绩点, 排名
    field_map = {
        'id_col': 'student_id_col', 'name_col': 'name_col',
        'class_col': 'class_col', 'course_count_col': 'course_count_col',
    }
    for source_field, target_field in field_map.items():
        idx = mapping.get(source_field)
        if isinstance(idx, int) and 0 <= idx < total:
            result[target_field] = columns[idx]

    configured_defs = mapping.get('course_definitions')
    if isinstance(configured_defs, list) and configured_defs:
        for item in configured_defs:
            if not isinstance(item, dict) or not item.get('enabled', True):
                continue
            score_idx = item.get('score_col')
            if not isinstance(score_idx, int) or not (0 <= score_idx < total):
                continue
            course_def = dict(item)
            result['course_defs'].append(course_def)
            result['course_cols'].append(columns[score_idx])
        return result

    course_start = mapping.get('course_start_col', 4)
    if not isinstance(course_start, int) or course_start < 0 or course_start >= total:
        course_start = 4
    course_end = max(course_start, total - 8)
    reserved = {result['student_id_col'], result['name_col'], result['class_col'], result['course_count_col']}
    for i in range(course_start, course_end):
        col_str = str(columns[i]).strip()
        if columns[i] not in reserved and col_str and col_str != 'nan' and 'Unnamed' not in col_str:
            result['course_cols'].append(columns[i])
            parsed = parse_course_header(col_str)
            result['course_defs'].append({
                'name': parsed.get('course_name') or col_str,
                'score_col': i, 'value_col': None,
                'credit': parsed.get('credit', 0),
                'is_pe': parsed.get('is_pe', False), 'enabled': True,
            })

    return result


def _safe_str(val) -> str:
    """Safely convert a value to string, cleaning float student IDs."""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ''
    # Clean float student IDs: 221352107.0 → 221352107
    if isinstance(val, float) and val == int(val) and val > 0:
        return str(int(val))
    s = str(val).strip()
    # Also handle "221352107.0" string format
    if s.endswith('.0') and s[:-2].isdigit() and len(s[:-2]) >= 6:
        s = s[:-2]
    return s


def _parse_score(value) -> float | None:
    """Parse a score value. Converts grades (优/良/中/及格/合格) to numeric.
    Returns None only for truly empty/missing values.
    """
    if value is None:
        return None
    if isinstance(value, (int, float)):
        if pd.isna(value):
            return None
        return float(value)

    val_str = str(value).strip()
    if not val_str or val_str.lower() in ('nan', 'none', ''):
        return None

    # Try direct numeric
    try:
        return float(val_str)
    except ValueError:
        pass

    # Grade mapping
    return SCORE_MAPPING.get(val_str, None)
