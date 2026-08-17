"""Annual GPA and comprehensive ranking consolidation.

The annual workspace accepts the two semester workbooks produced by the app.
GPA is credit-weighted when both semester credit totals are available; legacy
tables without credit totals fall back to an arithmetic mean. Comprehensive
scores always use the arithmetic mean of the available semesters.
"""

from __future__ import annotations

import os
import re
import zipfile
from collections import defaultdict

import openpyxl
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

from backend.utils.class_utils import filter_students_by_exact_program, group_by_program_grade
from backend.utils.rank_calculator import calculate_ranking
from backend.utils.excel_writer import unique_path
from backend.utils.progress_reporter import ProgressReporter


_ID_HEADERS = ('学号', '学生学号', '学生号')
_NAME_HEADERS = ('姓名', '学生姓名', '名字')
_CLASS_HEADERS = ('班级', '行政班级', '学生行政班级', '班别')
_GPA_HEADERS = ('学分绩点', '平均学分绩点', '平均学分绩', '绩点', '成绩')
_CREDIT_HEADERS = ('总学分', '获得学分', '所得学分', '学分合计')
_COMP_HEADERS = ('综合测评', '综测', '综合测评分', '综测分')
_MAX_XLSX_BYTES = 128 * 1024 * 1024
_MAX_XLSX_UNCOMPRESSED_BYTES = 1024 * 1024 * 1024
_MAX_XLSX_COMPRESSION_RATIO = 2000
_MAX_WORKSHEETS = 128
_MAX_ROWS_PER_SHEET = 200_000
_MAX_COLUMNS_PER_SHEET = 512
_MAX_CELLS = 5_000_000


def process_annual_gpa(
    semester1_path: str,
    semester2_path: str,
    output_dir: str,
    academic_year: str = '',
    major_filter: str = '',
    progress: ProgressReporter | None = None,
) -> dict:
    """Merge two semester GPA workbooks and create class/program rankings."""
    return _process_annual(
        'gpa', semester1_path, semester2_path, output_dir,
        academic_year=academic_year, major_filter=major_filter, progress=progress,
    )


def process_annual_comprehensive(
    semester1_path: str,
    semester2_path: str,
    output_dir: str,
    academic_year: str = '',
    major_filter: str = '',
    progress: ProgressReporter | None = None,
) -> dict:
    """Merge two semester comprehensive workbooks and create two rankings."""
    return _process_annual(
        'comprehensive', semester1_path, semester2_path, output_dir,
        academic_year=academic_year, major_filter=major_filter, progress=progress,
    )


def _process_annual(kind: str, semester1_path: str, semester2_path: str,
                     output_dir: str, academic_year: str = '',
                     major_filter: str = '',
                     progress: ProgressReporter | None = None) -> dict:
    progress = progress or ProgressReporter()
    _validate_inputs(semester1_path, semester2_path, output_dir)
    year = _normalise_academic_year(academic_year, semester1_path, semester2_path)
    score_label = '学分绩点' if kind == 'gpa' else '综合测评'

    progress.update(8, '正在读取第一学期数据...')
    first = _extract_semester_records(semester1_path, kind)
    progress.update(32, '正在读取第二学期数据...')
    second = _extract_semester_records(semester2_path, kind)
    if not first and not second:
        raise ValueError(f'两份文件中都没有识别到有效的{score_label}数据')

    progress.update(55, f'正在合并学年{score_label}...')
    students, diagnostics = _merge_semesters(first, second, kind)
    if not students:
        raise ValueError(f'没有可用于学年{score_label}排名的学生')
    if str(major_filter or '').strip():
        students = filter_students_by_exact_program(students, major_filter)
        if not students:
            raise ValueError(
                f'当前专业“{major_filter}”在两学期文件中没有匹配学生；'
                '班级应采用“顿河信241”这类完整名称，请检查专业设置与班级列'
            )
        diagnostics = _merge_diagnostics(
            first, second, kind, {student['学号'] for student in students},
        )

    class_groups = defaultdict(list)
    for student in students:
        class_groups[student['班级'] or '未识别班级'].append(student)
    program_groups = group_by_program_grade(students)

    progress.update(72, '正在计算班级与专业排名...')
    class_rankings = _rank_groups(class_groups, score_label)
    program_rankings = _rank_groups(program_groups, score_label)

    os.makedirs(output_dir, exist_ok=True)
    prefix = f'{year}学年' if year else '学年'
    if kind == 'gpa':
        class_name = f'{prefix}绩点班级排名百分比.xlsx'
        program_name = f'{prefix}学分绩点百分比.xlsx'
    else:
        class_name = f'{prefix}综测班级排名百分比.xlsx'
        program_name = f'{prefix}综测百分比.xlsx'
    class_path = unique_path(os.path.join(output_dir, class_name))
    program_path = unique_path(os.path.join(output_dir, program_name))

    progress.update(82, '正在生成班级排名表...')
    _write_ranking_workbook(class_path, class_rankings, kind, scope='class')
    progress.update(93, '正在生成专业排名表...')
    _write_ranking_workbook(program_path, program_rankings, kind, scope='program')
    progress.done('学年排名计算完成！')

    return {
        'success': True,
        'kind': kind,
        'academic_year': year,
        'student_count': len(students),
        'class_count': len(class_rankings),
        'program_count': len(program_rankings),
        'matched_count': diagnostics['matched_count'],
        'first_only_count': diagnostics['first_only_count'],
        'second_only_count': diagnostics['second_only_count'],
        'credit_weighted_count': diagnostics['credit_weighted_count'],
        'mean_fallback_count': diagnostics['mean_fallback_count'],
        'class_mismatch_count': diagnostics['class_mismatch_count'],
        'major_filter': str(major_filter or '').strip(),
        'output1': class_path,
        'output2': program_path,
    }


def _validate_inputs(first: str, second: str, output_dir: str) -> None:
    if not first or not os.path.isfile(first):
        raise ValueError('请选择有效的第一学期 Excel 文件')
    if not second or not os.path.isfile(second):
        raise ValueError('请选择有效的第二学期 Excel 文件')
    if os.path.abspath(first) == os.path.abspath(second):
        raise ValueError('第一学期和第二学期不能选择同一个文件')
    if any(os.path.splitext(path)[1].lower() != '.xlsx' for path in (first, second)):
        raise ValueError('学年汇总仅支持 .xlsx 文件，请先将旧版 .xls 另存为 .xlsx')
    if not output_dir:
        raise ValueError('请选择输出目录')


def _normalise_academic_year(value: str, *paths: str) -> str:
    candidates = [str(value or '')] + [os.path.basename(path) for path in paths]
    for candidate in candidates:
        match = re.search(r'(20\d{2})\s*[-—至]\s*(20\d{2})', candidate)
        if match and int(match.group(2)) == int(match.group(1)) + 1:
            return f'{match.group(1)}-{match.group(2)}'
    if str(value or '').strip():
        raise ValueError('学年格式应为“2024-2025”')
    return ''


def _clean_header(value) -> str:
    return re.sub(r'[\s\n\r（）()：:]+', '', str(value or '')).casefold()


def _header_index(headers: list, aliases: tuple[str, ...]) -> int | None:
    clean = [_clean_header(value) for value in headers]
    wanted = [_clean_header(value) for value in aliases]
    for index, value in enumerate(clean):
        if value in wanted:
            return index
    for index, value in enumerate(clean):
        if any(alias and alias in value for alias in wanted):
            return index
    return None


def _clean_student_id(value) -> str:
    if value is None or isinstance(value, bool):
        return ''
    if isinstance(value, (int, float)):
        if isinstance(value, float) and not value.is_integer():
            return ''
        value = str(int(value))
    text = str(value).strip()
    if text.endswith('.0'):
        text = text[:-2]
    return text if re.fullmatch(r'\d{6,20}', text) else ''


def _number(value) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number if number == number else None


def _safe_excel_text(value) -> str:
    """Keep workbook-sourced text from being interpreted as a formula."""
    text = str(value or '')
    return "'" + text if text.startswith(('=', '+', '-', '@')) else text


def _validate_workbook_archive(path: str) -> None:
    if os.path.getsize(path) > _MAX_XLSX_BYTES:
        raise ValueError('Excel 文件过大，已停止处理')
    try:
        with zipfile.ZipFile(path) as archive:
            total = 0
            for info in archive.infolist():
                total += info.file_size
                if total > _MAX_XLSX_UNCOMPRESSED_BYTES:
                    raise ValueError('Excel 文件解压规模超限，已停止处理')
                if info.file_size and (
                    not info.compress_size
                    or info.file_size / info.compress_size > _MAX_XLSX_COMPRESSION_RATIO
                ):
                    raise ValueError('Excel 文件压缩比例异常，已停止处理')
    except zipfile.BadZipFile as exc:
        raise ValueError('Excel 文件损坏或格式不正确') from exc


def _validate_workbook_shape(workbook) -> None:
    if len(workbook.worksheets) > _MAX_WORKSHEETS:
        raise ValueError('Excel 工作表数量超限，已停止处理')
    total_cells = 0
    for ws in workbook.worksheets:
        rows, columns = int(ws.max_row or 0), int(ws.max_column or 0)
        if rows > _MAX_ROWS_PER_SHEET or columns > _MAX_COLUMNS_PER_SHEET:
            raise ValueError(f'工作表“{ws.title}”规模超限，已停止处理')
        total_cells += rows * columns
        if total_cells > _MAX_CELLS:
            raise ValueError('Excel 单元格总量超限，已停止处理')


def _find_header(ws, score_aliases: tuple[str, ...]) -> tuple[int, list] | None:
    for row_number, values in enumerate(
        ws.iter_rows(min_row=1, max_row=min(ws.max_row, 12), values_only=True), 1
    ):
        headers = list(values)
        has_identity = _header_index(headers, _ID_HEADERS) is not None
        has_score = _header_index(headers, score_aliases) is not None
        if has_identity and has_score:
            return row_number, headers
    return None


def _read_values_index(workbook, score_aliases: tuple[str, ...]) -> dict:
    if '_values' not in workbook.sheetnames:
        return {}
    ws = workbook['_values']
    found = _find_header(ws, score_aliases)
    if not found:
        return {}
    header_row, headers = found
    id_col = _header_index(headers, _ID_HEADERS)
    name_col = _header_index(headers, _NAME_HEADERS)
    class_col = _header_index(headers, _CLASS_HEADERS)
    score_col = _header_index(headers, score_aliases)
    result = {}
    for values in ws.iter_rows(min_row=header_row + 1, values_only=True):
        sid = _clean_student_id(values[id_col] if id_col < len(values) else None)
        if not sid:
            continue
        result[sid] = {
            'id': sid,
            'name': str(values[name_col] or '').strip()
                    if name_col is not None and name_col < len(values) else '',
            'class': str(values[class_col] or '').strip()
                     if class_col is not None and class_col < len(values) else '',
            'score': _number(values[score_col] if score_col < len(values) else None),
            'credits': None,
        }
    return result


def _extract_semester_records(path: str, kind: str) -> dict:
    score_aliases = _GPA_HEADERS if kind == 'gpa' else _COMP_HEADERS
    _validate_workbook_archive(path)
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        _validate_workbook_shape(workbook)
    except Exception:
        workbook.close()
        raise
    values_index = _read_values_index(workbook, score_aliases)
    records = {}

    for ws in workbook.worksheets:
        if ws.title == '_values':
            continue
        found = _find_header(ws, score_aliases)
        if not found:
            continue
        header_row, headers = found
        id_col = _header_index(headers, _ID_HEADERS)
        name_col = _header_index(headers, _NAME_HEADERS)
        class_col = _header_index(headers, _CLASS_HEADERS)
        score_col = _header_index(headers, score_aliases)
        credit_col = _header_index(headers, _CREDIT_HEADERS) if kind == 'gpa' else None
        for values in ws.iter_rows(min_row=header_row + 1, values_only=True):
            sid = _clean_student_id(values[id_col] if id_col < len(values) else None)
            if not sid:
                continue
            cached = values_index.get(sid, {})
            score = _number(values[score_col] if score_col < len(values) else None)
            if score is None:
                score = cached.get('score')
            if score is None:
                continue
            explicit_class = (str(values[class_col] or '').strip()
                              if class_col is not None and class_col < len(values) else '')
            class_name = explicit_class or cached.get('class') or ws.title
            name = (str(values[name_col] or '').strip()
                    if name_col is not None and name_col < len(values) else '') or cached.get('name', '')
            credits = (_number(values[credit_col] if credit_col < len(values) else None)
                       if credit_col is not None else None)
            candidate = {'id': sid, 'name': name, 'class': class_name,
                         'score': score, 'credits': credits}
            existing = records.get(sid)
            if existing is None or (candidate['credits'] and not existing.get('credits')):
                records[sid] = candidate

    for sid, cached in values_index.items():
        if cached.get('score') is not None and sid not in records:
            records[sid] = dict(cached)
    workbook.close()
    return records


def _merge_diagnostics(first: dict, second: dict, kind: str,
                       student_ids: set[str]) -> dict:
    diagnostics = {
        'matched_count': 0, 'first_only_count': 0, 'second_only_count': 0,
        'credit_weighted_count': 0, 'mean_fallback_count': 0,
        'class_mismatch_count': 0,
    }
    for sid in student_ids:
        one, two = first.get(sid), second.get(sid)
        if one and two:
            diagnostics['matched_count'] += 1
        elif one:
            diagnostics['first_only_count'] += 1
        elif two:
            diagnostics['second_only_count'] += 1
        if one and two and one.get('class') and two.get('class') and one['class'] != two['class']:
            diagnostics['class_mismatch_count'] += 1
        available = [item for item in (one, two) if item and item.get('score') is not None]
        if kind == 'gpa' and len(available) == 2:
            if all((item.get('credits') or 0) > 0 for item in available):
                diagnostics['credit_weighted_count'] += 1
            else:
                diagnostics['mean_fallback_count'] += 1
    return diagnostics


def _merge_semesters(first: dict, second: dict, kind: str) -> tuple[list, dict]:
    merged = []
    for sid in sorted(set(first) | set(second)):
        one, two = first.get(sid), second.get(sid)
        available = [item for item in (one, two) if item and item.get('score') is not None]
        if not available:
            continue
        if kind == 'gpa' and len(available) == 2 and all((item.get('credits') or 0) > 0 for item in available):
            total_credits = sum(item['credits'] for item in available)
            score = sum(item['score'] * item['credits'] for item in available) / total_credits
        else:
            score = sum(item['score'] for item in available) / len(available)

        latest = two or one
        earlier = one or two
        class_name = str(latest.get('class') or earlier.get('class') or '').strip()
        name = str(latest.get('name') or earlier.get('name') or '').strip()
        merged.append({
            '学号': sid,
            '姓名': name,
            '班级': class_name,
            'class_name': class_name,
            ('学分绩点' if kind == 'gpa' else '综合测评'): round(score, 2),
        })
    return merged, _merge_diagnostics(first, second, kind, set(first) | set(second))


def _rank_groups(groups: dict, score_label: str) -> dict:
    ranked = {}
    for group_name in sorted(groups, key=str):
        rows = [dict(row) for row in groups[group_name]]
        ranked[group_name] = calculate_ranking(rows, score_label, desc=True)
    return ranked


def _short_class_sheet_name(class_name: str) -> str:
    return str(class_name or '未识别班级').strip()


def _safe_sheet_name(value: str, used: set[str]) -> str:
    base = re.sub(r'[\\/*?:\[\]]', '_', str(value or '未命名'))[:31] or '未命名'
    candidate = base
    index = 2
    while candidate in used:
        suffix = f'({index})'
        candidate = f'{base[:31-len(suffix)]}{suffix}'
        index += 1
    used.add(candidate)
    return candidate


def _write_ranking_workbook(path: str, groups: dict, kind: str, scope: str) -> None:
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    used_names = set()
    thin = Side(style='thin', color='B0B0B0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(name='SimSun', size=10, bold=True)
    data_font = Font(name='SimSun', size=10)
    center = Alignment(horizontal='center', vertical='center')

    if kind == 'gpa':
        rank_header = '班级排名' if scope == 'class' else '专业排名'
        columns = ['学号', '班级', '姓名', '学分绩点', rank_header, '绩点百分比']
        score_key, percentage_header = '学分绩点', '绩点百分比'
    else:
        rank_header = '班级排名' if scope == 'class' else '排名'
        columns = ['班级', '学号', '姓名', '综合测评', rank_header, '百分比']
        score_key, percentage_header = '综合测评', '百分比'

    for group_name, students in groups.items():
        visible_name = _short_class_sheet_name(group_name) if scope == 'class' else str(group_name)
        ws = workbook.create_sheet(_safe_sheet_name(visible_name, used_names))
        for col, header in enumerate(columns, 1):
            cell = ws.cell(1, col, header)
            cell.font = header_font
            cell.alignment = center
            cell.border = border
        for row_index, student in enumerate(students, 2):
            values = {
                '学号': student.get('学号', ''),
                '班级': _safe_excel_text(student.get('班级', '')),
                '姓名': _safe_excel_text(student.get('姓名', '')),
                score_key: student.get(score_key, 0),
                rank_header: student.get('排名', 0),
                percentage_header: student.get('百分比', 0),
            }
            for col, header in enumerate(columns, 1):
                value = values.get(header, '')
                if header == '学号':
                    value = str(value)
                cell = ws.cell(row_index, col, value)
                cell.font = data_font
                cell.alignment = center
                cell.border = border
                if header == '学号':
                    cell.number_format = '@'
                elif header == score_key:
                    cell.number_format = '0.00'
                elif header == percentage_header:
                    cell.number_format = '0.00%'
                elif header == rank_header:
                    cell.number_format = '0'
        widths = [14, 16, 12, 13, 12, 13] if kind == 'gpa' else [16, 14, 12, 13, 12, 13]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f'A1:F{max(1, ws.max_row)}'
    workbook.save(path)
    workbook.close()
