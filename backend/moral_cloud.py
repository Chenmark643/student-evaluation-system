"""Build one cloud-ready moral workbook from mixed A/B local outputs."""

from __future__ import annotations

import os
import re
import tempfile
from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter

from backend.utils.class_utils import parse_class_name


def _text(value) -> str:
    return re.sub(r"\s+", "", str(value or ""))


def _class_sheet_name(value) -> str:
    name = str(value or "").strip()
    return name[:31] if parse_class_name(name).get("grade") else ""


def _header_info(sheet) -> tuple[int, int | None]:
    for row_index in range(1, min(sheet.max_row, 20) + 1):
        headers = [_text(sheet.cell(row_index, col).value) for col in range(1, sheet.max_column + 1)]
        if not any("姓名" in header for header in headers):
            continue
        class_col = next(
            (index for index, header in enumerate(headers, 1) if header in {"班级", "行政班级", "班级名称"}),
            None,
        )
        return row_index, class_col
    return 1, None


def _translated_value(source_cell, destination_coordinate: str):
    value = source_cell.value
    if source_cell.data_type == "f" and isinstance(value, str):
        try:
            return Translator(value, origin=source_cell.coordinate).translate_formula(destination_coordinate)
        except (TypeError, ValueError):
            return value
    return value


def _copy_cell(source_cell, destination_cell) -> None:
    destination_cell.value = _translated_value(source_cell, destination_cell.coordinate)
    if source_cell.has_style:
        destination_cell._style = copy(source_cell._style)
    if source_cell.number_format:
        destination_cell.number_format = source_cell.number_format
    destination_cell.font = copy(source_cell.font)
    destination_cell.fill = copy(source_cell.fill)
    destination_cell.border = copy(source_cell.border)
    destination_cell.alignment = copy(source_cell.alignment)
    destination_cell.protection = copy(source_cell.protection)
    if source_cell.comment:
        destination_cell.comment = copy(source_cell.comment)
    if source_cell.hyperlink:
        destination_cell._hyperlink = copy(source_cell.hyperlink)


def _copy_dimensions(source, destination, max_col: int) -> None:
    for col_index in range(1, max_col + 1):
        letter = get_column_letter(col_index)
        destination.column_dimensions[letter] = copy(source.column_dimensions[letter])
    destination.sheet_format = copy(source.sheet_format)
    destination.sheet_properties = copy(source.sheet_properties)
    destination.sheet_view.showGridLines = source.sheet_view.showGridLines
    destination.freeze_panes = source.freeze_panes
    destination.page_margins = copy(source.page_margins)
    destination.page_setup = copy(source.page_setup)
    destination.print_options = copy(source.print_options)


def _copy_rows(source, destination, source_rows: list[int], header_row: int) -> None:
    max_col = source.max_column
    destination_row = 1
    row_map = {}
    for source_row in source_rows:
        row_map[source_row] = destination_row
        for col_index in range(1, max_col + 1):
            _copy_cell(
                source.cell(source_row, col_index),
                destination.cell(destination_row, col_index),
            )
        if source_row in source.row_dimensions:
            destination.row_dimensions[destination_row] = copy(source.row_dimensions[source_row])
        destination_row += 1
    _copy_dimensions(source, destination, max_col)
    for merged in source.merged_cells.ranges:
        if merged.max_row <= header_row and all(row in row_map for row in range(merged.min_row, merged.max_row + 1)):
            destination.merge_cells(
                start_row=row_map[merged.min_row],
                start_column=merged.min_col,
                end_row=row_map[merged.max_row],
                end_column=merged.max_col,
            )
    if source.auto_filter.ref:
        destination.auto_filter.ref = (
            f"A{row_map.get(header_row, header_row)}:"
            f"{get_column_letter(max_col)}{destination.max_row}"
        )


def prepare_moral_cloud_bundle(paths: list[str]) -> dict:
    """Merge visible class sheets; split aggregate A sheets by their class column."""
    source_paths = []
    for raw_path in paths or []:
        path = Path(str(raw_path or "")).expanduser().resolve()
        if path.is_file() and path.suffix.lower() in {".xlsx", ".xlsm"} and path not in source_paths:
            source_paths.append(path)
    if not source_paths:
        raise ValueError("没有可用于云端汇总的德育结果文件")

    bundle = openpyxl.Workbook()
    bundle.remove(bundle.active)
    skipped = []
    source_by_class = {}
    for path in source_paths:
        workbook = openpyxl.load_workbook(path, data_only=False, read_only=False)
        try:
            for sheet in workbook.worksheets:
                if sheet.sheet_state != "visible":
                    continue
                direct_class = _class_sheet_name(sheet.title)
                header_row, class_col = _header_info(sheet)
                groups = {}
                if direct_class:
                    groups[direct_class] = list(range(1, sheet.max_row + 1))
                elif class_col:
                    for row_index in range(header_row + 1, sheet.max_row + 1):
                        class_name = _class_sheet_name(sheet.cell(row_index, class_col).value)
                        if class_name:
                            groups.setdefault(class_name, list(range(1, header_row + 1))).append(row_index)
                if not groups:
                    skipped.append(f"{path.name} / {sheet.title}")
                    continue
                for class_name, rows in groups.items():
                    if class_name in bundle.sheetnames:
                        del bundle[class_name]
                    destination = bundle.create_sheet(class_name)
                    _copy_rows(sheet, destination, rows, header_row)
                    source_by_class[class_name] = path.name
        finally:
            workbook.close()

    if not bundle.sheetnames:
        bundle.close()
        raise ValueError("这些文件中没有识别到按班级命名的工作表或“班级”列")

    bundle.active = 0
    bundle.calculation.fullCalcOnLoad = True
    bundle.calculation.forceFullCalc = True
    bundle.calculation.calcMode = "auto"
    output_dir = Path(tempfile.mkdtemp(prefix="doncollege-moral-cloud-"))
    output_path = output_dir / "德育云端汇总-待同步.xlsx"
    bundle.save(output_path)
    bundle.close()
    return {
        "success": True,
        "output": os.fspath(output_path),
        "source_count": len(source_paths),
        "class_count": len(source_by_class),
        "classes": sorted(source_by_class),
        "source_by_class": source_by_class,
        "skipped": skipped,
    }
