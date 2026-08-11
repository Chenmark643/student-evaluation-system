"""Secretary toolbox: deterministic applicant qualification audit."""
from __future__ import annotations
import os, re
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from config import SCORE_MAPPING


def _norm(v): return re.sub(r'\s+', '', str(v or '')).casefold()
def _text(v): return '' if pd.isna(v) else str(v).strip()
def _num(v):
    if pd.isna(v) or _text(v) == '': return None
    try: return float(v)
    except (TypeError, ValueError): return SCORE_MAPPING.get(_text(v))


def _sheets(path):
    if not path or not os.path.isfile(path): return {}
    engine = 'xlrd' if os.path.splitext(path)[1].lower() == '.xls' else 'openpyxl'
    xl = pd.ExcelFile(path, engine=engine)
    try: return {n: pd.read_excel(xl, sheet_name=n, header=None) for n in xl.sheet_names}
    finally: xl.close()


def _header(df, keywords):
    best, score = 0, -1
    for i in range(min(15, len(df))):
        vals = [_norm(v) for v in df.iloc[i].tolist()]
        current = sum(1 for words in keywords.values() if any(any(_norm(w) in v for w in words) for v in vals))
        if current > score: best, score = i, current
    headers = [_text(v) for v in df.iloc[best].tolist()]
    fields = {}
    for key, words in keywords.items():
        fields[key] = next((i for i,h in enumerate(headers) if h and any(_norm(w) in _norm(h) for w in words)), None)
    return best, headers, fields


IDENTITY = {'id':['学号','学生号'], 'name':['姓名'], 'class':['班级','行政班']}


def _read_applicants(path):
    result = []
    for sheet, df in _sheets(path).items():
        schema={**IDENTITY,'declared_gpa':['学分绩点','平均学分绩点'],'declared_comp':['综合测评','综测'],'declared_quality':['素拓分','素质拓展分'],'declared_work':['学生工作测评','工作测评']}
        hi, _, f = _header(df, schema)
        if f['name'] is None: continue
        for _, row in df.iloc[hi+1:].iterrows():
            name = _text(row.iloc[f['name']])
            sid = _text(row.iloc[f['id']]) if f['id'] is not None else ''
            cls = _text(row.iloc[f['class']]) if f['class'] is not None else ''
            if name and name not in ('姓名','合计'):
                result.append({'id':sid, 'name':name, 'class_name':cls, 'source_sheet':sheet,
                    'declared':{key.replace('declared_',''):_text(row.iloc[idx]) if idx is not None else '' for key,idx in f.items() if key.startswith('declared_')}})
    return result


def _key(sid, name, cls=''):
    return ('id', _norm(sid)) if sid and len(re.sub(r'\D','',sid)) >= 6 else ('name', _norm(name), _norm(cls))


def _read_scores(path):
    records = {}
    schema = {**IDENTITY, 'gpa':['学分绩点','平均学分绩点'], 'rank':['专业排名','班级排名','排名'], 'percent':['百分比','排名比例']}
    stat_words = ('学分绩点','排名','百分比','平均分','总分','总学分','获得学分','不及格门数')
    for sheet, df in _sheets(path).items():
        hi, headers, f = _header(df, schema)
        if f['name'] is None: continue
        credit_row = hi+1 if hi+1 < len(df) and '学分' in _text(df.iloc[hi+1,0]) else None
        start = 4 if len(headers)>4 else len(headers)
        course_indices = []
        for i in range(start, len(headers)):
            if headers[i] and any(w in headers[i].replace(' ','') for w in stat_words): break
            if headers[i]: course_indices.append(i)
        for _, row in df.iloc[hi+1+(1 if credit_row is not None else 0):].iterrows():
            name = _text(row.iloc[f['name']]); sid = _text(row.iloc[f['id']]) if f['id'] is not None else ''
            cls = _text(row.iloc[f['class']]) if f['class'] is not None else sheet
            if not name: continue
            courses=[]
            for i in course_indices:
                score=_num(row.iloc[i])
                if score is not None: courses.append({'name':headers[i], 'score':score, 'is_pe':'体育' in headers[i]})
            rec={'id':sid,'name':name,'class_name':cls,'courses':courses}
            for field in ('gpa','rank','percent'):
                if f[field] is not None: rec[field]=_num(row.iloc[f[field]])
            records[_key(sid,name,cls)] = rec
            if sid: records[('name',_norm(name),_norm(cls))]=rec
    return records


def _read_simple_metric(path, metric_words):
    records={}; schema={**IDENTITY,'value':metric_words,'rank':['专业排名','班级排名','排名'],'percent':['百分比','排名比例']}
    for sheet,df in _sheets(path).items():
        hi,_,f=_header(df,schema)
        if f['name'] is None: continue
        sheet_records=[]
        for _,row in df.iloc[hi+1:].iterrows():
            name=_text(row.iloc[f['name']]); sid=_text(row.iloc[f['id']]) if f['id'] is not None else ''; cls=_text(row.iloc[f['class']]) if f['class'] is not None else sheet
            if not name: continue
            rec={'id':sid,'name':name,'class_name':cls,'value':_num(row.iloc[f['value']]) if f['value'] is not None else None,'rank':_num(row.iloc[f['rank']]) if f['rank'] is not None else None,'percent':_num(row.iloc[f['percent']]) if f['percent'] is not None else None}
            sheet_records.append(rec)
            records[_key(sid,name,cls)]=rec
            if sid: records[('name',_norm(name),_norm(cls))]=rec
        total=len(sheet_records)
        for rec in sheet_records:
            if rec['percent'] is None and rec['rank'] is not None and total:
                rec['percent']=round(rec['rank']/total*100,2)
    return records


def _read_work(paths):
    """Read textual or fill-color based work evaluations.

    Official sheets use yellow (#FFFF00) for 优秀 and green (#92D050) for 良好.
    Formatting is intentionally read with openpyxl because pandas discards styles.
    """
    by_name={}; skipped={'姓名','团委','学生会','自律委员会','主席团','比重','部门','班级'}
    for path in paths or []:
        if not path or not os.path.isfile(path): continue
        wb=openpyxl.load_workbook(path,data_only=True,read_only=False)
        try:
            for ws in wb.worksheets:
                for row in ws.iter_rows():
                    vals=[_text(cell.value) for cell in row]
                    if not vals: continue
                    # Both supplied formats place the person's name in column A.
                    name=vals[0].strip() if vals else ''
                    if not name or name in skipped or name.endswith(('委员会','中心')): continue
                    explicit='优秀' if '优秀' in vals[2:] else ('良好' if '良好' in vals[2:] else '')
                    color_grade=''
                    for cell in row:
                        if cell.fill.fill_type != 'solid': continue
                        color=cell.fill.fgColor
                        rgb=(color.rgb or '')[-6:].upper() if color.type == 'rgb' else ''
                        if rgb == 'FFFF00': color_grade='优秀'; break
                        if rgb == '92D050' and not color_grade: color_grade='良好'
                    grade=explicit or color_grade
                    if grade:
                        by_name.setdefault(_norm(name),[]).append({'grade':grade,'sheet':ws.title,
                            'department':vals[1] if len(vals)>1 else '',
                            'position':vals[2] if len(vals)>2 else '',
                            'source':os.path.basename(path),'detected_by':'文字' if explicit else '底色'})
        finally:
            wb.close()
    return by_name


def _lookup(dataset, applicant):
    return dataset.get(_key(applicant['id'],applicant['name'],applicant['class_name'])) or dataset.get(('name',_norm(applicant['name']),_norm(applicant['class_name'])))


def audit_applicants(config: dict) -> dict:
    applicants=_read_applicants(config.get('applicant_file',''))
    if not applicants: return {'success':False,'error':'申报名单中未识别到姓名'}
    scores=_read_scores(config.get('score_file',''))
    academic_class=_read_simple_metric(config.get('class_ranking_file',''),['学分绩点','平均学分绩点'])
    academic_major=_read_simple_metric(config.get('major_ranking_file',''),['学分绩点','平均学分绩点'])
    comp_class=_read_simple_metric(config.get('class_comp_file',''),['综合测评','综测'])
    comp_major=_read_simple_metric(config.get('major_comp_file',''),['综合测评','综测'])
    quality=_read_simple_metric(config.get('quality_file',''),['素拓分','素质拓展分','拓展分'])
    work=_read_work(config.get('work_files',[])); rules=config.get('rules',{})
    results=[]
    def pct_ok(value, threshold):
        if value is None: return False
        normalized=value*100 if 0 <= value <= 1 else value
        return normalized <= threshold
    for a in applicants:
        s=_lookup(scores,a); acr=_lookup(academic_class,a); amr=_lookup(academic_major,a); ccr=_lookup(comp_class,a); cmr=_lookup(comp_major,a); qr=_lookup(quality,a); wr=work.get(_norm(a['name']),[])
        courses=(s or {}).get('courses',[]); failed=[c for c in courses if c['score']<60]; below75=[c for c in courses if c['score']<75]
        non_pe=[c for c in courses if not c['is_pe']]; pe=[c for c in courses if c['is_pe']]
        values={'gpa':(acr or amr or s or {}).get('value',(s or {}).get('gpa')),'academic_class_percent':(acr or {}).get('percent'),'academic_major_percent':(amr or {}).get('percent'),'comp_class_percent':(ccr or {}).get('percent'),'comp_major_percent':(cmr or {}).get('percent'),'quality':(qr or {}).get('value'),'work_grade':wr[0]['grade'] if len(wr)==1 else ('待确认' if len(wr)>1 else ''),'failed_count':len(failed),'below75_count':len(below75),'min_course':min([c['score'] for c in non_pe],default=None),'pe_score':min([c['score'] for c in pe],default=None)}
        checks=[]
        def add(key,label,actual,passed,detail=''):
            if rules.get(key,{}).get('enabled',False): checks.append({'key':key,'label':label,'actual':actual,'passed':passed,'detail':detail})
        add('no_fail','无挂科',values['failed_count'],None if not s else values['failed_count']==0,'、'.join(c['name'] for c in failed))
        add('no_below75','无低于75课程',values['below75_count'],None if not s else values['below75_count']==0,'、'.join(f"{c['name']} {c['score']:g}" for c in below75))
        t=rules.get('gpa',{}).get('threshold',75); add('gpa',f'学分绩点≥{t}',values['gpa'],None if values['gpa'] is None else values['gpa']>=t)
        t=rules.get('academic_class_rank',{}).get('threshold',25); add('academic_class_rank',f'班级成绩前{t}%',values['academic_class_percent'],None if values['academic_class_percent'] is None else pct_ok(values['academic_class_percent'],t))
        t=rules.get('academic_major_rank',{}).get('threshold',25); add('academic_major_rank',f'专业成绩前{t}%',values['academic_major_percent'],None if values['academic_major_percent'] is None else pct_ok(values['academic_major_percent'],t))
        t=rules.get('comp_class_rank',{}).get('threshold',30); add('comp_class_rank',f'班级综测前{t}%',values['comp_class_percent'],None if values['comp_class_percent'] is None else pct_ok(values['comp_class_percent'],t))
        t=rules.get('comp_major_rank',{}).get('threshold',30); add('comp_major_rank',f'专业综测前{t}%',values['comp_major_percent'],None if values['comp_major_percent'] is None else pct_ok(values['comp_major_percent'],t))
        allowed=rules.get('work',{}).get('allowed',['优秀','良好']); add('work','学生工作测评',values['work_grade'],None if not values['work_grade'] or values['work_grade']=='待确认' else values['work_grade'] in allowed)
        add('quality','存在素拓',values['quality'],None if values['quality'] is None else values['quality']>0)
        t=rules.get('min_course',{}).get('threshold',70); add('min_course',f'单科≥{t}',values['min_course'],None if values['min_course'] is None else values['min_course']>=t)
        t=rules.get('pe',{}).get('threshold',75); add('pe',f'体育≥{t}',values['pe_score'],None if values['pe_score'] is None else values['pe_score']>=t)
        declared=a.get('declared',{})
        comparisons=[]
        for field,label,actual in [('gpa','申报表学分绩点',values['gpa']),('quality','申报表素拓分',values['quality']),('work','申报表工作测评',values['work_grade'])]:
            expected=declared.get(field,'')
            if expected!='':
                if field=='work': same=_norm(expected)==_norm(actual)
                else:
                    try: same=actual is not None and abs(float(expected)-float(actual))<0.011
                    except (TypeError,ValueError): same=False
                comparisons.append({'key':'correspondence_'+field,'label':label+'对应','actual':actual,'expected':expected,'passed':same,'detail':f'申报表：{expected}；源文件：{actual if actual is not None else "缺失"}'})
        checks.extend(comparisons)
        status='待确认' if not s or any(c['passed'] is None for c in checks) else ('通过' if all(c['passed'] for c in checks) else '不通过')
        results.append({**a,'status':status,'values':values,'checks':checks,'matched':{'成绩':bool(s),'班级成绩排名':bool(acr),'专业成绩排名':bool(amr),'班级综测':bool(ccr),'专业综测':bool(cmr),'素拓':bool(qr),'工作测评':bool(wr)}})
    return {'success':True,'students':results,'counts':{k:sum(1 for r in results if r['status']==k) for k in ('通过','不通过','待确认')}}


def export_audit_report(students: list, output_path: str) -> dict:
    wb=openpyxl.Workbook(); ws=wb.active; ws.title='资格审核'
    rule_labels=[]
    for student in students:
        for check in student.get('checks',[]):
            if check.get('label') not in rule_labels: rule_labels.append(check.get('label'))
    headers=['选择','学号','姓名','班级','结论']+rule_labels+['异常详情']
    ws.append(headers)
    for student in students:
        checks={c.get('label'):c for c in student.get('checks',[])}
        details=[]
        row=['是',student.get('id',''),student.get('name',''),student.get('class_name',''),student.get('status','')]
        for label in rule_labels:
            check=checks.get(label)
            row.append('未审核' if not check else ('待确认' if check.get('passed') is None else ('通过' if check.get('passed') else '不通过')))
            if check and not check.get('passed') and check.get('detail'): details.append(f"{label}：{check['detail']}")
        row.append('；'.join(details)); ws.append(row)
    for cell in ws[1]:
        cell.font=Font(bold=True,color='FFFFFF'); cell.fill=PatternFill('solid',fgColor='0B6F6B'); cell.alignment=Alignment(horizontal='center')
    for row in ws.iter_rows(min_row=2):
        status=row[4]
        color='DFF3E8' if status.value=='通过' else ('FBE1E1' if status.value=='不通过' else 'FFF1D5')
        status.fill=PatternFill('solid',fgColor=color)
    ws.freeze_panes='A2'; ws.auto_filter.ref=ws.dimensions
    widths=[8,15,12,14,10]+[15]*len(rule_labels)+[45]
    for i,width in enumerate(widths,1): ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width=width
    os.makedirs(os.path.dirname(output_path),exist_ok=True); wb.save(output_path); wb.close()
    return {'success':True,'output':output_path,'student_count':len(students)}
