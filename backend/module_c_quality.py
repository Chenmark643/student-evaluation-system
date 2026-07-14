from __future__ import annotations
"""
Module C: Quality Development Score Calculation (素质拓展分计算)

Interactive form-based scoring with activity memory and threshold capping.
v2.1: Thresholds now link to bonus categories. Score merging keeps original
individual scores — only the total gets capped (no proportional scaling).
"""

import os
import json
from copy import deepcopy

from backend.utils.excel_writer import write_multi_sheet_xlsx, write_values_sheet
from backend.utils.progress_reporter import ProgressReporter
from config import (
    ACTIVITY_MAPPINGS_FILE, QUALITY_CATEGORIES, QUALITY_GRADES,
)
from backend.quality_presets import (
    OFFICIAL_THRESHOLDS, extract_user_mappings, merge_official_with_user,
)

# Custom thresholds file
CUSTOM_THRESHOLDS_FILE = os.path.join(os.path.dirname(ACTIVITY_MAPPINGS_FILE),
                                       'custom_thresholds.json')


# ============================================================
# Activity Mappings CRUD
# ============================================================

def load_user_activity_mappings() -> dict:
    if os.path.exists(ACTIVITY_MAPPINGS_FILE):
        try:
            with open(ACTIVITY_MAPPINGS_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except (json.JSONDecodeError, IOError):
            pass
    return {}


def load_activity_mappings() -> dict:
    """Return the official catalog overlaid with user-created/edited rows."""
    return merge_official_with_user(load_user_activity_mappings())


def save_activity_mappings(mappings: dict):
    os.makedirs(os.path.dirname(ACTIVITY_MAPPINGS_FILE), exist_ok=True)
    with open(ACTIVITY_MAPPINGS_FILE, 'w', encoding='utf-8') as f:
        json.dump(extract_user_mappings(mappings), f, ensure_ascii=False, indent=2)


def get_activity_suggestion(activity_name: str) -> dict | None:
    mappings = load_activity_mappings()
    if activity_name in mappings:
        return mappings[activity_name]
    return None


def record_activity(activity_name: str, category: str, grade: str, score: float):
    mappings = load_user_activity_mappings()
    mappings[activity_name] = {
        'category': category, 'default_grade': grade,
        'default_score': score, 'last_used': '',
    }
    save_activity_mappings(mappings)


def add_activity_mapping(name: str, category: str, grade: str, score: float) -> dict:
    mappings = load_user_activity_mappings()
    mappings[name] = {
        'category': category, 'default_grade': grade,
        'default_score': score, 'last_used': '',
    }
    save_activity_mappings(mappings)
    return load_activity_mappings()


def delete_activity_mapping(name: str) -> dict:
    mappings = load_user_activity_mappings()
    if name in mappings:
        del mappings[name]
        save_activity_mappings(mappings)
    return load_activity_mappings()


CUSTOM_CATEGORIES_FILE = os.path.join(os.path.dirname(ACTIVITY_MAPPINGS_FILE),
                                       'custom_categories.json')


def get_categories() -> list:
    cats = list(QUALITY_CATEGORIES)
    if os.path.exists(CUSTOM_CATEGORIES_FILE):
        try:
            with open(CUSTOM_CATEGORIES_FILE, 'r', encoding='utf-8') as f:
                for c in json.load(f):
                    if c not in cats:
                        cats.append(c)
        except Exception:
            pass
    return cats


def add_custom_category(name: str) -> list:
    cats = get_categories()
    if name not in cats:
        custom = []
        if os.path.exists(CUSTOM_CATEGORIES_FILE):
            try:
                with open(CUSTOM_CATEGORIES_FILE, 'r', encoding='utf-8') as f:
                    custom = json.load(f)
            except Exception:
                pass
        custom.append(name)
        os.makedirs(os.path.dirname(CUSTOM_CATEGORIES_FILE), exist_ok=True)
        with open(CUSTOM_CATEGORIES_FILE, 'w', encoding='utf-8') as f:
            json.dump(custom, f, ensure_ascii=False, indent=2)
        cats.append(name)
    return cats


def remove_custom_category(name: str) -> list:
    if name in QUALITY_CATEGORIES:
        return get_categories()
    if os.path.exists(CUSTOM_CATEGORIES_FILE):
        try:
            with open(CUSTOM_CATEGORIES_FILE, 'r', encoding='utf-8') as f:
                custom = json.load(f)
            custom = [c for c in custom if c != name]
            with open(CUSTOM_CATEGORIES_FILE, 'w', encoding='utf-8') as f:
                json.dump(custom, f, ensure_ascii=False, indent=2)
        except Exception:
            pass
    return [c for c in get_categories() if c != name]


def get_grades_for_category(category: str) -> list:
    return QUALITY_GRADES.get(category, ['国家级', '省级', '市级', '校级', '院级'])


# ============================================================
# Threshold System (v2.1 — category-linked)
# ============================================================
# Threshold model: [{name, max, categories: [cat_names]}]
# - name: display name (e.g., "志愿类上限", "国家级证书上限")
# - max: maximum score
# - categories: list of bonus categories this threshold applies to
#   For single-category thresholds: categories = ["志愿类"]
#   For group thresholds: categories = ["A类", "B类", "C类", "D类"]

def _build_default_thresholds() -> list:
    """Return an isolated copy so UI edits cannot mutate official rules."""
    return deepcopy(OFFICIAL_THRESHOLDS)


def get_thresholds() -> list:
    """Get all thresholds (default + custom) as a list."""
    thresholds = _build_default_thresholds()

    if os.path.exists(CUSTOM_THRESHOLDS_FILE):
        try:
            with open(CUSTOM_THRESHOLDS_FILE, 'r', encoding='utf-8') as f:
                custom = json.load(f)
                thresholds.extend(custom)
        except (json.JSONDecodeError, IOError):
            pass
    return thresholds


def get_thresholds_dict() -> dict:
    """Get thresholds as a dict keyed by name (for bridge compatibility)."""
    result = {}
    for t in get_thresholds():
        result[t['name']] = {
            'max': t['max'],
            'categories': t['categories'],
        }
    return result


def save_custom_thresholds(thresholds: list):
    """Save custom thresholds (only the non-default ones)."""
    os.makedirs(os.path.dirname(CUSTOM_THRESHOLDS_FILE), exist_ok=True)
    default_names = {t['name'] for t in _build_default_thresholds()}
    custom = [t for t in thresholds if t['name'] not in default_names]
    with open(CUSTOM_THRESHOLDS_FILE, 'w', encoding='utf-8') as f:
        json.dump(custom, f, ensure_ascii=False, indent=2)


def add_custom_threshold(name: str, max_score: float, categories: list,
                          mode: str = 'sum') -> list:
    """Add a custom threshold that applies to specific bonus categories.

    Args:
        name: Display name (e.g., "国家级证书上限")
        max_score: Maximum combined score (absolute ceiling)
        categories: List of bonus category names this threshold covers
        mode: 'sum' (default) = sum all items then cap
              'max_item' = use highest individual score as cap

    Returns:
        Updated threshold list
    """
    thresholds = get_thresholds()
    thresholds.append({
        'name': name,
        'max': float(max_score),
        'categories': list(categories),
        'mode': mode if mode in ('sum', 'max_item') else 'sum',
    })
    save_custom_thresholds(thresholds)
    return thresholds


def remove_custom_threshold(name: str) -> list:
    """Remove a custom threshold by name."""
    default_names = {t['name'] for t in _build_default_thresholds()}
    if name in default_names:
        return get_thresholds()
    thresholds = get_thresholds()
    thresholds = [t for t in thresholds if t['name'] != name]
    save_custom_thresholds(thresholds)
    return thresholds


# ============================================================
# Score Calculation (v2.1 — keep original scores, cap totals only)
# ============================================================

def _build_effective_score_groups(activities: list, thresholds: list) -> list:
    """Group activities by threshold and calculate each group's effective score.

    An activity belongs to at most one threshold (the first matching rule).  This
    keeps overlapping custom rules deterministic and, more importantly, gives
    calculation and Excel export the exact same grouping semantics.
    """
    assigned = set()
    groups = []

    for threshold in thresholds:
        categories = set(threshold.get('categories', []))
        matched = [
            (index, activity)
            for index, activity in enumerate(activities)
            if index not in assigned and activity.get('category', '') in categories
        ]
        if not matched:
            continue

        assigned.update(index for index, _ in matched)
        scores = [float(activity.get('score', 0) or 0) for _, activity in matched]
        raw_score = sum(scores)
        maximum = float(threshold.get('max', float('inf')))
        mode = threshold.get('mode', 'sum')
        if mode == 'max_item':
            effective_score = min(max(scores, default=0.0), maximum)
        else:
            effective_score = min(raw_score, maximum)

        groups.append({
            'first_index': matched[0][0],
            'activities': [activity for _, activity in matched],
            'threshold': threshold.get('name', ''),
            'categories': list(threshold.get('categories', [])),
            'mode': mode,
            'raw': round(raw_score, 2),
            'effective': round(max(0.0, effective_score), 2),
            'is_threshold': True,
        })

    for index, activity in enumerate(activities):
        if index in assigned:
            continue
        score = round(float(activity.get('score', 0) or 0), 2)
        groups.append({
            'first_index': index,
            'activities': [activity],
            'threshold': '',
            'categories': [activity.get('category', '')],
            'mode': 'sum',
            'raw': score,
            'effective': score,
            'is_threshold': False,
        })

    groups.sort(key=lambda group: group['first_index'])
    return groups

def calculate_quality_scores(
    student_activities: dict,
    thresholds: list = None,
) -> dict:
    """Calculate quality development scores with threshold capping.

    v2.1: Individual activity scores are kept as-is.
    Only the TOTAL is capped when a threshold is exceeded.
    No proportional scaling of individual scores.

    Args:
        student_activities: {student_key: [{activity, category, grade, score}]}
        thresholds: List of threshold dicts [{name, max, categories}]

    Returns:
        {student_key: {activities, total, category_totals, capped_details}}
    """
    if thresholds is None:
        thresholds = get_thresholds()

    results = {}
    for student_key, activities in student_activities.items():
        # Sum by individual bonus category
        category_totals = {}
        for act in activities:
            cat = act.get('category', '')
            score = act.get('score', 0)
            category_totals[cat] = category_totals.get(cat, 0.0) + score

        score_groups = _build_effective_score_groups(activities, thresholds)
        capped_details = []
        for group in score_groups:
            if group['is_threshold'] and group['raw'] > group['effective']:
                capped_details.append({
                    'threshold': group['threshold'],
                    'categories': group['categories'],
                    'mode': group['mode'],
                    'raw': group['raw'],
                    'capped': group['effective'],
                    'excess': round(group['raw'] - group['effective'], 2),
                })

        final_total = round(sum(group['effective'] for group in score_groups), 2)

        results[student_key] = {
            'activities': activities,
            'total': final_total,
            'category_totals': {k: round(v, 2) for k, v in category_totals.items()},
            'capped_details': capped_details,
        }

    return results


# ============================================================
# Export with Merged Cells (v2.1 — no proportional scaling)
# ============================================================

def export_quality_merged(
    roster: dict,
    quality_data: dict,
    output_path: str,
    thresholds: list = None,
) -> dict:
    """Export quality scores with merged cells and formulas.

    v2.1: Individual 加分 scores are kept as original values.
    The 拓展分 column shows the capped total (just the number, not a formula).
    The cap is noted in the sheet for transparency.

    Output format per sheet:
      Row 1: Title "XX班拓展分加分统计"
      Row 2: Headers (学号 姓名 加分项目 等级 加分 拓展分)
      Rows 3+: Data with merged cells, original scores, capped total
      Last row: Total merged
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    if thresholds is None:
        thresholds = get_thresholds()

    class_students = {}
    for sid, info in roster.items():
        cls = info.get('class', '其他')
        if cls not in class_students:
            class_students[cls] = []
        class_students[cls].append(sid)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    total_student_count = 0
    values_data = []  # For hidden _values sheet
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'))

    for class_name in sorted(class_students.keys()):
        ws = wb.create_sheet(title=str(class_name)[:31])
        sids = class_students[class_name]
        total_student_count += len(sids)

        # Row 1: Title
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
        title_cell = ws.cell(row=1, column=1,
                             value=f'{class_name}班拓展分加分统计')
        title_cell.font = Font(name='SimSun', size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')

        # Row 2: Headers
        headers = ['学号', '姓名', '加分项目', '等级', '加分', '拓展分']
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=2, column=ci, value=h)
            cell.font = Font(name='SimSun', size=10, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        current_row = 3
        total_start_row = current_row

        for sid in sids:
            info = roster.get(sid, {})
            name = info.get('name', '')
            activities = quality_data.get(sid, [])

            if not activities:
                for ci, val in enumerate([sid, name, '', '', 0, 0], 1):
                    cell = ws.cell(row=current_row, column=ci, value=val)
                    cell.font = Font(name='SimSun', size=10)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border
                    if ci == 1:  # 学号 column
                        cell.number_format = '@'
                current_row += 1
                values_data.append({
                    '学号': sid, '姓名': name, '班级': class_name, '拓展分': 0.0,
                })
                continue

            # Calculate once using the same threshold groups that are rendered.
            score_groups = _build_effective_score_groups(activities, thresholds)
            capped_total = round(sum(group['effective'] for group in score_groups), 2)
            values_data.append({
                '学号': sid, '姓名': name, '班级': class_name,
                '拓展分': capped_total,
            })

            start_row = current_row
            # Write each threshold as one score group.  Multiple categories covered
            # by the same cap are merged together and show one effective value.
            for group in score_groups:
                acts = group['activities']
                group_start = current_row

                for ai, act in enumerate(acts):
                    row = current_row + ai
                    ws.cell(row=row, column=3, value=act.get('activity', ''))
                    ws.cell(row=row, column=4, value=act.get('grade', ''))
                    ws.cell(row=row, column=5, value=0 if group['is_threshold'] else group['effective'])
                    for ci in range(3, 6):
                        c = ws.cell(row=row, column=ci)
                        c.font = Font(name='SimSun', size=10)
                        c.alignment = Alignment(horizontal='center', vertical='center')
                        c.border = thin_border

                group_end = current_row + len(acts) - 1
                current_row = group_end + 1

                if group['is_threshold']:
                    if len(acts) > 1:
                        ws.merge_cells(start_row=group_start, start_column=5,
                                       end_row=group_end, end_column=5)
                    ws.cell(row=group_start, column=5, value=group['effective'])
                    ws.cell(row=group_start, column=5).font = Font(name='SimSun', size=10)
                    ws.cell(row=group_start, column=5).alignment = Alignment(horizontal='center', vertical='center')
                    ws.cell(row=group_start, column=5).border = thin_border

            end_row = current_row - 1
            total_rows = end_row - start_row + 1

            # Merge student ID / name / 拓展分
            if total_rows > 1:
                ws.merge_cells(start_row=start_row, start_column=1,
                               end_row=end_row, end_column=1)
                ws.merge_cells(start_row=start_row, start_column=2,
                               end_row=end_row, end_column=2)
                ws.merge_cells(start_row=start_row, start_column=6,
                               end_row=end_row, end_column=6)

            c_sid = ws.cell(row=start_row, column=1, value=sid)
            c_sid.number_format = '@'
            ws.cell(row=start_row, column=2, value=name)

            # Always store the effective value so downstream files do not depend
            # on Excel recalculating formulas after export.
            ws.cell(row=start_row, column=6, value=capped_total)

            for ci in [1, 2, 6]:
                c = ws.cell(row=start_row, column=ci)
                c.font = Font(name='SimSun', size=10)
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.border = thin_border

        # Total row
        total_end_row = current_row - 1
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=4)
        total_cell = ws.cell(row=current_row, column=1, value='合计')
        total_cell.font = Font(name='SimSun', size=10, bold=True)
        total_cell.alignment = Alignment(horizontal='center', vertical='center')

        ws.cell(row=current_row, column=5,
                value=f'=SUM({get_column_letter(5)}{total_start_row}:{get_column_letter(5)}{total_end_row})')
        ws.cell(row=current_row, column=6,
                value=f'=SUM({get_column_letter(6)}{total_start_row}:{get_column_letter(6)}{total_end_row})')

        for ci in range(1, 7):
            c = ws.cell(row=current_row, column=ci)
            c.font = Font(name='SimSun', size=10, bold=True)
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = thin_border

        for ci, w in enumerate([14, 10, 20, 12, 8, 10], 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

        ws.freeze_panes = 'A3'

    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else '.', exist_ok=True)

    # Write hidden _values sheet for downstream Module D
    if values_data:
        write_values_sheet(wb, values_data, '拓展分')

    wb.save(output_path)
    wb.close()

    return {
        'success': True,
        'student_count': total_student_count,
        'class_count': len(class_students),
        'output': output_path,
    }


def export_quality_scores(
    class_data: dict,
    output_path: str,
    thresholds: list = None,
    progress: ProgressReporter = None,
) -> dict:
    if progress is None:
        progress = ProgressReporter()

    progress.update(10, '正在计算素拓分数...')
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    sheets_data = {}
    total_students = 0

    for class_name, students in class_data.items():
        sheet_rows = []
        for student_key, data in students.items():
            name = data.get('name', student_key)
            activities = data.get('activities', [])

            scored = calculate_quality_scores({student_key: activities}, thresholds)
            total = scored.get(student_key, {}).get('total', 0.0)

            if activities:
                for act in activities:
                    sheet_rows.append({
                        '学号': student_key,
                        '姓名': name,
                        '加分项目': act.get('activity', ''),
                        '等级': act.get('grade', ''),
                        '加分': act.get('score', 0),
                        '拓展分': total,
                    })
            else:
                sheet_rows.append({
                    '学号': student_key, '姓名': name,
                    '加分项目': '', '等级': '', '加分': 0, '拓展分': 0.0,
                })

        sheets_data[class_name] = sheet_rows
        total_students += len(students)

    progress.update(70, f'正在导出 {len(sheets_data)} 个班级的数据...')
    write_multi_sheet_xlsx(output_path, sheets_data, title='素质拓展分统计', auto_filter=True)
    progress.done('素拓分数导出完成！')

    return {
        'success': True,
        'student_count': total_students,
        'class_count': len(sheets_data),
        'output': output_path,
    }
