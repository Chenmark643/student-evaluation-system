"""
Module B: Moral Education Score (德育分计算)

Uses GPA output as roster (学号/姓名/班级).
Processes multiple input files per category with correct deduction signs.
Supports file preview for column mapping selection.
Outputs per-class sheets with Excel formulas.
"""

import os
import re
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from backend.parsers.xls_reader import read_xlsx_sheets
from backend.utils.progress_reporter import ProgressReporter
from backend.utils.excel_writer import write_values_sheet, unique_path


# Output column headers
MORAL_HEADERS = [
    '学号', '姓名',
    '基础分', '评议分', '晚寝负责人',
    '早晚自习出勤', '课堂出勤', '出勤总',
    '宿舍卫生', '教室卫生', '卫生总',
    '团课出勤', '青年大学习', '通报批评', '违纪情况',
    '德育分',
]

IDX = {h: i for i, h in enumerate(MORAL_HEADERS)}
BASE_SCORE = 80


def process_moral_education(
    roster_path: str,
    absence_files: list = None,
    class_absence_files: list = None,
    dormitory_files: list = None,
    classroom_files: list = None,
    org_class_files: list = None,
    review_scores: dict = None,
    output_dir: str = None,
    progress: ProgressReporter = None,
    column_mappings: dict = None,
    manual_scores: dict = None,
    selected_columns: list = None,
    grade_filter: str = 'all',
) -> dict:
    """Process moral education scores.

    Args:
        roster_path: Path to GPA output (学分绩点.xlsx)
        absence_files: 早晚自习 absence files
        class_absence_files: 课堂出勤 absence files
        dormitory_files: 宿舍卫生 files
        classroom_files: 教室卫生 files
        org_class_files: 团课出勤 files
        review_scores: Manual 评议分 per student ID
        output_dir: Output directory
        progress: ProgressReporter
        column_mappings: Dict of {filepath: {sheet: {field: col_index}}}
        manual_scores: Dict of {student_id: {field_name: value}}
        selected_columns: List of column headers to output
        grade_filter: 'all' or 'XX级' to filter output
    """
    if progress is None:
        progress = ProgressReporter()

    # Reset MORAL_HEADERS to default (in case custom headers were added in previous run)
    global MORAL_HEADERS, IDX
    MORAL_HEADERS = [
        '学号', '姓名',
        '基础分', '评议分', '晚寝负责人',
        '早晚自习出勤', '课堂出勤', '出勤总',
        '宿舍卫生', '教室卫生', '卫生总',
        '团课出勤', '青年大学习', '通报批评', '违纪情况',
        '德育分',
    ]
    IDX = {h: i for i, h in enumerate(MORAL_HEADERS)}

    # --- Phase 1: Load roster ---
    progress.update(5, '正在从学分绩点表加载花名册...')
    roster = _load_roster(roster_path)
    if not roster:
        raise ValueError('无法从学分绩点文件读取花名册，请检查文件。')

    # Initialize student data
    students = {}
    for sid, info in roster.items():
        students[sid] = {
            '学号': sid,
            '姓名': info['name'],
            '班级': info['class'],
            '基础分': BASE_SCORE,
            '评议分': review_scores.get(sid, 0) if review_scores else 0,
            '晚寝负责人': 0,
            '早晚自习出勤': 0,
            '课堂出勤': 0,
            '宿舍卫生': 0,
            '教室卫生': 0,
            '团课出勤': 0,
            '青年大学习': 0,
            '通报批评': 0,
            '违纪情况': 0,
        }

    # Apply manual scores (including custom fields)
    if manual_scores:
        for sid, fields in manual_scores.items():
            if sid in students:
                for field, val in fields.items():
                    try:
                        fval = float(val)
                        # Add field to student dict even if it's a custom/new field
                        students[sid][field] = fval
                        # Auto-add custom field to MORAL_HEADERS if not already present
                        if field not in MORAL_HEADERS:
                            # Insert before 德育分
                            idx = MORAL_HEADERS.index('德育分') if '德育分' in MORAL_HEADERS else len(MORAL_HEADERS)
                            MORAL_HEADERS.insert(idx, field)
                            IDX[field] = idx
                            # Update IDX for shifted columns
                            for h, i in IDX.items():
                                if i >= idx and h != field:
                                    IDX[h] = i + 1
                    except (ValueError, TypeError):
                        pass

    # Build name+class index for faster lookup
    name_class_index = {}
    for sid, data in students.items():
        key = (data['班级'], data['姓名'].strip())
        name_class_index[key] = sid
        # Also index by name only (for loose matching)
        name_key = data['姓名'].strip()
        if name_key not in name_class_index:
            name_class_index[('', name_key)] = sid

    # --- Phase 2: Process input files ---

    if absence_files:
        progress.update(20, '正在处理早晚自习数据...')
        for fp in absence_files:
            if fp and os.path.exists(fp):
                _process_absence_file_v2(fp, students, name_class_index, '早晚自习出勤',
                                         column_mappings.get(fp, {}) if column_mappings else {})

    if class_absence_files:
        progress.update(35, '正在处理课堂出勤数据...')
        for fp in class_absence_files:
            if fp and os.path.exists(fp):
                _process_absence_file_v2(fp, students, name_class_index, '课堂出勤',
                                         column_mappings.get(fp, {}) if column_mappings else {})

    if dormitory_files:
        progress.update(50, '正在处理宿舍卫生数据...')
        for fp in dormitory_files:
            if fp and os.path.exists(fp):
                _process_hygiene_file_v2(fp, students, name_class_index, '宿舍卫生',
                                         column_mappings.get(fp, {}) if column_mappings else {})

    if classroom_files:
        progress.update(60, '正在处理教室卫生数据...')
        for fp in classroom_files:
            if fp and os.path.exists(fp):
                _process_hygiene_file_v2(fp, students, name_class_index, '教室卫生',
                                         column_mappings.get(fp, {}) if column_mappings else {})

    if org_class_files:
        progress.update(70, '正在处理团课出勤数据...')
        for fp in org_class_files:
            if fp and os.path.exists(fp):
                _process_org_class_file_v2(fp, students, name_class_index, '团课出勤',
                                           column_mappings.get(fp, {}) if column_mappings else {})

    progress.update(80, '正在生成德育分表...')

    # --- Phase 3: Write output ---
    # Determine output columns
    if selected_columns and len(selected_columns) >= 2:
        output_headers = [h for h in selected_columns if h in MORAL_HEADERS]
        if len(output_headers) < 2:
            output_headers = list(MORAL_HEADERS)
    else:
        output_headers = list(MORAL_HEADERS)

    # Filter by grade if specified
    class_groups = {}
    for sid, data in students.items():
        cls = data['班级']
        # Grade filter
        if grade_filter and grade_filter != 'all':
            match = re.search(r'(\d{2})\d{1,2}$', str(cls))
            grade_key = f"{match.group(1)}级" if match else ''
            if grade_key != grade_filter:
                continue
        if cls not in class_groups:
            class_groups[cls] = []
        class_groups[cls].append(data)

    for cls in class_groups:
        class_groups[cls].sort(key=lambda s: str(s['学号']))

    os.makedirs(output_dir, exist_ok=True)
    output_path = unique_path(os.path.join(output_dir, '德育分.xlsx'))

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    values_data = []  # For hidden _values sheet
    for class_name, class_students in class_groups.items():
        ws = wb.create_sheet(title=str(class_name)[:31])
        _write_moral_sheet(ws, class_students, output_headers,
                          values_data=values_data, class_name=class_name)

    # Write hidden _values sheet for downstream Module D
    if values_data:
        write_values_sheet(wb, values_data, '德育分')

    wb.save(output_path)
    wb.close()

    # Count actual students in output (after grade filter)
    total_students = sum(len(v) for v in class_groups.values())

    progress.done('德育分计算完成！')

    return {
        'success': True,
        'student_count': total_students,
        'class_count': len(class_groups),
        'output': output_path,
    }


def _load_roster(roster_path: str) -> dict:
    """Load student roster from GPA output file.

    Reads sheet names as class names AND also tries to extract class
    from column 2 (行政班级) if available.
    """
    roster = {}
    sheets = {}
    try:
        sheets = read_xlsx_sheets(roster_path)
    except Exception as e:
        print(f'[WARN] _load_roster: failed to read {roster_path}: {e}')
        return roster

    if not sheets:
        print(f'[WARN] _load_roster: no sheets found in {roster_path}')
        return roster

    for sheet_name, df in sheets.items():
        if df.empty or len(df.columns) < 2:
            continue

        # Detect if column 2 contains class info (行政班级)
        has_class_col = False
        if len(df.columns) > 2:
            col2_header = str(df.columns[2]).strip() if df.columns[2] is not None else ''
            has_class_col = any(kw in col2_header for kw in ['班级', '行政班', '专业'])

        for idx, row in df.iterrows():
            sid = str(row.iloc[0]).strip() if not pd.isna(row.iloc[0]) else ''
            # Clean float-format IDs
            if sid.endswith('.0') and sid[:-2].isdigit() and len(sid[:-2]) >= 6:
                sid = sid[:-2]
            if not sid or sid in ('学号', '学分', '总学分', 'nan', ''):
                continue
            name = str(row.iloc[1]).strip() if len(df.columns) > 1 and not pd.isna(row.iloc[1]) else ''
            if not name or name in ('nan', 'None', ''):
                continue

            # Determine class: prefer data column over sheet name
            cls = str(sheet_name)
            if has_class_col and len(row) > 2:
                col2_val = str(row.iloc[2]).strip() if not pd.isna(row.iloc[2]) else ''
                if col2_val and col2_val not in ('nan', 'None', ''):
                    cls = col2_val

            roster[sid] = {'name': name, 'class': cls}

    if not roster:
        print(f'[WARN] _load_roster: loaded 0 students from {roster_path}')
    return roster


# ============================================================
# V2 Processors — improved matching with name+class fallback
# ============================================================

def _normalize_name(name: str) -> str:
    """Normalize a name for comparison: strip spaces, unify full/half-width."""
    if not name:
        return ''
    n = str(name).strip()
    # Remove all whitespace
    n = re.sub(r'\s+', '', n)
    # Normalize parentheses
    n = n.replace('（', '(').replace('）', ')')
    return n


def _normalize_class(cls: str) -> str:
    """Normalize class name: strip 班 suffix, unify separators."""
    if not cls:
        return ''
    c = str(cls).strip()
    c = re.sub(r'\s+', '', c)
    c = c.replace('（', '(').replace('）', ')')
    return c


def _find_student_by_any(students: dict, name_class_index: dict,
                         sid: str, name: str, cls: str) -> str | None:
    """Multi-strategy student lookup, now with aggressive name normalization.

    Priority:
    1. Exact student ID match (normalized)
    2. Name + class exact match (normalized)
    3. Name-only match (normalized, collect all candidates)
    4. Name + class fuzzy (core extraction, digit matching)
    5. Name match with digit-group class matching
    """
    # Strategy 1: Exact ID (normalized)
    if sid:
        sid_str = str(sid).strip()
        if sid_str.endswith('.0') and sid_str[:-2].isdigit():
            sid_str = sid_str[:-2]
        if sid_str in students:
            return sid_str
        sid_int = sid_str.lstrip('0')
        for s in students:
            if s.lstrip('0') == sid_int:
                return s

    name_norm = _normalize_name(name) if name else ''

    # Strategy 2: Name + class exact (normalized)
    if name_norm and cls:
        cls_norm = _normalize_class(cls)
        # Try exact match on normalized values
        for (c, n), sid_match in name_class_index.items():
            if _normalize_name(n) == name_norm and _normalize_class(c) == cls_norm:
                return sid_match
        # Try without 班 suffix
        cls_no_ban = re.sub(r'班$', '', cls_norm)
        for (c, n), sid_match in name_class_index.items():
            c_no_ban = re.sub(r'班$', '', _normalize_class(c))
            if _normalize_name(n) == name_norm and c_no_ban == cls_no_ban:
                return sid_match

    # Strategy 3: Name only — collect ALL candidates, try to disambiguate
    if name_norm:
        candidates = []
        for (c, n), sid_match in name_class_index.items():
            if _normalize_name(n) == name_norm:
                candidates.append((c, sid_match))
        if len(candidates) == 1:
            return candidates[0][1]
        if len(candidates) > 1 and cls:
            # Try to narrow down using class
            cls_norm = _normalize_class(cls)
            for c, sid_match in candidates:
                if cls_norm in _normalize_class(c) or _normalize_class(c) in cls_norm:
                    return sid_match

    # Strategy 4: Class digit matching (e.g., 241 matches 241 in different formats)
    if name_norm and cls:
        cls_norm = _normalize_class(cls)
        cls_digits = re.findall(r'\d+', cls_norm)
        for (c, n), sid_match in name_class_index.items():
            if _normalize_name(n) != name_norm:
                continue
            c_norm = _normalize_class(c)
            if c_norm == cls_norm:
                return sid_match
            # Check if digit sequences match
            c_digits = re.findall(r'\d+', c_norm)
            if cls_digits and c_digits and cls_digits == c_digits:
                return sid_match
            # Partial containment
            if len(cls_norm) >= 4 and len(c_norm) >= 4:
                if cls_norm in c_norm or c_norm in cls_norm:
                    return sid_match

    return None


def _process_absence_file_v2(filepath: str, students: dict, name_class_index: dict,
                              field: str, col_mapping: dict = None):
    """Process absence file with improved format detection.

    Handles multiple formats:
    - Format A (旷课汇总): Class header rows interleaved with student data rows
      Headers: None/班级, 姓名, 学号, 班级, 所缺课程, 任课教师, 缺课学时, 合计学时, 缺课日期
    - Format B (信息部晚自习): 班级, 姓名, 合计学时, 晚自习扣分数
    - Format C: Simple name + deduction values
    """
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception:
        return

    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Detect format by analyzing first 10 rows
        fmt = _detect_absence_format(rows[:min(10, len(rows))])

        if fmt == 'summary':  # 旷课统计汇总
            _parse_absence_summary(rows, students, name_class_index, field)
        elif fmt == 'discipline':  # 信息部格式 (班级, 姓名, 合计学时, 扣分数)
            _parse_absence_discipline(rows, students, name_class_index, field)
        elif fmt == 'simple':  # (class, name, hours) or just (name, hours)
            _parse_absence_simple(rows, students, name_class_index, field, str(sn))
        else:
            # Fallback: try all parsers, use the one that matches most
            _parse_absence_generic(rows, students, name_class_index, field, str(sn))

    wb.close()


def _detect_absence_format(rows: list) -> str:
    """Detect the format of an absence file."""
    has_class_col = False
    has_student_id = False
    has_course_name = False
    has_hours_header = False

    for row in rows:
        if not row:
            continue
        cells = [str(c).strip() if c is not None else '' for c in row]
        row_text = ' '.join(cells)

        if '学号' in row_text:
            has_student_id = True
        if '所缺课程' in row_text or '任课教师' in row_text:
            has_course_name = True
        if '合计学时' in row_text or '缺课学时' in row_text:
            has_hours_header = True
        if '班级' in row_text:
            has_class_col = True
        if '扣分数' in row_text or '扣分' in row_text:
            return 'discipline'

    if has_course_name and has_hours_header:
        return 'summary'
    if has_class_col and has_hours_header:
        return 'discipline'
    return 'simple'


def _parse_absence_summary(rows: list, students: dict, name_class_index: dict, field: str):
    """Parse 旷课统计汇总 format.

    Interleaved format:
    - Row: [class_name, None, None, ...] (class header)
    - Row: [None, '姓名', '学号', '班级', '所缺课程', '任课教师', '缺课学时', '合计学时', ...] (sub-header)
    - Row: [None, name, student_id, class, course, teacher, absent_hours, total_hours, ...] (data)
    """
    current_class = ''
    for row in rows:
        if not row:
            continue

        col0 = str(row[0]).strip() if row[0] is not None else ''
        col1 = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ''

        # Skip sub-header rows
        if col1 in ('姓名', '学号', ''):
            # Check if col0 is a class name (has digits)
            if col0 and any(c.isdigit() for c in col0) and '姓名' not in col0 and '学号' not in col0:
                if not col0.isdigit() or len(col0) < 9:
                    current_class = col0
            continue

        # Data row
        name = col1
        student_id = ''
        cls = current_class

        # Find student ID (long digit sequence)
        for i, cell in enumerate(row):
            if cell is None:
                continue
            cell_str = str(cell).strip()
            if cell_str.isdigit() and len(cell_str) >= 9:
                student_id = cell_str
                break

        # Find class name from row (col 3 typically)
        if len(row) > 3 and row[3] is not None:
            row_cls = str(row[3]).strip()
            if row_cls and any(c.isdigit() for c in row_cls):
                cls = row_cls

        # Find hours (缺课学时 typically col 6, 合计学时 col 7)
        hours = 0.0
        for i in [6, 7]:
            if len(row) > i and row[i] is not None:
                try:
                    val = float(row[i])
                    if val > 0:
                        hours = val
                        break
                except (ValueError, TypeError):
                    pass

        if name and name not in ('nan', 'None', ''):
            matched = _find_student_by_any(students, name_class_index, student_id, name, cls)
            if matched:
                deduction = -abs(float(hours) * 2) if hours > 0 else 0
                students[matched][field] += deduction


def _parse_absence_discipline(rows: list, students: dict, name_class_index: dict, field: str):
    """Parse 信息部 format: 班级, 姓名, 合计学时, 晚自习扣分数."""
    # Find header row
    header_idx = -1
    for i, row in enumerate(rows):
        if row and any('姓名' in str(c) for c in row if c):
            header_idx = i
            break
    if header_idx < 0:
        return

    headers = [str(c).strip() if c else '' for c in rows[header_idx]]
    class_col = _find_header_col(headers, ['班级'])
    name_col = _find_header_col(headers, ['姓名'])
    hours_col = _find_header_col(headers, ['合计学时', '学时'])
    deduction_col = _find_header_col(headers, ['扣分数', '扣分'])

    if name_col is None:
        return

    current_class = ''
    for row in rows[header_idx + 1:]:
        if not row:
            continue

        name = str(row[name_col]).strip() if len(row) > name_col and row[name_col] is not None else ''
        if not name or name in ('姓名', 'nan', 'None', ''):
            # Could be a class header
            col0 = str(row[0]).strip() if row[0] is not None else ''
            if col0 and any(c.isdigit() for c in col0) and '姓名' not in col0:
                current_class = col0
            continue

        cls = ''
        if class_col is not None and len(row) > class_col and row[class_col] is not None:
            cls = str(row[class_col]).strip()
        if not cls:
            cls = current_class

        # Try deduction column first, then hours * 2
        deduction = 0.0
        if deduction_col is not None and len(row) > deduction_col and row[deduction_col] is not None:
            try:
                deduction = float(row[deduction_col])
            except (ValueError, TypeError):
                pass
        if deduction == 0.0 and hours_col is not None and len(row) > hours_col and row[hours_col] is not None:
            try:
                hours = float(row[hours_col])
                deduction = hours * 2.0
            except (ValueError, TypeError):
                pass

        if deduction != 0:
            matched = _find_student_by_any(students, name_class_index, None, name, cls)
            if matched:
                students[matched][field] += -abs(deduction)


def _parse_absence_simple(rows: list, students: dict, name_class_index: dict,
                          field: str, sheet_name: str):
    """Parse simple format with class headers and name/value pairs."""
    current_class = sheet_name
    for row in rows:
        if not row:
            continue

        # Check for class header
        col0 = str(row[0]).strip() if row[0] is not None else ''
        if col0 and any(c.isdigit() for c in col0) and len(col0) < 15:
            # Check if it looks like a class name (not a student ID)
            has_non_digit = any(not c.isdigit() for c in col0)
            if has_non_digit:
                current_class = col0
                continue

        # Process of name-value pairs in this row
        for i in range(0, len(row) - 1, 2):
            name_cell = row[i] if i < len(row) else None
            val_cell = row[i + 1] if i + 1 < len(row) else None
            if not name_cell:
                continue
            name = str(name_cell).strip()
            if not name or name in ('姓名', '学号', 'nan', 'None', ''):
                continue
            try:
                val = float(val_cell) if val_cell is not None else 0
            except (ValueError, TypeError):
                continue
            if val != 0:
                matched = _find_student_by_any(students, name_class_index, None, name, current_class)
                if matched:
                    students[matched][field] += -abs(val)


def _parse_absence_generic(rows: list, students: dict, name_class_index: dict,
                           field: str, sheet_name: str):
    """Generic fallback parser - tries to extract any name+number pairs."""
    current_class = sheet_name
    for row in rows:
        if not row:
            continue
        name = None
        hours = 0.0
        student_id = None

        for i, cell in enumerate(row):
            if cell is None:
                continue
            cell_str = str(cell).strip()

            # Detect class header
            if i == 0 and cell_str and any(c.isdigit() for c in cell_str) and len(cell_str) < 15:
                has_alpha = any(not c.isdigit() for c in cell_str)
                if has_alpha:
                    current_class = cell_str
                continue

            # Skip headers
            if cell_str in ('姓名', '学号', '班级', '课程', '合计学时', '缺课学时', '扣分'):
                continue

            # Student ID
            if cell_str.isdigit() and len(cell_str) >= 9:
                student_id = cell_str
            # Name
            elif not cell_str.isdigit() and 2 <= len(cell_str) <= 5 and not any(
                kw in cell_str for kw in ['班级', '课程', '教师', '学时', '日期', '扣分', '统计', '学年']
            ):
                if name is None:
                    name = cell_str
            # Hours
            try:
                val = float(cell_str)
                if 0 < val < 100 and hours == 0:
                    hours = val
            except ValueError:
                pass

        if name or student_id:
            matched = _find_student_by_any(students, name_class_index, student_id, name, current_class)
            if matched and hours > 0:
                deduction = -abs(float(hours) * 2)
                students[matched][field] += deduction


# ============================================================
# V2 Hygiene processor
# ============================================================

def _process_hygiene_file_v2(filepath: str, students: dict, name_class_index: dict,
                              field: str, col_mapping: dict = None):
    """Process hygiene file with improved matching.

    Formats:
    - Format A (22/23级): 公寓楼, 宿舍号, 床位号, 姓名, 班级, 得分情况 (NO student IDs)
    - Format B (24级): 公寓楼, 宿舍号/名称, 床位号, 姓名, 班级, 住宿人数, 得分情况
    - Format C (信息部教室卫生): 班级, 姓名, 教室分, 寝室分, 教室卫生总
    """
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception:
        return

    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue

        # Find header row
        header_idx = -1
        for i, row in enumerate(rows):
            if row and any('姓名' in str(c) for c in row if c):
                header_idx = i
                break
        if header_idx < 0:
            continue

        headers = [str(c).strip() if c else '' for c in rows[header_idx]]
        name_col = _find_header_col(headers, ['姓名'])
        class_col = _find_header_col(headers, ['班级'])
        score_col = _find_header_col(headers, ['得分情况', '得分', '分数', '卫生总', '卫生得分'])

        if name_col is None:
            continue

        # Determine if this is 信息部 format (has both 教室分 and 寝室分)
        has_personal_score = any('教室分' in h or '个人分' in h for h in headers)
        has_dorm_score = any('寝室分' in h or '宿舍分' in h for h in headers)

        current_class = str(sn).strip()
        used_names = set()  # Track processed names to avoid double-counting

        for row in rows[header_idx + 1:]:
            if not row:
                continue

            name = str(row[name_col]).strip() if len(row) > name_col and row[name_col] is not None else ''
            if not name or name in ('姓名', '班级', 'nan', 'None', ''):
                # Check for class header in col 0
                col0 = str(row[0]).strip() if row[0] is not None else ''
                if col0 and any(c.isdigit() for c in col0) and len(col0) < 15:
                    current_class = col0
                continue

            cls = current_class
            if class_col is not None and len(row) > class_col and row[class_col] is not None:
                row_cls = str(row[class_col]).strip()
                if row_cls and row_cls != 'nan':
                    cls = row_cls

            # Get score
            score = 0.0
            if score_col is not None and len(row) > score_col and row[score_col] is not None:
                try:
                    score = float(row[score_col])
                except (ValueError, TypeError):
                    score = 0.0

            # For 信息部 format: 教室分→教室卫生, 寝室分→宿舍卫生 (don't sum them)
            if has_personal_score and has_dorm_score:
                if field == '教室卫生':
                    # Use only 教室分, not the combined total
                    personal_col = _find_header_col(headers, ['教室分', '个人分'])
                    if personal_col is not None and len(row) > personal_col and row[personal_col] is not None:
                        try:
                            score = float(row[personal_col])
                        except (ValueError, TypeError):
                            score = 0.0
                    else:
                        score = 0.0
                elif field == '宿舍卫生':
                    # Use only 寝室分/宿舍分
                    dorm_col = _find_header_col(headers, ['寝室分', '宿舍分'])
                    if dorm_col is not None and len(row) > dorm_col and row[dorm_col] is not None:
                        try:
                            score = float(row[dorm_col])
                        except (ValueError, TypeError):
                            score = 0.0
                    else:
                        score = 0.0

            if score <= 0:
                continue

            # Avoid double-counting same student in same sheet
            dedup_key = (cls, name)
            if dedup_key in used_names:
                continue
            used_names.add(dedup_key)

            matched = _find_student_by_any(students, name_class_index, None, name, cls)
            if matched:
                students[matched][field] = max(students[matched][field], score)


# ============================================================
# V2 Org Class processor
# ============================================================

def _process_org_class_file_v2(filepath: str, students: dict, name_class_index: dict,
                                field: str, col_mapping: dict = None):
    """Process 团课出勤 file with improved format detection.

    Formats:
    - Format A (团课旷课统计): Title row, then class headers with (name, deduction) pairs
    - Format B: Simple name + deduction columns
    """
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception:
        return

    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue

        current_class = ''
        found_any_data = False

        for row in rows:
            if not row:
                continue

            # Check if first cell looks like a class header
            col0 = str(row[0]).strip() if row[0] is not None else ''
            col0_is_class = False
            if col0 and not col0.isdigit() and any(c.isdigit() for c in col0) and len(col0) < 20:
                if _looks_like_class_name(col0):
                    current_class = col0
                    found_any_data = True
                    col0_is_class = True  # DON'T skip — process data in this row too

            # Skip title rows (only before any data found)
            row_text = ' '.join([str(c).strip() for c in row if c is not None])
            if not found_any_data and ('统计表' in row_text or '旷课' in row_text or '团课' in row_text):
                continue

            # Process (name, value) pairs in current row
            # Try pairs: (1,2), (3,4), (5,6)
            # If col0 was a class header, skip pair (0,1) as col0 is class name
            pairs_to_check = [(1, 2), (3, 4)]
            if not col0_is_class:
                pairs_to_check.append((0, 1))

            for name_idx, val_idx in pairs_to_check:
                if len(row) > max(name_idx, val_idx):
                    name_cell = row[name_idx]
                    val_cell = row[val_idx]
                    if name_cell is None:
                        continue
                    name = str(name_cell).strip()
                    if not name or name in ('姓名', '学号', 'nan', 'None', '', '班级'):
                        continue
                    # Skip if name looks like a number
                    if name.replace('.', '').replace('-', '').isdigit():
                        continue
                    # Skip if name looks like a class name
                    if _looks_like_class_name(name):
                        continue
                    try:
                        val = float(val_cell) if val_cell is not None else 0
                    except (ValueError, TypeError):
                        continue
                    if val != 0:
                        matched = _find_student_by_any(students, name_class_index, None, name, current_class)
                        if matched:
                            students[matched][field] += -abs(val)


def _looks_like_class_name(text: str) -> bool:
    """Check if text looks like a class name (contains both Chinese and digits)."""
    has_chinese = any('一' <= c <= '鿿' for c in text)
    has_digit = any(c.isdigit() for c in text)
    return has_chinese and has_digit


# ============================================================
# File Preview - get sheets and headers for column mapping
# ============================================================

def preview_file_sheets(filepath: str) -> dict:
    """Preview sheets and column headers in an Excel file.

    Returns:
        {sheet_name: {headers: [str], sample_rows: [[str, ...], ...]}}
    """
    result = {}
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    except Exception:
        return result

    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            result[sn] = {'headers': [], 'sample_rows': []}
            continue

        # Find the best header row (skip empty rows at top)
        header_row = rows[0] if rows else []
        for i, row in enumerate(rows[:5]):
            non_empty = sum(1 for c in row if c is not None and str(c).strip())
            if non_empty >= 3:
                header_row = row
                break

        headers = [str(c).strip() if c is not None else '' for c in header_row]

        # Get sample rows (next 3 non-empty rows)
        sample_rows = []
        for row in rows[1:]:
            if len(sample_rows) >= 5:
                break
            row_vals = [str(c).strip() if c is not None else '' for c in row]
            if any(v for v in row_vals):
                sample_rows.append(row_vals[:len(headers)])

        result[sn] = {'headers': headers, 'sample_rows': sample_rows}

    wb.close()
    return result


# ============================================================
# Utility functions
# ============================================================

def _find_header_col(headers: list, keywords: list) -> int | None:
    """Find column index matching any keyword."""
    for kw in keywords:
        for i, h in enumerate(headers):
            if kw in h:
                return i
    return None


# ============================================================
# Excel writer
# ============================================================

def _write_moral_sheet(ws, class_students: list, output_headers: list = None,
                      values_data: list = None, class_name: str = ''):
    """Write one class sheet with proper headers and Excel formulas.
    v2.3: Plain white background, black text, SimSun 10pt, all centered.
    Supports column filtering via output_headers.

    Args:
        values_data: Optional list to append {学号, 姓名, 班级, 德育分} for _values sheet.
        class_name: Class name for _values sheet entries.
    """
    if output_headers is None:
        output_headers = list(MORAL_HEADERS)

    # Pre-compute scoring columns for the 德育分 formula.
    # Exclude non-scoring and subtotal columns (出勤总/卫生总 are formula-only).
    scoring_cols = []  # (col_index_1based, header_name, column_letter)
    for ci, hdr in enumerate(output_headers, 1):
        if hdr not in ('学号', '姓名', '德育分', '出勤总', '卫生总'):
            scoring_cols.append((ci, hdr, get_column_letter(ci)))

    num_cols = len(output_headers)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal='center', vertical='center')
    header_font = Font(name='SimSun', size=10, bold=True)
    data_font = Font(name='SimSun', size=10)
    formula_font = Font(name='SimSun', size=10, color='000000')

    # Row 1: Headers
    for ci, hdr in enumerate(output_headers, 1):
        cell = ws.cell(row=1, column=ci, value=hdr)
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    for ri, s in enumerate(class_students):
        row = ri + 2
        # Write columns in the order of output_headers
        for ci, hdr in enumerate(output_headers, 1):
            cell = ws.cell(row=row, column=ci)
            cell.font = data_font
            cell.alignment = center_align
            cell.border = thin_border

            if hdr == '学号':
                cell.value = s['学号']
                cell.number_format = '@'  # Text format, prevent scientific notation
            elif hdr == '姓名':
                cell.value = s['姓名']
            elif hdr == '基础分':
                cell.value = s['基础分']
            elif hdr == '评议分':
                cell.value = s.get('评议分', 0)
            elif hdr == '晚寝负责人':
                cell.value = s.get('晚寝负责人', 0)
            elif hdr == '早晚自习出勤':
                cell.value = s['早晚自习出勤']
            elif hdr == '课堂出勤':
                cell.value = s['课堂出勤']
            elif hdr == '出勤总':
                # Find positions of 早晚自习出勤 and 课堂出勤 in output
                zw_idx = None
                kt_idx = None
                for oi, oh in enumerate(output_headers, 1):
                    if oh == '早晚自习出勤': zw_idx = oi
                    elif oh == '课堂出勤': kt_idx = oi
                if zw_idx and kt_idx:
                    cell.value = f'={get_column_letter(zw_idx)}{row}+{get_column_letter(kt_idx)}{row}'
                    cell.font = formula_font
                else:
                    cell.value = (s['早晚自习出勤'] or 0) + (s['课堂出勤'] or 0)
            elif hdr == '宿舍卫生':
                cell.value = s.get('宿舍卫生', 0)
            elif hdr == '教室卫生':
                cell.value = s.get('教室卫生', 0)
            elif hdr == '卫生总':
                ss_idx = None
                js_idx = None
                for oi, oh in enumerate(output_headers, 1):
                    if oh == '宿舍卫生': ss_idx = oi
                    elif oh == '教室卫生': js_idx = oi
                if ss_idx and js_idx:
                    cell.value = f'={get_column_letter(ss_idx)}{row}+{get_column_letter(js_idx)}{row}'
                    cell.font = formula_font
                else:
                    cell.value = (s.get('宿舍卫生', 0) or 0) + (s.get('教室卫生', 0) or 0)
            elif hdr == '团课出勤':
                cell.value = s.get('团课出勤', 0)
            elif hdr == '青年大学习':
                cell.value = s.get('青年大学习', 0)
            elif hdr == '通报批评':
                cell.value = s.get('通报批评', 0)
            elif hdr == '违纪情况':
                cell.value = s.get('违纪情况', 0)
            elif hdr == '德育分':
                # Write Excel formula summing all scoring columns for auditability.
                # Skip 学号, 姓名, 德育分, 出勤总, 卫生总 (subtotals are formula-only).
                # The _values sheet stores the pre-computed value for downstream use.
                formula_parts = [f'{letter}{row}' for _, _, letter in scoring_cols]
                # v7.1: Cap total at 0 — scores below 0 become 0
                cell.value = '=MAX(0,' + '+'.join(formula_parts) + ')' if formula_parts else '0'
                cell.font = formula_font
                cell.number_format = '0.00'

                # Also compute numeric total for _values sheet (capped at 0)
                if values_data is not None:
                    numeric_total = sum(
                        float(s.get(hdr_name, 0) or 0)
                        for _, hdr_name, _ in scoring_cols
                    )
                    values_data.append({
                        '学号': s.get('学号', ''),
                        '姓名': s.get('姓名', ''),
                        '班级': class_name,
                        '德育分': round(max(0, numeric_total), 2),
                    })
            else:
                # Custom/unknown column — write value if available, else 0
                cell.value = s.get(hdr, 0)

    # Column widths
    for ci in range(1, num_cols + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 12

    if '学号' in output_headers:
        ws.column_dimensions[get_column_letter(output_headers.index('学号') + 1)].width = 14
    if '姓名' in output_headers:
        ws.column_dimensions[get_column_letter(output_headers.index('姓名') + 1)].width = 10

    ws.freeze_panes = 'A2'
