import os
import shutil
import tempfile
import unittest
from pathlib import Path

import openpyxl

from backend.moral_templates import (
    MORAL_PROJECT_TEMPLATES,
    analyze_moral_project_template,
    analyze_moral_project_templates,
    copy_moral_project_templates,
)
from backend.moral_vnext import build_moral_fresh_preview, process_moral_fresh


ROOT = Path(__file__).resolve().parents[1]
TEMPLATE_DIR = ROOT / "outputs" / "moral-project-templates"


class MoralProjectTemplateTests(unittest.TestCase):
    def setUp(self):
        self.temp = tempfile.TemporaryDirectory()
        self.addCleanup(self.temp.cleanup)
        self.roster = os.path.join(self.temp.name, "roster.xlsx")
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "顿河信251"
        sheet.append(["学号", "姓名", "行政班级"])
        sheet.append(["2025001", "陈雨昂", "顿河信251"])
        sheet.append(["2025002", "李明", "顿河信251"])
        workbook.save(self.roster)
        workbook.close()

    def _filled_template(self, rows):
        source = TEMPLATE_DIR / "德育项目模板-团课出勤.xlsx"
        target = Path(self.temp.name) / "团课出勤-已填写.xlsx"
        shutil.copy2(source, target)
        workbook = openpyxl.load_workbook(target)
        sheet = workbook["数据录入"]
        for index, values in enumerate(rows, 5):
            for col, value in enumerate(values, 1):
                sheet.cell(index, col, value)
        workbook.save(target)
        workbook.close()
        return str(target)

    def test_all_reserved_project_templates_are_present_and_recognizable(self):
        paths = [str(TEMPLATE_DIR / project["filename"]) for project in MORAL_PROJECT_TEMPLATES]
        result = analyze_moral_project_templates(paths)
        self.assertEqual(len(paths), 10)
        self.assertEqual(result["valid_count"], 10)
        self.assertEqual(result["error_count"], 0)
        self.assertEqual({item["project_name"] for item in result["files"]},
                         {item["name"] for item in MORAL_PROJECT_TEMPLATES})

    def test_duplicate_rows_are_allowed_and_add_deduct_share_original_column(self):
        path = self._filled_template([
            ["顿河信251", "陈雨昂", 8, None, "活动奖励"],
            ["顿河信251", "陈雨昂", None, 3, "团课缺勤"],
        ])
        analysis = analyze_moral_project_template(path)
        self.assertTrue(analysis["success"])
        self.assertEqual(analysis["row_count"], 2)
        self.assertEqual(analysis["duplicate_count"], 1)
        mapping = {"数据录入": {
            "enabled": True, "header_row": 3, "name_col": 1, "class_col": 0,
            "score_col": 2, "row_actions": {},
        }}
        deduct_mapping = {"数据录入": {**mapping["数据录入"], "score_col": 3, "row_actions": {}}}
        config = {
            "mode": "fresh",
            "roster_path": self.roster,
            "major_filter": "顿河信",
            "items": [
                {"id": "template-add", "name": "团课出勤", "direction": "add", "value_mode": "amount",
                 "standard_template": True, "sources": [{"path": path, "standard_template": True, "mappings": mapping}]},
                {"id": "template-deduct", "name": "团课出勤", "direction": "deduct", "value_mode": "amount",
                 "standard_template": True, "sources": [{"path": path, "standard_template": True, "mappings": deduct_mapping}]},
            ],
            "scoring": {"base": 80, "min": 0, "max": 115},
            "output_dir": self.temp.name,
        }
        preview = build_moral_fresh_preview(config)
        student = next(row for row in preview["students"] if row["name"] == "陈雨昂")
        self.assertFalse(preview["needs_review"])
        self.assertEqual(student["add_total"], 8)
        self.assertEqual(student["deduct_total"], 3)
        self.assertEqual(student["final"], 85)

        result = process_moral_fresh(config)
        workbook = openpyxl.load_workbook(result["output"], data_only=False)
        try:
            sheet = workbook["顿河信251"]
            headers = [cell.value for cell in sheet[1]]
            self.assertEqual(headers.count("团课出勤"), 1)
            self.assertEqual(sheet.cell(2, headers.index("团课出勤") + 1).value, 5)
            self.assertTrue(str(sheet.cell(2, headers.index("德育分") + 1).value).startswith("=MIN(115,MAX(0,"))
        finally:
            workbook.close()

    def test_same_row_cannot_contain_both_add_and_deduct(self):
        path = self._filled_template([["顿河信251", "陈雨昂", 5, 3, "错误示例"]])
        result = analyze_moral_project_template(path)
        self.assertFalse(result["success"])
        self.assertIn("同一行不能同时填写加分和扣分", result["error"])

    def test_negative_values_are_rejected(self):
        path = self._filled_template([["顿河信251", "陈雨昂", None, -3, "错误示例"]])
        result = analyze_moral_project_template(path)
        self.assertFalse(result["success"])
        self.assertIn("不要输入负号", result["error"])

    def test_copy_one_and_all_templates(self):
        single_dir = os.path.join(self.temp.name, "single")
        all_dir = os.path.join(self.temp.name, "all")
        os.makedirs(single_dir)
        os.makedirs(all_dir)
        single = copy_moral_project_templates("league_class", single_dir)
        all_templates = copy_moral_project_templates("all", all_dir)
        self.assertTrue(single["success"])
        self.assertEqual(single["count"], 1)
        self.assertTrue(all_templates["success"])
        self.assertEqual(all_templates["count"], 10)


class MoralProjectTemplateUiContractTests(unittest.TestCase):
    def test_fresh_route_has_template_center_and_batch_actions(self):
        script = (ROOT / "web" / "js" / "modules" / "moral.js").read_text(encoding="utf-8")
        self.assertIn("标准模板中心", script)
        self.assertIn("批量导入已填模板", script)
        self.assertIn("moralOpenTemplateBatch", script)
        self.assertIn("批量加分", script)
        self.assertIn("批量扣分", script)
        self.assertIn("standard_template", script)

    def test_fresh_route_allows_optional_template_projects(self):
        script = (ROOT / "web" / "js" / "modules" / "moral.js").read_text(encoding="utf-8")
        self.assertIn("选择本次需要的项目", script)
        self.assertIn("只校验已勾选项目", script)
        self.assertIn("function moralSetTemplateProjectSelected", script)
        self.assertIn("function moralFreshConfiguredItems", script)
        self.assertIn("items:configuredItems", script)
        self.assertIn("if (config.items.some", script)
        self.assertIn("当前未选择计分项目，将按基础分生成", script)


if __name__ == "__main__":
    unittest.main()
