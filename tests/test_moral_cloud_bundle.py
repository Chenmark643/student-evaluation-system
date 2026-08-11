import os
import tempfile
import unittest

import openpyxl

from backend.moral_cloud import prepare_moral_cloud_bundle


class MoralCloudBundleTests(unittest.TestCase):
    def setUp(self):
        self.temp = tempfile.TemporaryDirectory()
        self.addCleanup(self.temp.cleanup)

    def test_mixed_a_and_b_outputs_become_one_per_class_workbook(self):
        continuation = os.path.join(self.temp.name, "continuation.xlsx")
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "总分"
        sheet.append(["班级", "姓名", "团课扣分", "最终得分"])
        sheet.append(["顿河信251", "甲", 5, "=MIN(115,MAX(0,125-C2))"])
        sheet.append(["顿河信252", "乙", 8, "=MIN(115,MAX(0,100-C3))"])
        sheet["A1"].font = openpyxl.styles.Font(name="宋体", size=11)
        hidden = workbook.create_sheet("_德育新增设置")
        hidden.sheet_state = "hidden"
        workbook.save(continuation)
        workbook.close()

        fresh = os.path.join(self.temp.name, "fresh.xlsx")
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "顿河信231"
        sheet.append(["学号", "姓名", "基础分", "德育分"])
        sheet.append(["2023001", "丙", 80, "=MIN(115,MAX(0,SUM(C2)))"])
        settings = workbook.create_sheet("计分设置")
        settings.sheet_state = "hidden"
        workbook.save(fresh)
        workbook.close()

        result = prepare_moral_cloud_bundle([continuation, fresh])
        self.assertTrue(result["success"])
        self.assertEqual(result["source_count"], 2)
        self.assertEqual(set(result["classes"]), {"顿河信231", "顿河信251", "顿河信252"})

        bundle = openpyxl.load_workbook(result["output"], data_only=False)
        try:
            self.assertEqual(set(bundle.sheetnames), {"顿河信231", "顿河信251", "顿河信252"})
            self.assertEqual(bundle["顿河信251"]["D2"].value, "=MIN(115,MAX(0,125-C2))")
            self.assertEqual(bundle["顿河信252"]["D2"].value, "=MIN(115,MAX(0,100-C2))")
            self.assertEqual(bundle["顿河信251"]["A1"].font.name, "宋体")
            self.assertTrue(all(sheet.sheet_state == "visible" for sheet in bundle.worksheets))
        finally:
            bundle.close()

    def test_later_file_replaces_the_same_class(self):
        paths = []
        for index, score in enumerate((80, 95), 1):
            path = os.path.join(self.temp.name, f"result-{index}.xlsx")
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "顿河信251"
            sheet.append(["姓名", "德育分"])
            sheet.append(["甲", score])
            workbook.save(path)
            workbook.close()
            paths.append(path)

        result = prepare_moral_cloud_bundle(paths)
        bundle = openpyxl.load_workbook(result["output"], data_only=False)
        try:
            self.assertEqual(bundle["顿河信251"]["B2"].value, 95)
            self.assertEqual(result["source_by_class"]["顿河信251"], "result-2.xlsx")
        finally:
            bundle.close()


if __name__ == "__main__":
    unittest.main()
