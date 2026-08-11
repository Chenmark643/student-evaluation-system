"""Tests for the configurable moral continuation workflow."""

import os
import tempfile
import unittest

import openpyxl

from backend.moral_vnext import build_moral_fresh_preview, build_moral_preview, list_moral_students, process_moral_fresh, process_moral_vnext


class MoralVNextTests(unittest.TestCase):
    def setUp(self):
        self.temp = tempfile.TemporaryDirectory()
        self.addCleanup(self.temp.cleanup)
        self.existing = os.path.join(self.temp.name, "existing.xlsx")
        self.item = os.path.join(self.temp.name, "org.xlsx")
        self.roster = os.path.join(self.temp.name, "roster.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "总分"
        ws.append(["班级", "姓名", "总扣分", "总加分", "最终得分"])
        ws.append(["顿河信251", "甲", 10, 20, 115])  # raw = 125
        ws.append(["顿河信251", "乙", 20, 0, 95])
        wb.save(self.existing)
        wb.close()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "顿河信251"
        ws.append(["姓名", "班级", "扣分情况"])
        ws.append(["甲", "顿河信251", -5])
        ws.append(["乙", "顿河信251", -5])
        wb.save(self.item)
        wb.close()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "顿河信251"
        ws.append(["学号", "姓名", "行政班级"])
        ws.append(["2025001", "甲", "顿河信251"])
        ws.append(["2025002", "乙", "顿河信251"])
        wb.save(self.roster)
        wb.close()

    def config(self, basis="raw"):
        return {
            "existing": {"path": self.existing, "mappings": {"总分": {
                "enabled": True, "header_row": 0, "name_col": 1, "class_col": 0,
                "deduction_col": 2, "addition_col": 3, "score_col": 4,
            }}},
            "items": [{
                "id": "org", "name": "团课出勤", "direction": "deduct", "value_mode": "signed",
                "sources": [{"path": self.item, "mappings": {"顿河信251": {
                    "enabled": True, "header_row": 0, "name_col": 0, "class_col": 1, "score_col": 2,
                }}}],
            }],
            "scoring": {"base": 115, "min": 0, "max": 115, "continuation_basis": basis},
            "output_dir": self.temp.name,
        }

    def test_user_can_keep_uncapped_overflow_buffer(self):
        preview = build_moral_preview(self.config("raw"))
        first = next(row for row in preview["students"] if row["name"] == "甲")
        self.assertEqual(first["existing_raw"], 125)
        self.assertEqual(first["existing_display"], 115)
        self.assertEqual(first["final"], 115)
        self.assertEqual(preview["summary"]["basis_difference_count"], 1)

    def test_user_can_restart_from_displayed_capped_score(self):
        preview = build_moral_preview(self.config("display"))
        first = next(row for row in preview["students"] if row["name"] == "甲")
        self.assertEqual(first["final"], 110)

    def test_export_keeps_formulas_for_every_student(self):
        result = process_moral_vnext(self.config("raw"))
        self.assertTrue(result["success"])
        wb = openpyxl.load_workbook(result["output"], data_only=False)
        try:
            ws = wb["总分"]
            headers = [cell.value for cell in ws[1]]
            self.assertEqual(headers, ["班级", "姓名", "团课出勤（扣分）", "总扣分", "总加分", "最终得分"])
            self.assertNotIn("更新后最终得分", headers)
            final_col = headers.index("最终得分") + 1
            for row in (2, 3):
                self.assertTrue(str(ws.cell(row, final_col).value).startswith("=MIN(115,MAX(0,"))
                self.assertNotIn("_德育新增设置", str(ws.cell(row, final_col).value))
                self.assertTrue(str(ws.cell(row, headers.index("总扣分") + 1).value).startswith("="))
                self.assertTrue(str(ws.cell(row, headers.index("总加分") + 1).value).startswith("="))
            values = wb["_values"]
            rows = list(values.iter_rows(values_only=True))
            self.assertIn((None, "甲", "顿河信251", 115), rows)
            self.assertEqual(wb["_德育新增设置"].sheet_state, "hidden")
        finally:
            wb.close()

    def test_original_workbook_structure_and_values_are_preserved(self):
        wb = openpyxl.load_workbook(self.existing)
        detail = wb.create_sheet("原有明细")
        detail["A1"] = "不得改动"
        wb["总分"]["H1"].fill = openpyxl.styles.PatternFill("solid", fgColor="FFFFFF")
        wb.save(self.existing)
        wb.close()

        result = process_moral_vnext(self.config("raw"))
        wb = openpyxl.load_workbook(result["output"], data_only=False)
        try:
            self.assertEqual(wb["原有明细"]["A1"].value, "不得改动")
            self.assertEqual(wb["总分"]["C2"].value, 5)
            self.assertEqual(wb["总分"]["C1"].value, "团课出勤（扣分）")
            self.assertEqual(wb["总分"]["D1"].value, "总扣分")
            self.assertEqual(wb["总分"]["E1"].value, "总加分")
            self.assertEqual(wb["总分"]["F1"].value, "最终得分")
            self.assertNotIn("更新后最终得分", [cell.value for cell in wb["总分"][1]])
            settings = wb["_德育新增设置"]
            first_row = next(row for row in range(9, settings.max_row + 1) if str(settings.cell(row, 1).value).endswith("|甲"))
            self.assertEqual(settings.cell(first_row, 2).value, 10)
            self.assertEqual(settings.cell(first_row, 3).value, 20)
        finally:
            wb.close()

    def test_new_student_missing_from_existing_is_a_blocker(self):
        wb = openpyxl.load_workbook(self.item)
        wb["顿河信251"].append(["丙", "顿河信251", -5])
        wb.save(self.item)
        wb.close()
        preview = build_moral_preview(self.config("raw"))
        self.assertTrue(preview["needs_review"])
        self.assertTrue(any(issue["type"] == "missing_existing_student" for issue in preview["issues"]))

    def test_continuation_keeps_rows_with_missing_scores_so_they_can_be_resolved(self):
        workbook = openpyxl.load_workbook(self.existing)
        sheet = workbook["总分"]
        sheet["C3"] = None
        sheet["D3"] = None
        sheet["E3"] = None
        workbook.save(self.existing)
        workbook.close()

        config = self.config("raw")
        preview = build_moral_preview(config)
        student = next(row for row in preview["students"] if row["name"] == "乙")
        self.assertIsNone(student["existing_raw"])
        self.assertTrue(any(
            issue["type"] == "raw_basis_unavailable" and issue["student_key"] == student["key"]
            for issue in preview["issues"]
        ))

        config["overrides"] = {student["key"]: {"raw": 95, "display": 95}}
        resolved = build_moral_preview(config)
        self.assertFalse(resolved["needs_review"])
        self.assertEqual(next(row for row in resolved["students"] if row["name"] == "乙")["final"], 90)

    def test_display_score_can_be_used_to_fill_missing_raw_score(self):
        config = self.config("raw")
        config["existing"]["mappings"]["总分"].pop("deduction_col")
        config["existing"]["mappings"]["总分"].pop("addition_col")
        preview = build_moral_preview(config)
        student = next(row for row in preview["students"] if row["name"] == "甲")
        self.assertIsNone(student["existing_raw"])
        self.assertEqual(student["existing_display"], 115)

        config["overrides"] = {student["key"]: {"raw": student["existing_display"]}}
        resolved = build_moral_preview(config)
        self.assertFalse(any(
            issue["type"] == "raw_basis_unavailable" and issue["student_key"] == student["key"]
            for issue in resolved["issues"]
        ))
        self.assertEqual(next(row for row in resolved["students"] if row["name"] == "甲")["final"], 110)

    def test_continuation_filters_existing_materials_and_picker_by_selected_major(self):
        workbook = openpyxl.load_workbook(self.existing)
        workbook["总分"].append(["顿河土251", "外专业学生", 0, 0, 115])
        workbook.save(self.existing)
        workbook.close()

        workbook = openpyxl.load_workbook(self.item)
        workbook["顿河信251"].append(["外专业学生", "顿河土251", -9])
        workbook.save(self.item)
        workbook.close()

        config = self.config("raw")
        config["major_filter"] = "顿河信"
        preview = build_moral_preview(config)
        self.assertFalse(preview["needs_review"])
        self.assertEqual({row["class_name"] for row in preview["students"]}, {"顿河信251"})
        self.assertFalse(any(issue.get("student") == "外专业学生" for issue in preview["issues"]))
        picker = list_moral_students(config)
        self.assertEqual({row["class_name"] for row in picker["students"]}, {"顿河信251"})

    def test_custom_bounds_are_applied(self):
        config = self.config("raw")
        config["scoring"].update({"min": 20, "max": 100})
        preview = build_moral_preview(config)
        first = next(row for row in preview["students"] if row["name"] == "甲")
        self.assertEqual(first["final"], 100)

    def test_fresh_mode_supports_dynamic_add_and_deduct_items_with_bounds(self):
        config = {
            "mode": "fresh", "roster_path": self.roster,
            "items": self.config("raw")["items"],
            "scoring": {"base": 80, "min": 0, "max": 82},
            "output_dir": self.temp.name,
        }
        config["items"].append({
            "id": "bonus", "name": "评议", "direction": "add", "value_mode": "amount",
            "sources": [{"path": self.item, "mappings": {"顿河信251": {
                "enabled": True, "header_row": 0, "name_col": 0, "class_col": 1, "score_col": 2,
            }}}],
        })
        preview = build_moral_fresh_preview(config)
        first = next(row for row in preview["students"] if row["name"] == "甲")
        self.assertEqual(first["final"], 80)
        result = process_moral_fresh(config)
        wb = openpyxl.load_workbook(result["output"], data_only=False)
        try:
            ws = wb["顿河信251"]
            headers = [cell.value for cell in ws[1]]
            self.assertEqual(headers, [
                "学号", "姓名", "基础分", "评议分", "晚寝负责人",
                "早晚自习出勤", "课堂出勤", "出勤总",
                "宿舍卫生", "教室卫生", "卫生总",
                "团课出勤", "青年大学习", "通报批评", "违纪情况", "德育分",
            ])
            self.assertEqual(ws.cell(2, headers.index("团课出勤") + 1).value, -5)
            self.assertEqual(ws.cell(2, headers.index("评议分") + 1).value, 5)
            formula = ws.cell(2, headers.index("德育分") + 1).value
            self.assertTrue(formula.startswith("=MIN(82,MAX(0,"))
            self.assertNotIn("计分设置", formula)
            self.assertEqual(wb["计分设置"].sheet_state, "hidden")
        finally:
            wb.close()

    def test_fresh_output_matches_reference_workbook_format(self):
        config = {
            "mode": "fresh", "roster_path": self.roster,
            "items": [
                {"id": "org", "name": "团课出勤", "direction": "deduct", "value_mode": "amount", "sources": [], "manual_values": {}},
                {"id": "custom", "name": "志愿服务", "direction": "add", "value_mode": "amount", "sources": [], "manual_values": {}},
            ],
            "scoring": {"base": 80, "min": 0, "max": 115},
            "output_dir": self.temp.name,
        }
        result = process_moral_fresh(config)
        wb = openpyxl.load_workbook(result["output"], data_only=False)
        try:
            ws = wb["顿河信251"]
            headers = [cell.value for cell in ws[1]]
            self.assertEqual(headers[-2:], ["志愿服务", "德育分"])
            self.assertNotIn("班级", headers)
            self.assertNotIn("总加分", headers)
            self.assertNotIn("总扣分", headers)
            self.assertEqual(ws.row_dimensions[1].height, 28)
            self.assertEqual(ws.row_dimensions[2].height, 26)
            self.assertAlmostEqual(ws.column_dimensions["A"].width, 8.09, places=2)
            self.assertEqual(ws["A1"].font.name, "宋体")
            self.assertEqual(ws["A1"].font.sz, 11)
            self.assertFalse(ws["A1"].font.bold)
            self.assertEqual(ws["A2"].font.name, "宋体")
            self.assertEqual(ws["A2"].font.sz, 10)
            self.assertEqual(ws["A1"].alignment.horizontal, "center")
            self.assertTrue(ws["A1"].alignment.wrap_text)
            self.assertEqual(ws["A1"].fill.fgColor.rgb, "00FFFFFF")
            self.assertEqual(ws["A1"].border.left.style, "thin")
            self.assertIsNone(ws.freeze_panes)
            self.assertIsNone(ws.auto_filter.ref)
            self.assertTrue(str(ws.cell(2, len(headers)).value).startswith("=MIN("))
            self.assertEqual(wb.active.title, "顿河信251")
        finally:
            wb.close()

    def test_project_direction_can_change_later_without_losing_manual_values(self):
        students = list_moral_students({"mode": "fresh", "roster_path": self.roster})["students"]
        first_key = next(row["key"] for row in students if row["name"] == "甲")
        item = {
            "id": "attendance", "name": "团课出勤", "direction": "deduct",
            "value_mode": "amount", "sources": [], "manual_values": {first_key: 5},
        }
        config = {
            "mode": "fresh", "roster_path": self.roster, "items": [item],
            "scoring": {"base": 80, "min": 0, "max": 115}, "output_dir": self.temp.name,
        }
        deducted = build_moral_fresh_preview(config)
        self.assertEqual(next(row for row in deducted["students"] if row["name"] == "甲")["final"], 75)

        item["direction"] = "add"
        added = build_moral_fresh_preview(config)
        self.assertEqual(next(row for row in added["students"] if row["name"] == "甲")["final"], 85)
        self.assertEqual(item["manual_values"], {first_key: 5})

    def test_auto_value_mode_accepts_source_sign_without_extra_user_decision(self):
        config = self.config("raw")
        config["items"][0]["value_mode"] = "auto"
        config["items"][0]["direction"] = "add"
        preview = build_moral_preview(config)
        first = next(row for row in preview["students"] if row["name"] == "甲")
        self.assertEqual(first["add_total"], 5)
        self.assertFalse(any(issue["type"] == "sign_conflict" for issue in preview["issues"]))

    def test_real_sample_chen_yuang_remains_115_and_all_finals_are_bounded(self):
        existing = r"C:\Users\cheny\Downloads\顿河信251德育分.xlsx"
        item = r"C:\Users\cheny\Downloads\顿河信251团课.xlsx"
        if not os.path.isfile(existing) or not os.path.isfile(item):
            self.skipTest("user samples are not available")
        config = {
            "existing": {"path": existing, "mappings": {"总分": {
                "enabled": True, "header_row": 0, "name_col": 1, "class_col": 0,
                "deduction_col": 8, "addition_col": 13, "score_col": 14,
            }}},
            "items": [{
                "id": "org", "name": "团课出勤", "direction": "deduct", "value_mode": "signed",
                "sources": [{"path": item, "mappings": {"顿河信251": {
                    "enabled": True, "header_row": 3, "name_col": 0, "class_col": 1, "score_col": 2,
                }}}],
            }],
            "scoring": {"base": 115, "min": 0, "max": 115, "continuation_basis": "raw"},
            "output_dir": self.temp.name,
        }
        preview = build_moral_preview(config)
        chen = next(row for row in preview["students"] if row["name"] == "陈雨昂")
        self.assertEqual(chen["existing_raw"], 162)
        self.assertEqual(chen["deduct_total"], 0)
        self.assertEqual(chen["final"], 115)
        finals = [row["final"] for row in preview["students"] if row["final"] is not None]
        self.assertTrue(all(0 <= value <= 115 for value in finals))

    def test_batch_entry_accumulates_with_uploaded_item_value(self):
        config = self.config("raw")
        students = list_moral_students(config)["students"]
        first_key = next(row["key"] for row in students if row["name"] == "甲")
        config["items"][0]["manual_values"] = {first_key: 7}
        preview = build_moral_preview(config)
        first = next(row for row in preview["students"] if row["name"] == "甲")
        self.assertEqual(first["deduct_total"], 12)
        self.assertEqual(first["final"], 113)

    def test_fresh_batch_only_project_does_not_require_a_file(self):
        students = list_moral_students({"mode": "fresh", "roster_path": self.roster})["students"]
        first_key = next(row["key"] for row in students if row["name"] == "甲")
        config = {
            "mode": "fresh", "roster_path": self.roster,
            "items": [{"id": "award", "name": "评议", "direction": "add", "value_mode": "amount", "sources": [], "manual_values": {first_key: 8}}],
            "scoring": {"base": 80, "min": 0, "max": 85}, "output_dir": self.temp.name,
        }
        preview = build_moral_fresh_preview(config)
        first = next(row for row in preview["students"] if row["name"] == "甲")
        self.assertEqual(first["final"], 85)

    def test_fresh_mode_filters_roster_materials_and_batch_picker_by_major(self):
        roster_wb = openpyxl.load_workbook(self.roster)
        roster_wb["顿河信251"].append(["2025999", "外专业学生", "顿河土251"])
        roster_wb.save(self.roster)
        roster_wb.close()

        item_wb = openpyxl.load_workbook(self.item)
        item_wb["顿河信251"].append(["外专业学生", "顿河土251", -9])
        item_wb.save(self.item)
        item_wb.close()

        config = {
            "mode": "fresh", "roster_path": self.roster, "major_filter": "顿河信",
            "items": self.config("raw")["items"],
            "scoring": {"base": 80, "min": 0, "max": 115},
            "output_dir": self.temp.name,
        }
        preview = build_moral_fresh_preview(config)
        self.assertFalse(preview["needs_review"])
        self.assertEqual({row["class_name"] for row in preview["students"]}, {"顿河信251"})
        picker = list_moral_students(config)
        self.assertEqual({row["class_name"] for row in picker["students"]}, {"顿河信251"})

    def test_fresh_review_can_assign_an_unmatched_source_row_to_roster_student(self):
        item_wb = openpyxl.load_workbook(self.item)
        item_wb["顿河信251"]["A2"] = "甲同学"
        item_wb.save(self.item)
        item_wb.close()

        config = {
            "mode": "fresh", "roster_path": self.roster, "major_filter": "顿河信",
            "items": self.config("raw")["items"],
            "scoring": {"base": 80, "min": 0, "max": 115},
            "output_dir": self.temp.name,
        }
        first_preview = build_moral_fresh_preview(config)
        issue = next(issue for issue in first_preview["issues"] if issue["type"] == "missing_roster_student")
        self.assertEqual(issue["item_id"], "org")
        self.assertEqual(issue["source_index"], 0)
        self.assertEqual(issue["sheet_name"], "顿河信251")
        self.assertEqual(issue["excel_row"], 2)

        target = next(row for row in list_moral_students(config)["students"] if row["name"] == "甲")
        mapping = config["items"][0]["sources"][0]["mappings"]["顿河信251"]
        mapping["row_actions"] = {"2": {"action": "match", "student_key": target["key"]}}
        resolved = build_moral_fresh_preview(config)
        self.assertFalse(resolved["needs_review"])
        student = next(row for row in resolved["students"] if row["name"] == "甲")
        self.assertEqual(student["deduct_total"], 5)

    def test_fresh_review_suggests_a_single_character_name_typo_without_auto_matching(self):
        roster_wb = openpyxl.load_workbook(self.roster)
        roster_wb["顿河信251"]["B2"] = "张三"
        roster_wb.save(self.roster)
        roster_wb.close()
        item_wb = openpyxl.load_workbook(self.item)
        item_wb["顿河信251"]["A2"] = "张山"
        item_wb.save(self.item)
        item_wb.close()

        config = {
            "mode": "fresh", "roster_path": self.roster, "major_filter": "顿河信",
            "items": self.config("raw")["items"],
            "scoring": {"base": 80, "min": 0, "max": 115},
            "output_dir": self.temp.name,
        }
        preview = build_moral_fresh_preview(config)
        self.assertTrue(preview["needs_review"])
        issue = next(issue for issue in preview["issues"] if issue["type"] == "missing_roster_student")
        self.assertEqual(issue["suggested_student_key"], "id:2025001")
        self.assertEqual(issue["suggestions"][0]["name"], "张三")
        self.assertEqual(issue["suggestions"][0]["distance"], 1)

    def test_one_project_accepts_multiple_source_files(self):
        config = self.config("raw")
        config["items"][0]["sources"].append(dict(config["items"][0]["sources"][0]))
        preview = build_moral_preview(config)
        first = next(row for row in preview["students"] if row["name"] == "甲")
        self.assertEqual(first["deduct_total"], 10)

    def test_continuation_inserts_both_directions_before_existing_totals(self):
        config = self.config("raw")
        config["items"].append({
            "id": "bonus", "name": "评议", "direction": "add", "value_mode": "amount",
            "sources": [dict(config["items"][0]["sources"][0])],
        })
        result = process_moral_vnext(config)
        self.assertTrue(result["success"])
        wb = openpyxl.load_workbook(result["output"], data_only=False)
        try:
            ws = wb["总分"]
            headers = [cell.value for cell in ws[1]]
            self.assertEqual(headers, [
                "班级", "姓名", "团课出勤（扣分）", "总扣分",
                "评议（加分）", "总加分", "最终得分",
            ])
            self.assertEqual(ws["C2"].value, 5)
            self.assertEqual(ws["E2"].value, 5)
            self.assertTrue(str(ws["D2"].value).startswith("="))
            self.assertTrue(str(ws["F2"].value).startswith("="))
            self.assertTrue(str(ws["G2"].value).startswith("=MIN("))
        finally:
            wb.close()


if __name__ == "__main__":
    unittest.main()
