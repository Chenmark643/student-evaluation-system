import tempfile
import unittest
import re
import zipfile
from pathlib import Path
from unittest.mock import patch

import openpyxl

from backend.module_e_annual import (
    _safe_sheet_name,
    process_annual_comprehensive,
    process_annual_gpa,
)


class AnnualRankingTests(unittest.TestCase):
    def _gpa_workbook(self, path, rows, include_credits=True, values_sheet=True):
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        by_class = {}
        for row in rows:
            by_class.setdefault(row['class'], []).append(row)
        for class_name, students in by_class.items():
            ws = wb.create_sheet(class_name)
            headers = ['学号', '姓名', '课程门数']
            if include_credits:
                headers.append('总学分')
            headers.extend(['课程A', '学分绩点'])
            ws.append(headers)
            ws.append(['', '', '', '', 3, ''] if include_credits else ['', '', '', 3, ''])
            for student in students:
                data = [student['id'], student['name'], 1]
                if include_credits:
                    data.append(student['credits'])
                data.extend([student['score'], '=1+1'])
                ws.append(data)
        if values_sheet:
            ws = wb.create_sheet('_values')
            ws.append(['学号', '姓名', '班级', '学分绩点'])
            for student in rows:
                ws.append([student['id'], student['name'], student['class'], student['score']])
            ws.sheet_state = 'hidden'
        wb.save(path)

    def _comp_workbook(self, path, rows):
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        for class_name in sorted({row['class'] for row in rows}):
            ws = wb.create_sheet(class_name)
            ws.append(['综合测评表', '', '', ''])
            ws.append(['学号', '姓名', '学分绩点', '综合测评'])
            for student in [row for row in rows if row['class'] == class_name]:
                ws.append([student['id'], student['name'], 80, '=1+1'])
        ws = wb.create_sheet('_values')
        ws.append(['学号', '姓名', '班级', '综合测评'])
        for student in rows:
            ws.append([student['id'], student['name'], student['class'], student['score']])
        ws.sheet_state = 'hidden'
        wb.save(path)

    def test_gpa_uses_credit_weighted_mean_and_creates_two_rankings(self):
        with tempfile.TemporaryDirectory() as folder:
            root = Path(folder)
            first, second = root / '2024-2025-1.xlsx', root / '2024-2025-2.xlsx'
            self._gpa_workbook(first, [
                {'id': '240001', 'name': '甲', 'class': '顿河信241', 'score': 80, 'credits': 10},
                {'id': '240002', 'name': '乙', 'class': '顿河信241', 'score': 90, 'credits': 10},
                {'id': '240003', 'name': '丙', 'class': '顿河信242', 'score': 88, 'credits': 10},
            ])
            self._gpa_workbook(second, [
                {'id': '240001', 'name': '甲', 'class': '顿河信241', 'score': 95, 'credits': 20},
                {'id': '240002', 'name': '乙', 'class': '顿河信241', 'score': 90, 'credits': 20},
                {'id': '240003', 'name': '丙', 'class': '顿河信242', 'score': 88, 'credits': 20},
            ])

            result = process_annual_gpa(str(first), str(second), folder, '2024-2025')
            self.assertTrue(result['success'])
            self.assertEqual(result['credit_weighted_count'], 3)
            self.assertTrue(Path(result['output1']).exists())
            self.assertTrue(Path(result['output2']).exists())

            class_wb = openpyxl.load_workbook(result['output1'], data_only=True)
            self.assertIn('顿河信241', class_wb.sheetnames)
            ws = class_wb['顿河信241']
            by_id = {str(ws.cell(row, 1).value): row for row in range(2, ws.max_row + 1)}
            row = by_id['240001']
            self.assertEqual(ws.cell(row, 4).value, 90)
            self.assertEqual(ws.cell(row, 5).value, 1)
            self.assertEqual(ws.cell(row, 6).value, 0.5)

            program_wb = openpyxl.load_workbook(result['output2'], data_only=True)
            self.assertIn('顿河信24', program_wb.sheetnames)
            program_ws = program_wb['顿河信24']
            scores = [program_ws.cell(row, 4).value for row in range(2, program_ws.max_row + 1)]
            self.assertEqual(scores, [90, 90, 88])
            self.assertEqual([program_ws.cell(row, 5).value for row in range(2, 5)], [1, 1, 3])

    def test_gpa_falls_back_to_mean_for_legacy_tables_without_credits(self):
        with tempfile.TemporaryDirectory() as folder:
            root = Path(folder)
            first, second = root / 'first.xlsx', root / 'second.xlsx'
            self._gpa_workbook(first, [
                {'id': '240001', 'name': '甲', 'class': '国电241', 'score': 80, 'credits': 0},
            ], include_credits=False, values_sheet=False)
            self._gpa_workbook(second, [
                {'id': '240001', 'name': '甲', 'class': '国电241', 'score': 90, 'credits': 0},
            ], include_credits=False, values_sheet=False)
            # Replace formula placeholders with the numeric legacy score.
            for path, score in ((first, 80), (second, 90)):
                wb = openpyxl.load_workbook(path)
                ws = wb[wb.sheetnames[0]]
                ws.cell(3, ws.max_column, score)
                wb.save(path)

            result = process_annual_gpa(str(first), str(second), folder, '2024-2025')
            self.assertEqual(result['mean_fallback_count'], 1)
            wb = openpyxl.load_workbook(result['output1'], data_only=True)
            self.assertEqual(wb['国电241']['D2'].value, 85)

    def test_major_filter_limits_export_and_keeps_full_class_names(self):
        with tempfile.TemporaryDirectory() as folder:
            root = Path(folder)
            first, second = root / 'first.xlsx', root / 'second.xlsx'
            rows = [
                {'id': '240001', 'name': '甲', 'class': '顿河信241', 'score': 80, 'credits': 10},
                {'id': '240002', 'name': '乙', 'class': '顿河土241', 'score': 90, 'credits': 10},
                {'id': '240003', 'name': '丙', 'class': '顿河信息241', 'score': 95, 'credits': 10},
            ]
            self._gpa_workbook(first, rows)
            self._gpa_workbook(second, rows)

            result = process_annual_gpa(
                str(first), str(second), folder, '2024-2025', major_filter='顿河信'
            )
            self.assertEqual(result['student_count'], 1)
            self.assertEqual(result['major_filter'], '顿河信')
            self.assertEqual(result['matched_count'], 1)
            self.assertEqual(result['first_only_count'], 0)
            self.assertEqual(result['second_only_count'], 0)
            class_wb = openpyxl.load_workbook(result['output1'], data_only=True)
            self.assertEqual(class_wb.sheetnames, ['顿河信241'])
            program_wb = openpyxl.load_workbook(result['output2'], data_only=True)
            self.assertEqual(program_wb.sheetnames, ['顿河信24'])

            with self.assertRaisesRegex(ValueError, '没有匹配学生'):
                process_annual_gpa(
                    str(first), str(second), folder, '2024-2025', major_filter='不存在专业'
                )
            with self.assertRaisesRegex(ValueError, '没有匹配学生'):
                process_annual_gpa(
                    str(first), str(second), folder, '2024-2025', major_filter='信'
                )

    def test_desktop_bridge_passes_major_filter_to_annual_processing(self):
        from backend import bridge

        with patch.object(bridge, 'process_annual_gpa', return_value={'success': True}) as process_gpa:
            result = bridge.run_annual_gpa('s1.xlsx', 's2.xlsx', 'out', '2024-2025', '顿河信')
        self.assertTrue(result['success'])
        process_gpa.assert_called_once_with(
            's1.xlsx', 's2.xlsx', 'out',
            academic_year='2024-2025', major_filter='顿河信',
        )

        with patch.object(
            bridge, 'process_annual_comprehensive', return_value={'success': True}
        ) as process_comprehensive:
            result = bridge.run_annual_comprehensive(
                's1.xlsx', 's2.xlsx', 'out', '2024-2025', '顿河交'
            )
        self.assertTrue(result['success'])
        process_comprehensive.assert_called_once_with(
            's1.xlsx', 's2.xlsx', 'out',
            academic_year='2024-2025', major_filter='顿河交',
        )

        root = Path(__file__).resolve().parents[1]
        annual_ui = (root / 'web' / 'js' / 'modules' / 'annual.js').read_text(encoding='utf-8')
        self.assertIn('MajorScope.requireForExport()', annual_ui)
        self.assertIn('MajorScope.get())()', annual_ui)
        self.assertIn('导出专业限制', annual_ui)

    def test_comprehensive_averages_semesters_and_reports_missing_students(self):
        with tempfile.TemporaryDirectory() as folder:
            root = Path(folder)
            first, second = root / 'first.xlsx', root / 'second.xlsx'
            self._comp_workbook(first, [
                {'id': '230001', 'name': '甲', 'class': '顿河交231', 'score': 80},
                {'id': '230002', 'name': '乙', 'class': '顿河交231', 'score': 90},
            ])
            self._comp_workbook(second, [
                {'id': '230001', 'name': '甲', 'class': '顿河交231', 'score': 100},
                {'id': '230003', 'name': '丙', 'class': '顿河交231', 'score': 95},
            ])
            result = process_annual_comprehensive(str(first), str(second), folder, '2024-2025')
            self.assertEqual(result['matched_count'], 1)
            self.assertEqual(result['first_only_count'], 1)
            self.assertEqual(result['second_only_count'], 1)
            wb = openpyxl.load_workbook(result['output2'], data_only=True)
            ws = wb['顿河交23']
            values = {str(ws.cell(row, 2).value): ws.cell(row, 4).value for row in range(2, ws.max_row + 1)}
            self.assertEqual(values, {'230003': 95, '230001': 90, '230002': 90})

    def test_comprehensive_major_filter_is_exact_and_diagnostics_are_scoped(self):
        with tempfile.TemporaryDirectory() as folder:
            root = Path(folder)
            first, second = root / 'first.xlsx', root / 'second.xlsx'
            self._comp_workbook(first, [
                {'id': '230001', 'name': '甲', 'class': '顿河交231', 'score': 80},
                {'id': '230002', 'name': '乙', 'class': '顿河交通231', 'score': 90},
                {'id': '230003', 'name': '丙', 'class': '顿河土231', 'score': 70},
            ])
            self._comp_workbook(second, [
                {'id': '230001', 'name': '甲', 'class': '顿河交231', 'score': 100},
                {'id': '230002', 'name': '乙', 'class': '顿河交通231', 'score': 95},
            ])
            result = process_annual_comprehensive(
                str(first), str(second), folder, '2024-2025', major_filter='顿河交'
            )
            self.assertEqual(result['student_count'], 1)
            self.assertEqual(result['matched_count'], 1)
            self.assertEqual(result['first_only_count'], 0)
            self.assertEqual(result['second_only_count'], 0)
            self.assertEqual(result['class_count'], 1)
            self.assertEqual(result['program_count'], 1)

    def test_sheet_names_are_sanitized_truncated_and_deduplicated(self):
        used = set()
        first = _safe_sheet_name('顿河信/超长班级名称12345678901234567890', used)
        second = _safe_sheet_name('顿河信\\超长班级名称12345678901234567890', used)
        self.assertLessEqual(len(first), 31)
        self.assertLessEqual(len(second), 31)
        self.assertNotIn('/', first)
        self.assertNotIn('\\', second)
        self.assertNotEqual(first, second)

    def test_export_neutralizes_formula_like_name_and_class_text(self):
        with tempfile.TemporaryDirectory() as folder:
            root = Path(folder)
            first, second = root / 'first.xlsx', root / 'second.xlsx'
            malicious_name = '=HYPERLINK("https://example.invalid","点我")'
            rows = [
                {'id': '230001', 'name': malicious_name, 'class': '@顿河交231', 'score': 80},
            ]
            self._comp_workbook(first, rows)
            self._comp_workbook(second, rows)
            for path in (first, second):
                wb = openpyxl.load_workbook(path)
                for ws in wb.worksheets:
                    for row in ws.iter_rows():
                        for cell in row:
                            if cell.value in (malicious_name, '@顿河交231'):
                                cell.data_type = 's'
                wb.save(path)

            result = process_annual_comprehensive(str(first), str(second), folder, '2024-2025')
            output = openpyxl.load_workbook(result['output1'], data_only=False)
            ws = output[output.sheetnames[0]]
            self.assertEqual(ws['A2'].value, "'@顿河交231")
            self.assertEqual(ws['C2'].value, "'" + malicious_name)
            self.assertNotEqual(ws['C2'].data_type, 'f')

    def test_rejects_workbook_with_declared_excel_maximum_dimension(self):
        with tempfile.TemporaryDirectory() as folder:
            root = Path(folder)
            normal, oversized, second = root / 'normal.xlsx', root / 'oversized.xlsx', root / 'second.xlsx'
            rows = [{'id': '230001', 'name': '甲', 'class': '顿河交231', 'score': 80}]
            self._comp_workbook(normal, rows)
            self._comp_workbook(second, rows)
            with zipfile.ZipFile(normal) as source, zipfile.ZipFile(oversized, 'w') as target:
                for item in source.infolist():
                    data = source.read(item.filename)
                    if item.filename == 'xl/worksheets/sheet1.xml':
                        data = re.sub(
                            br'<dimension ref="[^"]+"',
                            b'<dimension ref="A1:XFD1048576"',
                            data,
                            count=1,
                        )
                    target.writestr(item, data)

            with self.assertRaisesRegex(ValueError, '规模超限'):
                process_annual_comprehensive(str(oversized), str(second), folder, '2024-2025')

    def test_rejects_same_file_and_bad_academic_year(self):
        with tempfile.TemporaryDirectory() as folder:
            path = Path(folder) / 'source.xlsx'
            self._comp_workbook(path, [
                {'id': '230001', 'name': '甲', 'class': '顿河土231', 'score': 80},
            ])
            with self.assertRaisesRegex(ValueError, '不能选择同一个文件'):
                process_annual_comprehensive(str(path), str(path), folder, '2024-2025')
            copy = Path(folder) / 'copy.xlsx'
            copy.write_bytes(path.read_bytes())
            with self.assertRaisesRegex(ValueError, '学年格式'):
                process_annual_comprehensive(str(path), str(copy), folder, '2024')


if __name__ == '__main__':
    unittest.main()
