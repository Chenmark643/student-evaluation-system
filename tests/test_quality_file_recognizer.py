import tempfile
import unittest
from pathlib import Path

import openpyxl

from backend.quality_file_recognizer import recognize_quality_bonus_file


class QualityFileRecognizerTests(unittest.TestCase):
    def setUp(self):
        self.roster = {
            '253301101': {'name': '王昱睿', 'class': '顿河交251'},
            '233303301': {'name': '吴会鑫', 'class': '顿河信233'},
            '253301329': {'name': '曹成毅', 'class': '顿河交253'},
            '243301225': {'name': '章洲宁', 'class': '顿河交242'},
        }

    def _workbook(self):
        temp_dir = tempfile.TemporaryDirectory()
        self.addCleanup(temp_dir.cleanup)
        path = Path(temp_dir.name) / '混合名单.xlsx'
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = '错表头'
        sheet.append(['序号', '序号', '姓名', '学号', '素拓加分'])
        sheet.append([1, '顿河交253', '曹成毅', 253301329, 0.4])

        suffix = workbook.create_sheet('姓名带次数')
        suffix.append(['姓名', '素拓加分'])
        suffix.append(['王昱睿10/10', 0.5])

        combined = workbook.create_sheet('班级姓名连写')
        combined.append(['运动会志愿者', None, None])
        combined.append(['志愿者', None, '志愿时长/h'])
        combined.append(['顿河信233吴会鑫', None, 16])
        combined.append(['顿河信233吴会鑫', None, 8])

        typo = workbook.create_sheet('错别字')
        typo.append(['班级', '姓名'])
        typo.append(['顿河交242', '章州宁'])

        conflicts = workbook.create_sheet('班级学号都错误')
        conflicts.append(['班级', '姓名', '学号'])
        conflicts.append(['顿河土259', '章洲宁', '999999999'])

        unknown = workbook.create_sheet('低相似度')
        unknown.append(['班级', '姓名', '学号'])
        unknown.append(['未知班级', '测试人员甲乙丙', '111111111'])
        workbook.save(path)
        workbook.close()
        return path

    def test_recognizes_standard_suffix_and_combined_identity_formats(self):
        result = recognize_quality_bonus_file(str(self._workbook()), self.roster)
        self.assertTrue(result['success'])
        by_source = {row['source_id']: row for row in result['rows']}

        standard = by_source['错表头:2']
        self.assertEqual(standard['matched_sid'], '253301329')
        self.assertEqual(standard['confidence'], 'high')
        self.assertEqual(standard['file_score'], 0.4)

        suffix = by_source['姓名带次数:2']
        self.assertEqual(suffix['matched_sid'], '253301101')
        self.assertTrue(suffix['selected'])

        combined = by_source['班级姓名连写:3']
        self.assertEqual(combined['matched_sid'], '233303301')
        self.assertIsNone(combined['file_score'])

    def test_duplicates_are_unselected_and_typos_require_review(self):
        result = recognize_quality_bonus_file(str(self._workbook()), self.roster)
        duplicates = [row for row in result['rows'] if row['status'] == 'duplicate']
        self.assertEqual(len(duplicates), 1)
        self.assertFalse(duplicates[0]['selected'])

        typo = next(row for row in result['rows'] if row['source_id'] == '错别字:2')
        self.assertEqual(typo['status'], 'review')
        self.assertFalse(typo['selected'])
        self.assertEqual(typo['candidates'][0]['sid'], '243301225')

    def test_candidates_search_the_full_roster_when_class_and_id_are_wrong(self):
        result = recognize_quality_bonus_file(str(self._workbook()), self.roster)
        by_source = {row['source_id']: row for row in result['rows']}

        conflict = by_source['班级学号都错误:2']
        self.assertEqual(conflict['status'], 'review')
        self.assertEqual(conflict['candidates'][0]['sid'], '243301225')
        self.assertEqual(conflict['candidates'][0]['name_similarity'], 1.0)
        self.assertIn('similarity', conflict['candidates'][0])
        self.assertGreaterEqual(len(conflict['candidates']), 4)

        unknown = by_source['低相似度:2']
        self.assertEqual(unknown['status'], 'unmatched')
        self.assertTrue(unknown['candidates'])
        self.assertIn('相似度最高', unknown['reason'])


if __name__ == '__main__':
    unittest.main()
