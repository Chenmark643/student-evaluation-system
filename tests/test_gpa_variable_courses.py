import tempfile
import unittest
from pathlib import Path

import openpyxl

from backend.gpa_course_audit import analyze_gpa_course_structure
from backend.import_studio import analyze_import_file
from backend.module_a_gpa import process_gpa_batch


class VariableCourseGpaTests(unittest.TestCase):
    def _make_source(self, folder):
        path = Path(folder) / 'source.xlsx'
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = '测试班'
        ws.append(['学号','姓名','班级','课程门数','数学','等级课程',None,'转专业课程','体育（1）','学分绩点'])
        ws.append(['总学分',None,None,None,3,2,None,1,1,None])
        ws.append(['250001','普通学生','测试251',2,80,'良',85,None,90,None])
        ws.append(['250002','转专业学生','测试251',3,70,'优',95,88,None,None])
        wb.save(path); return path

    def test_detects_per_student_course_sets_and_companion_columns(self):
        with tempfile.TemporaryDirectory() as folder:
            path = self._make_source(folder)
            result = analyze_gpa_course_structure(str(path), {'sheet_name':'测试班','header_row':0,'id_col':0,'name_col':1,'class_col':2,'course_count_col':3,'course_start_col':4})
            self.assertTrue(result['success'])
            self.assertEqual(len(result['courses']), 4)
            grade_course = next(c for c in result['courses'] if c['name'] == '等级课程')
            self.assertEqual(grade_course['value_col'], 6)
            self.assertEqual(grade_course['credit'], 2)
            self.assertEqual([s['detected_count'] for s in result['students']], [3,3])
            self.assertEqual(result['classes'][0]['name'], '测试251')
            self.assertEqual(result['classes'][0]['student_count'], 2)
            self.assertEqual(result['classes'][0]['groups'][0]['students'][0]['total_credits'], 6)

    def test_first_students_scores_are_never_used_as_credits(self):
        with tempfile.TemporaryDirectory() as folder:
            path=Path(folder)/'raw-without-credit-row.xlsx'
            wb=openpyxl.Workbook(); ws=wb.active
            ws.append(['学号','姓名','班级','课程门数','课程A','课程B','学分绩点'])
            ws.append(['250001','甲','测试251',2,77,88,80])
            wb.save(path)
            result=analyze_gpa_course_structure(str(path),{'header_row':0,'id_col':0,'name_col':1,'class_col':2,'course_count_col':3,'course_start_col':4})
            self.assertTrue(result['success'])
            self.assertTrue(all(course['credit'] == 0 for course in result['courses']))
            self.assertTrue(all(course['credit_source'] == '待人工确认' for course in result['courses']))

    def test_export_uses_confirmed_credits_and_each_students_actual_courses(self):
        with tempfile.TemporaryDirectory() as folder:
            path = self._make_source(folder)
            base={'sheet_name':'测试班','header_row':0,'id_col':0,'name_col':1,'class_col':2,'course_count_col':3,'course_start_col':4,'enabled':True}
            audit=analyze_gpa_course_structure(str(path),base); base['course_definitions']=audit['courses']
            result=process_gpa_batch([str(path)],folder,column_mappings={str(path):{'测试班':base}})
            wb=openpyxl.load_workbook(result['output1'],data_only=False); ws=wb['测试251']
            self.assertEqual(ws['C3'].value,3)
            self.assertEqual(ws['D3'].value,5)  # math + grade; PE excluded
            self.assertEqual(ws['C4'].value,3)
            self.assertEqual(ws['D4'].value,6)  # math + grade + transfer course
            self.assertIn('/D3',ws.cell(3,ws.max_column).value)
            self.assertIn('/D4',ws.cell(4,ws.max_column).value)

    def test_ranking_uses_the_same_gpa_as_the_main_table_for_grade_only_courses(self):
        with tempfile.TemporaryDirectory() as folder:
            path = Path(folder) / 'grade-only.xlsx'
            wb = openpyxl.Workbook(); ws = wb.active; ws.title = '测试班'
            ws.append(['学号','姓名','班级','课程门数','数学','等级课程','体育'])
            ws.append(['总学分',None,None,None,3,2,1])
            ws.append(['250001','测试学生','测试251',3,80,'良',90])
            wb.save(path)

            mapping = {
                'sheet_name':'测试班','header_row':0,'id_col':0,'name_col':1,
                'class_col':2,'course_count_col':3,'enabled':True,
                'course_definitions':[
                    {'name':'数学','score_col':4,'value_col':None,'credit':3,'is_pe':False,'enabled':True},
                    {'name':'等级课程','score_col':5,'value_col':None,'credit':2,'is_pe':False,'enabled':True},
                    {'name':'体育','score_col':6,'value_col':None,'credit':1,'is_pe':True,'enabled':True},
                ],
            }
            result = process_gpa_batch(
                [str(path)], folder,
                column_mappings={str(path):{'测试班':mapping}},
            )

            main_wb = openpyxl.load_workbook(result['output1'], data_only=False)
            main_ws = main_wb['测试251']
            self.assertEqual(main_ws.cell(3, main_ws.max_column).value, '=(E3*E$2+G3*F$2)/D3')

            rank_wb = openpyxl.load_workbook(result['output2'], data_only=True)
            rank_ws = rank_wb[rank_wb.sheetnames[0]]
            self.assertAlmostEqual(rank_ws.cell(2, 3).value, (80 * 3 + 85 * 2) / 5, places=6)

    def test_real_transfer_student_sample_keeps_individual_courses(self):
        path=Path(r'D:\Wechat\xwechat_files\wxid_rnn8vd8giljk22_1bb2\msg\file\2026-04\2025-2026-1学分绩点(1).xlsx')
        if not path.exists(): self.skipTest('real GPA sample is unavailable')
        result=analyze_gpa_course_structure(str(path),{'sheet_name':'国电241','header_row':0,'id_col':0,'name_col':1,'class_col':2,'course_count_col':3,'course_start_col':4})
        self.assertEqual(result['typical_course_count'],16)
        dong=next(s for s in result['students'] if s['name']=='董安琪')
        self.assertEqual(dong['declared_count'],18)
        self.assertEqual(dong['detected_count'],18)
        self.assertTrue(any(c['enrolled_count']==1 for c in result['courses']))


    def test_sheet_name_is_used_when_the_class_column_is_missing(self):
        with tempfile.TemporaryDirectory() as folder:
            path = Path(folder) / 'missing-class.xlsx'
            wb = openpyxl.Workbook(); ws = wb.active; ws.title = '顿河信251'
            ws.append(['学号', '姓名', '课程门数', '总学分', '数学'])
            ws.append(['', '', '', '', 3])
            ws.append(['253303101', '测试学生', 1, 3, 80])
            wb.save(path)
            result = analyze_gpa_course_structure(str(path), {
                'sheet_name': '顿河信251', 'header_row': 0, 'id_col': 0,
                'name_col': 1, 'class_col': None, 'course_count_col': 2,
                'course_start_col': 4,
            })
            self.assertEqual(result['classes'][0]['name'], '顿河信251')

    def test_real_file_detects_all_25_information_classes(self):
        candidates = list(Path(r'D:\Wechat\xwechat_files\wxid_rnn8vd8giljk22_1bb2\msg\file\2026-04').glob('*.xlsx'))
        path = next((p for p in candidates if p.stat().st_size == 242951), None)
        if path is None: self.skipTest('real GPA sample is unavailable')
        analysis = analyze_import_file(str(path), 'gpa_raw')
        names = set(analysis['recommended_sheets'])
        self.assertTrue({f'顿河信25{i}' for i in range(1, 7)}.issubset(names))

        info_251 = next(sheet for sheet in analysis['sheets'] if sheet['name'] == '顿河信251')
        self.assertEqual(info_251['suggested_mapping']['course_start_col'], 4)
        audit = analyze_gpa_course_structure(str(path), {
            **info_251['suggested_mapping'], 'sheet_name': '顿河信251',
            'header_row': info_251['header_row'],
        })
        self.assertEqual(len(audit['courses']), 11)
        self.assertEqual(audit['students'][0]['detected_count'], 11)
        self.assertEqual([course['credit'] for course in audit['courses'][:3]], [1, 3, 2])

        transport = next(sheet for sheet in analysis['sheets'] if sheet['name'] == '顿河交251')
        transport_audit = analyze_gpa_course_structure(str(path), {
            **transport['suggested_mapping'], 'sheet_name': '顿河交251',
            'header_row': transport['header_row'],
        })
        self.assertEqual(len(transport_audit['courses']), 9)
        self.assertNotIn('总学分（除体育）', [course['name'] for course in transport_audit['courses']])
        self.assertEqual([course['credit'] for course in transport_audit['courses'][:3]], [1, 2.5, 1])

    def test_student_difference_lists_specific_courses(self):
        with tempfile.TemporaryDirectory() as folder:
            path = self._make_source(folder)
            result = analyze_gpa_course_structure(str(path), {
                'sheet_name':'测试班','header_row':0,'id_col':0,'name_col':1,
                'class_col':2,'course_count_col':3,'course_start_col':4,
            })
            transfer = next(s for s in result['students'] if s['name'] == '转专业学生')
            self.assertEqual(transfer['extra_courses'], ['转专业课程'])
            self.assertEqual(transfer['missing_courses'], ['体育（1）'])


if __name__ == '__main__': unittest.main()
