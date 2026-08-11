import json
import tempfile
import unittest
from copy import copy
from pathlib import Path
from unittest.mock import Mock, patch

from openpyxl import Workbook

from backend import kdocs_sync


class KdocsSyncTests(unittest.TestCase):
    def test_kdocs_links_are_https_and_host_allowlisted(self):
        self.assertEqual(
            kdocs_sync._validated_kdocs_link("https://365.kdocs.cn/l/example"),
            "https://365.kdocs.cn/l/example",
        )
        for value in ("http://kdocs.cn/l/example", "https://kdocs.cn.evil.test/l/example"):
            with self.assertRaises(kdocs_sync.KdocsSyncError):
                kdocs_sync._validated_kdocs_link(value)

    def test_untrusted_formula_like_text_is_written_as_literal(self):
        workbook = Workbook()
        cell = workbook.active["A1"]
        for value in ("=HYPERLINK(\"https://example.test\")", "+1+1", "-1+1", "@SUM(A1:A2)"):
            cell.value = value
            cell.data_type = "s"
            self.assertEqual(kdocs_sync._cell_formula_payload(cell), "'" + value)
        cell.value = "=SUM(A1:A2)"
        cell.data_type = "f"
        self.assertEqual(kdocs_sync._cell_formula_payload(cell), "=SUM(A1:A2)")

    def test_version_tuple_ignores_prefix_and_orders_semantically(self):
        self.assertEqual(kdocs_sync._version_tuple("kdocs-cli v2.6.0"), (2, 6, 0))
        self.assertGreater(kdocs_sync._version_tuple("2.10.0"), kdocs_sync._version_tuple("2.9.9"))

    def test_cli_version_status_detects_available_official_update(self):
        completed = Mock(
            returncode=0,
            stdout="Current version: 2.5.21\nLatest version: 2.6.0\n",
            stderr="",
        )

        with patch.object(kdocs_sync, "_CLI_UPDATE_CACHE", None), \
                patch.object(kdocs_sync, "_find_cli", return_value="C:/kdocs-cli.exe"), \
                patch.object(kdocs_sync, "_cli_version", return_value="2.5.21"), \
                patch.object(kdocs_sync.subprocess, "run", return_value=completed) as run:
            result = kdocs_sync.get_cli_version_status(force=True)

        self.assertTrue(result["success"])
        self.assertTrue(result["update_available"])
        self.assertEqual(result["latest_version"], "2.6.0")
        self.assertEqual(run.call_args.args[0], ["C:/kdocs-cli.exe", "upgrade", "--check"])

    def test_upgrade_cli_installs_after_confirmation_and_reselects_latest(self):
        before = {
            "success": True,
            "current_version": "2.5.21",
            "latest_version": "2.6.0",
            "update_available": True,
        }
        completed = Mock(returncode=0, stdout="upgrade completed", stderr="")

        with patch.object(kdocs_sync, "get_cli_version_status", return_value=before), \
                patch.object(kdocs_sync, "_find_cli", return_value="C:/kdocs-cli.exe"), \
                patch.object(kdocs_sync, "_cli_version", return_value="2.6.0"), \
                patch.object(kdocs_sync, "_invalidate_cli_cache"), \
                patch.object(kdocs_sync.subprocess, "run", return_value=completed) as run:
            result = kdocs_sync.upgrade_cli()

        self.assertTrue(result["success"])
        self.assertTrue(result["updated"])
        self.assertEqual(result["current_version"], "2.6.0")
        self.assertEqual(run.call_args.args[0], ["C:/kdocs-cli.exe", "upgrade", "-y"])

    def test_bundled_cli_is_copied_to_persistent_user_storage_before_upgrade(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            bundle = root / "bundle"
            local = root / "local"
            bundle.mkdir()
            source = bundle / "kdocs-cli.exe"
            source.write_bytes(b"connector")

            with patch.object(kdocs_sync.sys, "frozen", True, create=True), \
                    patch.object(kdocs_sync.sys, "_MEIPASS", str(bundle), create=True), \
                    patch.object(kdocs_sync.os, "name", "nt"), \
                    patch.dict(kdocs_sync.os.environ, {"LOCALAPPDATA": str(local)}), \
                    patch.object(kdocs_sync, "_invalidate_cli_cache"):
                selected = Path(kdocs_sync._updatable_cli_path(str(source)))

            self.assertEqual(selected, local / "kdocs-cli" / "kdocs-cli.exe")
            self.assertEqual(selected.read_bytes(), b"connector")

    def test_upgrade_cli_refuses_to_run_during_cloud_sync(self):
        kdocs_sync._SYNC_LOCK.acquire()
        try:
            result = kdocs_sync.upgrade_cli()
        finally:
            kdocs_sync._SYNC_LOCK.release()

        self.assertFalse(result["success"])
        self.assertTrue(result["busy"])

    def test_update_ops_never_exceeds_kdocs_hundred_item_limit(self):
        operations = [{"opType": "formula", "rowFrom": i} for i in range(205)]

        with patch.object(kdocs_sync, "_api", return_value={}) as api:
            kdocs_sync._update_ops("file-1", 7, operations)

        self.assertEqual(api.call_count, 3)
        self.assertEqual(
            [len(call.args[2]["rangeData"]) for call in api.call_args_list],
            [100, 100, 5],
        )

    def test_run_can_discard_human_auth_output(self):
        completed = Mock(returncode=0, stdout="authorization completed\nnon-json output", stderr="")

        with patch.object(kdocs_sync, "_find_cli", return_value="kdocs-cli"), \
                patch.object(kdocs_sync.subprocess, "run", return_value=completed):
            result = kdocs_sync._run(["auth", "login"], parse_json=False)

        self.assertEqual(result, {})

    def test_kdocs_cli_runs_without_flashing_a_console_window_on_windows(self):
        completed = Mock(returncode=0, stdout='{"authenticated":true}', stderr="")

        with patch.object(kdocs_sync, "_find_cli", return_value="kdocs-cli"), \
                patch.object(kdocs_sync.os, "name", "nt"), \
                patch.object(kdocs_sync.subprocess, "run", return_value=completed) as run:
            kdocs_sync._run(["auth", "status"])

        self.assertEqual(
            run.call_args.kwargs["creationflags"],
            getattr(kdocs_sync.subprocess, "CREATE_NO_WINDOW", 0x08000000),
        )

    def test_login_discards_cli_output_then_checks_saved_credential(self):
        with patch.object(kdocs_sync, "_run", return_value={}) as run, \
                patch.object(kdocs_sync, "auth_status", return_value={"authenticated": True}):
            result = kdocs_sync.login()

        self.assertEqual(result, {"success": True, "authenticated": True})
        run.assert_called_once_with(
            ["auth", "login", "--oauth-timeout", "300000"],
            timeout=330,
            parse_json=False,
        )

    def test_json_parser_accepts_cli_prefix_text(self):
        parsed = kdocs_sync._json_from_output('notice\n{"code":0,"data":{"id":"abc"}}')
        self.assertEqual(parsed["data"]["id"], "abc")

    def test_extract_file_id_accepts_double_wrapped_cli_response(self):
        response = {
            "code": 0,
            "message": "success",
            "data": {
                "code": 0,
                "msg": "ok",
                "data": {"id": "cloud-file-123", "name": "gpa.xlsx"},
            },
        }

        self.assertEqual(kdocs_sync._extract_file_id(response), "cloud-file-123")

    def test_sheet_infos_accepts_cli_detail_envelope(self):
        response = {
            "code": 0,
            "data": {
                "code": 0,
                "data": {
                    "result": "ok",
                    "detail": {
                        "sheetsInfo": [
                            {"sheetId": 7, "sheetName": "顿河信243", "rowTo": 24, "colTo": 23}
                        ]
                    },
                },
            },
        }

        with patch.object(kdocs_sync, "_api", return_value=response):
            infos = kdocs_sync._sheet_infos("cloud-file-123")

        self.assertEqual(infos[0]["sheetId"], 7)
        self.assertEqual(infos[0]["sheetName"], "顿河信243")

    def test_add_sheet_omits_deprecated_type_parameter(self):
        created = {"sheetId": 9, "sheetName": "顿河土241"}
        with patch.object(kdocs_sync, "_api", return_value={}) as api, \
                patch.object(kdocs_sync, "_sheet_infos", return_value=[created]):
            result = kdocs_sync._add_sheet("cloud-file-123", "顿河土241")

        self.assertEqual(result, created)
        api.assert_called_once_with(
            "sheet",
            "add-sheet",
            {
                "file_id": "cloud-file-123",
                "name": "顿河土241",
                "end": True,
                "count": 1,
            },
        )

    def test_college_sheet_order_groups_program_grade_and_class(self):
        infos = [
            {"sheetId": 1, "sheetIdx": 0, "sheetName": "顿河信251", "isVisible": True},
            {"sheetId": 2, "sheetIdx": 1, "sheetName": "_values", "isVisible": False},
            {"sheetId": 3, "sheetIdx": 2, "sheetName": "顿河土242", "isVisible": True},
            {"sheetId": 4, "sheetIdx": 3, "sheetName": "顿河交251", "isVisible": True},
            {"sheetId": 5, "sheetIdx": 4, "sheetName": "顿河土241", "isVisible": True},
            {"sheetId": 6, "sheetIdx": 5, "sheetName": "国电241", "isVisible": True},
            {"sheetId": 7, "sheetIdx": 6, "sheetName": "自定义231", "isVisible": True},
        ]

        ordered = kdocs_sync._ordered_sheet_infos(infos)

        self.assertEqual(
            [item["sheetName"] for item in ordered],
            ["顿河交251", "顿河土241", "顿河土242", "顿河信251", "国电241", "自定义231", "_values"],
        )

    def test_ranking_sheet_order_uses_program_then_grade(self):
        infos = [
            {"sheetId": 1, "sheetIdx": 0, "sheetName": "顿河信25"},
            {"sheetId": 2, "sheetIdx": 1, "sheetName": "顿河交24"},
            {"sheetId": 3, "sheetIdx": 2, "sheetName": "顿河交23"},
            {"sheetId": 4, "sheetIdx": 3, "sheetName": "顿河土24"},
        ]

        ordered = kdocs_sync._ordered_sheet_infos(infos)

        self.assertEqual(
            [item["sheetName"] for item in ordered],
            ["顿河交23", "顿河交24", "顿河土24", "顿河信25"],
        )

    def test_reorder_workbook_moves_then_reads_back_for_verification(self):
        initial = [
            {"sheetId": 3, "sheetIdx": 0, "sheetName": "顿河信251"},
            {"sheetId": 1, "sheetIdx": 1, "sheetName": "顿河交251"},
            {"sheetId": 2, "sheetIdx": 2, "sheetName": "顿河土251"},
        ]
        verified = [
            {"sheetId": 1, "sheetIdx": 0, "sheetName": "顿河交251"},
            {"sheetId": 2, "sheetIdx": 1, "sheetName": "顿河土251"},
            {"sheetId": 3, "sheetIdx": 2, "sheetName": "顿河信251"},
        ]

        with patch.object(kdocs_sync, "_sheet_infos", side_effect=[initial, verified]) as sheet_infos, \
                patch.object(kdocs_sync, "_api", return_value={}) as api:
            result = kdocs_sync._reorder_workbook("file-1", retry_delays=())

        self.assertTrue(result["success"])
        self.assertEqual(result["sheet_order"], ["顿河交251", "顿河土251", "顿河信251"])
        self.assertEqual(result["moved_sheets"], 1)
        self.assertEqual(sheet_infos.call_count, 2)
        api.assert_called_once_with(
            "sheet",
            "update-worksheet",
            {
                "file_id": "file-1",
                "worksheet_id": 3,
                "move_sheet_id": 2,
                "move_type": "sheet_move_type_after",
            },
        )

    def test_api_raises_nested_service_error_code(self):
        response = {
            "code": 0,
            "data": {"code": 400006, "msg": "authentication expired", "data": None},
        }

        with patch.object(kdocs_sync, "_run", return_value=response):
            with self.assertRaises(kdocs_sync.KdocsSyncError) as raised:
                kdocs_sync._api("drive", "get-file-info", {"file_id": "file-1"})

        self.assertEqual(raised.exception.code, 400006)

    def test_sync_reports_expired_auth_as_login_required(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "gpa.xlsx"
            Workbook().save(path)
            expired = kdocs_sync.KdocsSyncError("authentication expired", code=400006)

            with patch.object(kdocs_sync, "auth_status", return_value={"authenticated": True}), \
                    patch.object(kdocs_sync, "_binding", return_value={"file_id": "file-1"}), \
                    patch.object(kdocs_sync, "_sync_existing", side_effect=expired):
                result = kdocs_sync.sync_workbook(str(path), "college-gpa-main-v1")

        self.assertFalse(result["success"])
        self.assertTrue(result["needs_login"])

    def test_cell_format_maps_generated_excel_style(self):
        workbook = Workbook()
        cell = workbook.active["A1"]
        font = copy(cell.font)
        font.name, font.sz, font.bold = "微软雅黑", 11, True
        cell.font = font
        alignment = copy(cell.alignment)
        alignment.horizontal, alignment.vertical, alignment.wrap_text = "center", "center", True
        cell.alignment = alignment
        cell.number_format = "0.00"

        xf = kdocs_sync._cell_format(cell)

        self.assertEqual(xf["font"]["name"], "微软雅黑")
        self.assertEqual(xf["font"]["dyHeight"], 220)
        self.assertTrue(xf["font"]["bls"])
        self.assertEqual(xf["alcH"], 2)
        self.assertEqual(xf["alcV"], 1)
        self.assertTrue(xf["wrap"])
        self.assertEqual(xf["numfmt"], "0.00")

    def test_sync_first_publish_uses_exact_workbook_upload(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "gpa.xlsx"
            workbook = Workbook()
            workbook.active["A1"] = "student_id"
            workbook.save(path)

            with patch.object(kdocs_sync, "auth_status", return_value={"authenticated": True}), \
                    patch.object(kdocs_sync, "_binding", return_value=None), \
                    patch.object(kdocs_sync, "_publish_new", return_value={"success": True, "created": True}) as publish:
                result = kdocs_sync.sync_workbook(str(path), "college-gpa-main-v1")

            self.assertEqual(result, {"success": True, "created": True})
            publish.assert_called_once_with(path.resolve(), "college-gpa-main-v1", None)

    def test_force_create_publishes_new_workbook_even_when_binding_exists(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "gpa.xlsx"
            Workbook().save(path)

            with patch.object(kdocs_sync, "auth_status", return_value={"authenticated": True}), \
                    patch.object(kdocs_sync, "_binding") as binding, \
                    patch.object(kdocs_sync, "_publish_new", return_value={"success": True, "created": True}) as publish, \
                    patch.object(kdocs_sync, "_sync_existing") as update:
                result = kdocs_sync.sync_workbook(
                    str(path), "college-gpa-main-v1", force_create=True
                )

            self.assertTrue(result["created"])
            binding.assert_not_called()
            update.assert_not_called()
            publish.assert_called_once_with(path.resolve(), "college-gpa-main-v1", None)

    def test_sync_workbook_reports_real_stages_to_callback(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "gpa.xlsx"
            Workbook().save(path)
            events = []

            def publish(local_path, cloud_key, progress_callback=None):
                progress_callback(35, "正在上传 Excel", "首次创建学院云表")
                progress_callback(90, "正在校验云端文件", "确认文件和访问链接")
                return {"success": True, "created": True}

            with patch.object(kdocs_sync, "auth_status", return_value={"authenticated": True}), \
                    patch.object(kdocs_sync, "_binding", return_value=None), \
                    patch.object(kdocs_sync, "_publish_new", side_effect=publish):
                result = kdocs_sync.sync_workbook(
                    str(path), "college-gpa-main-v1",
                    progress_callback=lambda percent, stage, detail="", **extra: events.append((percent, stage, detail, extra)),
                )

        self.assertTrue(result["success"])
        self.assertEqual(events[-1][0], 100)
        self.assertEqual(events[-1][1], "同步完成")
        self.assertIn("正在上传 Excel", [event[1] for event in events])

    def test_sync_existing_resumes_an_empty_sheet_as_new(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "gpa.xlsx"
            workbook = Workbook()
            workbook.active.title = "顿河土241"
            workbook.active["A1"] = "学号"
            workbook.save(path)

            empty_cloud_sheet = {
                "sheetId": 9,
                "sheetName": "顿河土241",
                "isEmpty": True,
                "rowTo": 0,
                "colTo": 0,
            }
            binding = {"file_id": "file-1", "link_url": "https://kdocs.cn/l/test", "name": "学院总表.xlsx"}
            summary = {"name": "顿河土241", "changed_cells": 1, "created": True}

            with patch.object(kdocs_sync, "_sheet_infos", return_value=[empty_cloud_sheet]), \
                    patch.object(kdocs_sync, "_sync_sheet", return_value=summary) as sync_sheet, \
                    patch.object(kdocs_sync, "_load_config", return_value={"version": 1, "bindings": {"gpa": dict(binding)}}), \
                    patch.object(kdocs_sync, "_save_config"):
                result = kdocs_sync._sync_existing(path, "gpa", binding)

        self.assertTrue(result["success"])
        self.assertEqual(result["created_sheets"], ["顿河土241"])
        self.assertTrue(sync_sheet.call_args.args[3])

    def test_sync_existing_skips_hidden_helper_sheet(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "gpa.xlsx"
            workbook = Workbook()
            workbook.active.title = "顿河土241"
            workbook.active["A1"] = "学号"
            helper = workbook.create_sheet("_values")
            helper.sheet_state = "hidden"
            helper["A1"] = "local-only"
            workbook.save(path)

            info = {"sheetId": 9, "sheetName": "顿河土241", "isEmpty": False, "rowTo": 0, "colTo": 0}
            binding = {"file_id": "file-1", "link_url": "https://kdocs.cn/l/test", "name": "学院总表.xlsx"}
            summary = {"name": "顿河土241", "changed_cells": 0, "created": False}
            with patch.object(kdocs_sync, "_sheet_infos", return_value=[info]), \
                    patch.object(kdocs_sync, "_sync_sheet", return_value=summary) as sync_sheet, \
                    patch.object(kdocs_sync, "_load_config", return_value={"version": 1, "bindings": {"gpa": dict(binding)}}), \
                    patch.object(kdocs_sync, "_save_config"):
                kdocs_sync._sync_existing(path, "gpa", binding)

        self.assertEqual(sync_sheet.call_count, 1)
        self.assertEqual(sync_sheet.call_args.args[1].title, "顿河土241")

    def test_format_operations_merge_adjacent_equal_styles(self):
        workbook = Workbook()
        sheet = workbook.active
        for column in range(1, 11):
            cell = sheet.cell(row=1, column=column)
            cell.value = column
            cell.font = copy(cell.font)
            cell.font = cell.font.copy(bold=True)

        operations = kdocs_sync._format_ops_for_sheet(sheet, 1, 10)

        self.assertEqual(len(operations), 1)
        self.assertEqual((operations[0]["colFrom"], operations[0]["colTo"]), (0, 9))

    def test_sync_overview_excludes_hidden_helper_names(self):
        binding = {"file_id": "file-1", "name": "学院总表.xlsx", "link_url": "https://kdocs.cn/l/test"}
        infos = [
            {"sheetName": "顿河信241"},
            {"sheetName": "顿河土241"},
            {"sheetName": "_values"},
        ]
        with patch.object(kdocs_sync, "_binding", return_value=binding), \
                patch.object(kdocs_sync, "_sheet_infos", return_value=infos):
            result = kdocs_sync.get_sync_overview("gpa", "顿河土")

        self.assertEqual(result["sheet_count"], 2)
        self.assertEqual(result["major_sheets"], ["顿河土241"])

    def test_binding_config_does_not_store_credentials(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            app_dir = Path(temp_dir) / "state"
            config_path = app_dir / "kdocs-workbooks.json"
            with patch.object(kdocs_sync, "APP_DIR", app_dir), patch.object(kdocs_sync, "CONFIG_PATH", config_path):
                kdocs_sync._save_config({
                    "version": 1,
                    "bindings": {"gpa": {"file_id": "file-1", "link_url": "https://kdocs.cn/l/test"}},
                })
                stored = json.loads(config_path.read_text(encoding="utf-8"))

            self.assertEqual(stored["bindings"]["gpa"]["file_id"], "file-1")
            self.assertNotIn("token", config_path.read_text(encoding="utf-8").lower())


if __name__ == "__main__":
    unittest.main()
